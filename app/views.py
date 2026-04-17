from django.shortcuts import get_object_or_404, render, redirect
from app.forms import *
from app.models import *
from useraccount.models import Role
from useraccount.models import CustomUser
from decimal import Decimal
from django.db.models import Sum, Count, Q, Avg
from datetime import datetime, timedelta, date
from django.utils import timezone
# import messages from python
from django.contrib import messages
from django.db.models.functions import Coalesce


# ─────────────────────────────────────────────────────────────────────────────
#  HELPER: which offices is this user allowed to work with?
# ─────────────────────────────────────────────────────────────────────────────

def get_user_allowed_offices(user):
    """
    Return a queryset of Office objects the user is permitted to operate.

    Rules:
    • Superuser:
        – Returns all offices that appear in UserOfficeAssignment for this user.
        – If the user has NO UserOfficeAssignment rows yet (e.g. admin just
          created), falls back to ALL offices so the admin is never locked out.
    • Regular staff:
        – Returns ONLY the offices in their UserOfficeAssignment rows.
        – Falls back to their office_allocation (if set) as a single-item list
          when no UserOfficeAssignment rows exist (graceful degradation).

    The result is always an Office queryset ordered by name.
    """
    assigned_qs = Office.objects.filter(
        user_assignments__user=user
    ).distinct().order_by('name')

    if user.is_superuser:
        if assigned_qs.exists():
            return assigned_qs
        # Superuser with no assignments yet → can see everything
        return Office.objects.all().order_by('name')

    # Regular staff
    if assigned_qs.exists():
        return assigned_qs

    # Fallback: only their primary office_allocation
    primary = getattr(user, 'office_allocation', None)
    if primary:
        return Office.objects.filter(pk=primary.pk)

    return Office.objects.none()


# ─────────────────────────────────────────────────────────────────────────────
#  HELPER: resolve the currently-selected office from the session
# ─────────────────────────────────────────────────────────────────────────────

def get_selected_office(request):
    """
    Return the Office the current user is scoped to, or None.

    Logic:
    1.  Superuser  → reads session['selected_office_id'].
        – Validates the stored id is inside get_user_allowed_offices().
        – Returns None ("All Branches") if no session value or id not allowed.
    2.  Regular user:
        – Tries session['selected_office_id'] first (so multi-office staff can
          switch between their allowed offices).
        – Validates against get_user_allowed_offices().
        – Falls back to their primary office_allocation.

    Never returns an office the user is not allowed to access.
    """
    allowed_qs = get_user_allowed_offices(request.user)

    if request.user.is_superuser:
        office_id = request.session.get('selected_office_id')
        if office_id:
            try:
                office = Office.objects.get(id=office_id)
                if allowed_qs.filter(pk=office.pk).exists():
                    return office
                else:
                    request.session.pop('selected_office_id', None)
                    request.session.pop('selected_office_name', None)
            except Office.DoesNotExist:
                request.session.pop('selected_office_id', None)
                request.session.pop('selected_office_name', None)

        # No session yet (first login) — default to office_allocation or first office
        primary = getattr(request.user, 'office_allocation', None)
        if primary and allowed_qs.filter(pk=primary.pk).exists():
            return primary
        return allowed_qs.first()  # never shows "Select Office" again

    # ── Regular (non-superuser) staff ────────────────────────────────────────
    office_id = request.session.get('selected_office_id')
    if office_id:
        try:
            office = Office.objects.get(id=office_id)
            if allowed_qs.filter(pk=office.pk).exists():
                return office
            else:
                # Stored office not in allowed set → clear & fall back
                request.session.pop('selected_office_id', None)
                request.session.pop('selected_office_name', None)
        except Office.DoesNotExist:
            request.session.pop('selected_office_id', None)
            request.session.pop('selected_office_name', None)

    # No valid session → use primary office_allocation
    primary = getattr(request.user, 'office_allocation', None)
    if primary and allowed_qs.filter(pk=primary.pk).exists():
        return primary

    # Last resort: first allowed office
    return allowed_qs.first()


# ─────────────────────────────────────────────────────────────────────────────
#  HELPER: base context every template-rendering view must include
# ─────────────────────────────────────────────────────────────────────────────

def get_base_context(request):
    """
    Return a dict with the context variables needed by base.html.

    Keys returned:
        offices              – all Office objects (legacy / fallback use)
        selected_office      – Office | None  (current active office)
        user_allowed_offices – QuerySet of offices the user may select
        is_hq_selected       – bool: True when the selected office name is "HQ"

    Usage in any view:
        context = {
            **get_base_context(request),
            'loans': loans_qs,
            ...
        }
        return render(request, 'app/my_template.html', context)
    """
    selected_office      = get_selected_office(request)
    user_allowed_offices = get_user_allowed_offices(request.user)
    offices              = Office.objects.all()

    # "HQ" detection — case-insensitive, strip whitespace
    is_hq_selected = bool(
        selected_office and selected_office.name.strip().upper() == 'HQ'
    )

    # filter_office is what every view must use for queryset filtering.
    # Rules:
    #   - HQ selected (any staff)  → filter_office = None  → all branches data
    #   - Superuser with no session  → filter_office = None  → all branches data
    #   - Specific branch selected → filter_office = that Office object
    filter_office = None if (is_hq_selected or selected_office is None) else selected_office

    return {
        'offices':              offices,
        'selected_office':      selected_office,
        'user_allowed_offices': user_allowed_offices,
        'is_hq_selected':       is_hq_selected,
        'filter_office':        filter_office,   # use THIS for queryset filtering
    }


# ─────────────────────────────────────────────────────────────────────────────
#  UPDATED: switch_branch — validates office against user's allowed set
# ─────────────────────────────────────────────────────────────────────────────

def switch_branch(request):
    """
    Save selected office to session and redirect back.

    Security: only offices in get_user_allowed_offices() are accepted.
    If an invalid office_id is passed it is silently ignored.

    HQ office: any staff assigned to HQ via UserOfficeAssignment can select it.
    When HQ is selected, get_base_context() sets filter_office=None so views
    automatically show all-branches data — no special case needed here.
    "All Branches" has been removed; HQ is the canonical way to see all data.
    """
    office_id = request.GET.get('office_id')
    allowed_qs = get_user_allowed_offices(request.user)

    if office_id:
        try:
            office = allowed_qs.get(id=office_id)   # raises if not allowed
            request.session['selected_office_id']   = office.id
            request.session['selected_office_name'] = office.name
        except Office.DoesNotExist:
            # Silently ignore — user tried to switch to an office they don't own
            pass

    # No else branch — "All Branches" removed. Select HQ office to see all data.

    next_url = request.GET.get('next', request.META.get('HTTP_REFERER', '/'))
    return redirect('index')


# ─────────────────────────────────────────────────────────────────────────────
#  UPDATED: base view (now uses get_base_context)
# ─────────────────────────────────────────────────────────────────────────────

def base(request):
    context = get_base_context(request)
    return render(request, 'app/base.html', context)
# ─────────────────────────────────────────────────────────────────────────────
#  UPDATED: index view
# ─────────────────────────────────────────────────────────────────────────────

def index(request):
    import datetime
    from decimal import Decimal

    today = date.today()

    # ── Date filter ───────────────────────────────────────────────────────────
    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            selected_date = None
    else:
        selected_date = None

    display_date = selected_date or today

    # ── Base context & office filter ──────────────────────────────────────────
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    # ── Determine user's primary office allocation safely ─────────────────────
    try:
        user_primary_office = request.user.office_allocation
    except Exception:
        user_primary_office = None

    # HQ rule: superuser OR null allocation OR explicitly HQ → see all branches
    is_hq_allocated = bool(
        request.user.is_superuser or
        not user_primary_office or
        (
            user_primary_office and
            hasattr(user_primary_office, 'name') and
            user_primary_office.name.strip().upper() == 'HQ'
        )
    )

    # ── Global stats (scoped to filter_office) ────────────────────────────────
    loans_qs      = LoanApplication.objects.all()
    repayments_qs = LoanRepayment.objects.all()
    clients_qs    = Client.objects.all()

    if filter_office:
        loans_qs      = loans_qs.filter(office=filter_office.name)
        repayments_qs = repayments_qs.filter(loan_application__office=filter_office.name)
        clients_qs    = clients_qs.filter(registered_office=filter_office)

    total_repayments = repayments_qs.aggregate(
        total=Sum('repayment_amount'))['total'] or Decimal('0.00')

    total_clients = clients_qs.count()

    total_due_loan = loans_qs.aggregate(
        total=Sum('repayment_amount_remaining'))['total'] or Decimal('0.00')

    repayment_rate = (
        (total_repayments / (total_repayments + total_due_loan)) * Decimal('100')
    ) if (total_repayments + total_due_loan) > 0 else Decimal('0.00')
    repayment_rate = repayment_rate.quantize(Decimal('0.01'))

    # ── Target offices for per-office summary cards ───────────────────────────
    if is_hq_allocated:
        # HQ / superuser / null allocation → always show all branches
        target_offices = Office.objects.exclude(name__iexact='HQ').order_by('name')
    elif filter_office:
        # Branch staff with a session office selected
        target_offices = Office.objects.filter(pk=filter_office.pk).exclude(name__iexact='HQ')
    elif user_primary_office:
        # Fallback: no session office yet, use their allocation
        target_offices = Office.objects.filter(pk=user_primary_office.pk).exclude(name__iexact='HQ')
    else:
        # Last resort: show all (better than showing nothing)
        target_offices = Office.objects.exclude(name__iexact='HQ').order_by('name')

    # ── Per-office summary data ───────────────────────────────────────────────
    loans_by_office        = []
    expenses_by_office     = []
    repayments_by_office   = []
    transactions_by_office = []

    for office in target_offices:
        # Loans
        r = LoanApplication.objects.filter(
            office=office.name,
            created_at__date=display_date,
        ).aggregate(total_amount=Sum('loan_amount'), count=Count('id'))
        loans_by_office.append({
            'office': office.name,
            'amount': r['total_amount'] or Decimal('0.00'),
            'count':  r['count'] or 0,
        })

        # Expenses
        r = Expense.objects.filter(
            office=office.name,
            expense_date=display_date,
        ).aggregate(total=Sum('amount'), count=Count('id'))
        expenses_by_office.append({
            'office': office.name,
            'amount': r['total'] or Decimal('0.00'),
            'count':  r['count'] or 0,
        })

        # Repayments
        r = LoanRepayment.objects.filter(
            loan_application__office__iexact=office.name,
            created_at__date=display_date,
        ).aggregate(total_amount=Sum('repayment_amount'), count=Count('id'))
        repayments_by_office.append({
            'office': office.name,
            'amount': r['total_amount'] or Decimal('0.00'),
            'count':  r['count'] or 0,
        })

        # Transactions (both incoming and outgoing)
        to_ = OfficeTransaction.objects.filter(
            office_to=office,
            transaction_date=display_date,
        ).aggregate(total_amount=Sum('amount'), count=Count('id'))
        fr_ = OfficeTransaction.objects.filter(
            office_from=office,
            transaction_date=display_date,
        ).aggregate(total_amount=Sum('amount'), count=Count('id'))
        transactions_by_office.append({
            'office': office.name,
            'amount': (to_['total_amount'] or Decimal('0.00')) +
                      (fr_['total_amount'] or Decimal('0.00')),
            'count':  (to_['count'] or 0) + (fr_['count'] or 0),
        })

    # ── Pre-computed totals for template ──────────────────────────────────────
    loans_total        = sum(item['amount'] for item in loans_by_office)
    expenses_total     = sum(item['amount'] for item in expenses_by_office)
    repayments_total   = sum(item['amount'] for item in repayments_by_office)
    transactions_total = sum(item['amount'] for item in transactions_by_office)

    return render(request, 'app/index.html', {
        **base_ctx,
        'total_repayments':        total_repayments,
        'total_clients':           total_clients,
        'total_due_loan':          total_due_loan,
        'repayment_rate':          repayment_rate,
        'loans_by_office':         loans_by_office,
        'expenses_by_office':      expenses_by_office,
        'repayments_by_office':    repayments_by_office,
        'transactions_by_office':  transactions_by_office,
        'loans_total':             loans_total,
        'expenses_total':          expenses_total,
        'repayments_total':        repayments_total,
        'transactions_total':      transactions_total,
        'today':                   today,
        'selected_date':           selected_date,
        'display_date':            display_date,
    })
    
    
def office_transactions(request):
    offices = Office.objects.all()
    context = {
        'offices': offices,
    }
    return render(request, 'app/office_transactions.html', context)

def staff_list(request):
    staff_members = CustomUser.objects.all().order_by('id')
    roles = Role.objects.all()
    context = {
        **get_base_context(request),
        'staff_members': staff_members,
        'roles':roles,
    }
    return render(request, 'app/staff_list.html', context)

def office(request):
    offices_list = Office.objects.all().order_by('-id')
    context = {
        **get_base_context(request),
        'offices_list': offices_list,
    }
    return render(request, 'app/office.html', context)

def office_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        region = request.POST.get('region')
        district = request.POST.get('district')
        ward = request.POST.get('ward')
        street = request.POST.get('street')
        founded_date = request.POST.get('founded_date')
        head_officer = None

        office = Office(
            name=name,
            region=region,
            district=district,
            ward=ward,
            street=street,
            founded_date=founded_date if founded_date else None,
            head_officer=head_officer
        )
        office.save()
        return redirect('office')
    return render(request, 'app/office.html')

def office_update(request, office_id):
    office = get_object_or_404(Office, id=office_id)

    if request.method == 'POST':
        office.name = request.POST.get('name')
        office.region = request.POST.get('region')
        office.district = request.POST.get('district')
        office.ward = request.POST.get('ward')
        office.street = request.POST.get('street')

        founded_date = request.POST.get('founded_date')
        office.founded_date = founded_date if founded_date else None

        office.save()
        return redirect('office')

    context = {
        'office': office
    }
    return render(request, 'app/office.html', context)

def customer_plan_calculator(request):
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/customer_plan_calculator.html', context)

def loan_calculator(request):
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/loan_calculator.html', context)


def clients(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    clients_qs = Client.objects.all().order_by('-id')
    if filter_office:
        clients_qs = clients_qs.filter(registered_office=filter_office)

    # IDs of clients who have at least one loan application
    clients_with_loans = set(
        LoanApplication.objects.values_list('client_id', flat=True).distinct()
    )

    return render(request, 'app/clients.html', {
        **base_ctx,
        'clients': clients_qs,
        'clients_with_loans': list(clients_with_loans),
    })

from django.http import JsonResponse
from django.views.decorators.http import require_POST   
@require_POST
def client_delete(request, client_id):
    try:
        client = Client.objects.get(id=client_id)
    except Client.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Client not found.'}, status=404)

    # Block delete if client has any loan application ever
    if client.loan_applications.exists():
        return JsonResponse({'success': False, 'error': 'Cannot delete a client who has loan history.'}, status=400)

    client.delete()
    return JsonResponse({'success': True})


def client_add(request):
    form = ClientForm()
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            checkno = form.cleaned_data.get('checkno', '').strip()

            if checkno:
                existing = Client.objects.filter(checkno=checkno).first()
                if existing:
                    registered_office = getattr(existing, 'registered_office', None)
                    office_name = registered_office.name if registered_office else 'Unknown Office'
                    form.add_error(
                        'checkno',
                        f"Client with check number '{checkno}' already exists. "
                        f"Registered at: {office_name}."
                    )
                    # ✅ Fixed: must render client_add.html, NOT clients.html
                    return render(request, 'app/client_add.html', {
                        **get_base_context(request),
                        'form': form,
                    })

            client = form.save(commit=False)
            client.registered_by     = request.user
            client.registered_office = get_selected_office(request)
            client.save()
            return redirect('process_loan_partb', pk=client.pk)

    return render(request, 'app/client_add.html', {
        **get_base_context(request),
        'form': form,
    })


def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)

    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            checkno = form.cleaned_data.get('checkno', '').strip()

            if checkno:
                existing = Client.objects.filter(checkno=checkno).exclude(pk=pk).first()
                if existing:
                    registered_office = getattr(existing, 'registered_office', None)
                    office_name = registered_office.name if registered_office else 'Unknown Office'
                    form.add_error(
                        'checkno',
                        f"Check number '{checkno}' is already used by "
                        f"{existing.firstname} {existing.lastname} "
                        f"at {office_name}."
                    )
                    return render(request, 'app/client_edit.html', {
                        **get_base_context(request),
                        'client': client,
                        'form': form,
                    })

            form.save()
            return redirect('clients')

    return render(request, 'app/client_edit.html', {
        **get_base_context(request),
        'client': client,
    })


def check_checkno(request):
    """
    AJAX endpoint for live checkno validation.
    Pass ?exclude_id=<pk> on the edit page to skip the current client.

    urls.py:
        path('check-checkno/', views.check_checkno, name='check_checkno'),
        path('clients/<int:pk>/edit/', views.client_edit, name='client_edit'),
    """
    checkno    = request.GET.get('checkno', '').strip()
    exclude_id = request.GET.get('exclude_id', None)

    if not checkno:
        return JsonResponse({'exists': False})

    qs = Client.objects.filter(checkno=checkno)
    if exclude_id:
        qs = qs.exclude(pk=exclude_id)

    existing = qs.first()
    if existing:
        registered_office = getattr(existing, 'registered_office', None)
        office_name = registered_office.name if registered_office else 'Unknown Office'
        return JsonResponse({
            'exists': True,
            'client': f"{existing.firstname} {existing.lastname}",
            'office': office_name,
        })

    return JsonResponse({'exists': False})

def office_transaction(request):
    transactions = OfficeTransaction.objects.all().order_by('-id')
    offices = Office.objects.all()
    context = {
        'transactions': transactions,
        'offices': offices,
    }
    return render(request, 'app/office_transaction.html', context)

def office_transaction_add(request):
    if request.method == 'POST':
        office_to_id = request.POST.get('office_to')
        transaction_type = request.POST.get('transaction_type')
        amount = request.POST.get('amount')

        # Get office assigned to logged-in user
        office_from = get_selected_office(request)

        if not office_from:
            return render(request, 'app/office_transaction.html', {
                'error': 'You are not assigned to any office.'
            })

        office_to = get_object_or_404(Office, id=office_to_id)

        OfficeTransaction.objects.create(
            office_from=office_from,
            office_to=office_to,
            transaction_type=transaction_type,
            amount=amount,
            processed_by=request.user
        )

        return redirect('office_transaction')

    return render(request, 'app/office_transaction.html')

def bank_cash_transaction(request):
    transactions = BankCashTransaction.objects.all().order_by('-id')
    offices = Office.objects.all()
    context = {
        'transactions': transactions,
        'offices': offices,
    }
    return render(request, 'app/bank_cash_transaction.html', context)



def loans(request):
    base_ctx        = get_base_context(request)
    selected_office = base_ctx['selected_office']

    loans_qs = LoanApplication.objects.all()
    if selected_office:
        loans_qs = loans_qs.filter(processed_by__office_allocation=selected_office)

    clients_qs = Client.objects.all().order_by('-id')

    context = {
        **base_ctx,
        'loans':   loans_qs,
        'clients': clients_qs,
    }
    return render(request, 'app/loans.html', context)


def fomu_mkopo(request, loan_id=None):
    """
    View for loan application form - can be used for both new and existing loans
    If loan_id is provided, it loads existing loan data
    """
    loan = None
    client = None
    base_ctx        = get_base_context(request)
    
    # If loan_id is provided, get that specific loan
    if loan_id:
        loan = get_object_or_404(LoanApplication, id=loan_id)
        client = loan.client
    
    context = {
        **base_ctx,
        'loan': loan,
        'client': client,
        'current_user': request.user,
        'today': timezone.now().date(),
    }
    return render(request, 'app/fomu_mkopo.html', context)


def fomu_mkopo_new(request):
    """
    View for creating a new loan application
    You can pass a client_id to pre-fill client data
    """
    client_id = request.GET.get('client_id')
    client = None
    
    if client_id:
        client = get_object_or_404(Client, id=client_id)
    
    context = {
        'client': client,
        'current_user': request.user,
        'today': timezone.now().date(),
        'is_new': True,
    }
    return render(request, 'app/fomu_mkopo.html', context)

# def loan_application(request):
#     if request.method == 'POST':
#         client_id = request.POST.get('client')
#         loan_amount = request.POST.get('amount')
#         loan_purpose = request.POST.get('purpose')
#         loan_type = request.POST.get('loan_type')
#         interest_rate = request.POST.get('interest_rate')
#         payment_period_months = request.POST.get('term_months')
#         application_date = request.POST.get('application_date')
#         # fund_source = request.POST.get('fund_source')
#         transaction_method = request.POST.get('transaction_method')
#         processed_by = request.user

#         try:
#             loan_amount_decimal = Decimal(loan_amount)
#             client = get_object_or_404(Client, id=client_id)

#             # Check if client has existing pending loan
#             existing_loan = LoanApplication.objects.filter(
#                 client=client,
#                 status='Pending'
#             ).first()
#             if existing_loan:
#                 messages.error(request, f"Client {client} has an existing pending loan application. Cannot apply for a new loan until the existing one is processed.")
#                 return redirect('loans')

#             # Get branch office from fund_source or fallback to user's office
#             branch_office = processed_by.office_allocation
#             if not branch_office:
#                 messages.error(request, 'No branch office found. Please select a fund source.')
#                 return redirect('loans')

#             # Get latest branch balance
#             branch_balance = BranchBalance.objects.filter(branch=branch_office).last()
#             if not branch_balance:
#                 messages.error(request, f'No balance record found for {branch_office.name}.')
#                 return redirect('loans')

#             # Check sufficient balance & deduct based on transaction method
#             if transaction_method == 'cash':
#                 if branch_balance.office_balance < loan_amount_decimal:
#                     messages.error(request, f'Insufficient cash balance. Available: {branch_balance.office_balance}')
#                     return redirect('loans')
#                 BranchBalance.objects.create(
#                     branch=branch_office,
#                     office_balance=branch_balance.office_balance - loan_amount_decimal,
#                     bank_balance=branch_balance.bank_balance,
#                     updated_by=processed_by,
#                 )
#             else:
#                 if branch_balance.bank_balance < loan_amount_decimal:
#                     messages.error(request, f'Insufficient bank balance. Available: {branch_balance.bank_balance}')
#                     return redirect('loans')
#                 BranchBalance.objects.create(
#                     branch=branch_office,
#                     office_balance=branch_balance.office_balance,
#                     bank_balance=branch_balance.bank_balance - loan_amount_decimal,
#                     updated_by=processed_by,
#                 )

#             # Create loan application
#             new_loan = LoanApplication.objects.create(
#                 client=client,
#                 loan_amount=loan_amount_decimal,
#                 loan_purpose=loan_purpose,
#                 loan_type=loan_type,
#                 interest_rate=interest_rate,
#                 payment_period_months=payment_period_months,
#                 application_date=application_date,
#                 processed_by=processed_by,
#                 office=branch_office.name,
#             )

#             messages.success(request, f'Loan application for {client} processed successfully.')
#             return redirect('loan_receipt', loan_id=new_loan.id)

#         except Exception as e:
#             messages.error(request, f'Error processing loan application: {str(e)}')

#     return render(request, 'app/loan_application.html')


def loan_application_receipt(request, loan_id):
    from decimal import Decimal

    loan = get_object_or_404(LoanApplication, id=loan_id)
    repayments = LoanRepayment.objects.filter(
        loan_application=loan
    ).order_by('-id')

    total_paid = sum(
        (repayment.repayment_amount for repayment in repayments),
        Decimal('0.00')
    )

    installment_schedule = calculate_installment_schedule(loan)

    context = {
        'loan':                 loan,
        'repayments':           repayments,
        'total_paid':           total_paid,
        'installment_schedule': installment_schedule,
        **get_base_context(request),
    }
    return render(request, 'app/loan_application_receipt.html', context)

from decimal import Decimal, ROUND_HALF_UP, ROUND_DOWN

def calculate_installment_schedule(loan):
    """
    Builds the monthly repayment schedule using the flat-rate formula:
        Total Interest  = (I/100) * P
        Monthly Payment = (P + Total Interest) / N

    Rounding strategy (floor to nearest 1,000):
        - Middle/last months get the clean rounded-DOWN amount
        - First installment absorbs the remainder (always largest)

    First installment date comes from loan.first_repayment_date which is
    already set by the model using the rule:
        application_date.day <= 18  → same month, 28th
        application_date.day >  18  → next month, 28th
    """
    schedule = []

    P     = Decimal(str(loan.loan_amount))
    N     = loan.payment_period_months
    today = date.today()

    total_interest = loan.total_interest_amount
    total_return   = loan.total_repayment_amount

    def floor_1000(val):
        """Round DOWN to nearest 1,000"""
        return (val / Decimal('1000')).to_integral_value(rounding=ROUND_DOWN) * Decimal('1000')

    # Middle/last months — clean rounded-DOWN amounts
    rounded_principal = floor_1000(P             / Decimal(N))
    rounded_interest  = floor_1000(total_interest / Decimal(N))
    rounded_monthly   = rounded_principal + rounded_interest

    # First installment absorbs the remainder (always the largest)
    first_principal = P              - (rounded_principal * (N - 1))
    first_interest  = total_interest - (rounded_interest  * (N - 1))
    first_monthly   = first_principal + first_interest

    current_balance = total_return
    current_date    = loan.first_repayment_date

    for i in range(1, N + 1):
        is_first = (i == 1)

        row_principal = first_principal if is_first else rounded_principal
        row_interest  = first_interest  if is_first else rounded_interest
        row_payment   = first_monthly   if is_first else rounded_monthly

        current_balance = (current_balance - row_payment).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
        if abs(current_balance) < Decimal('0.50'):
            current_balance = Decimal('0.00')

        is_current = (
            current_date.year  == today.year and
            current_date.month == today.month
        )

        schedule.append({
            'phase':             i,
            'due_date':          current_date,
            'principal':         row_principal,
            'interest':          row_interest,
            'total_payment':     row_payment,
            'remaining_balance': current_balance,
            'is_current':        is_current,
        })

        current_date = current_date + relativedelta(months=1)

    return schedule

    
def loan_application_update(request, loan_id):
    loan = get_object_or_404(LoanApplication, id=loan_id)
    # check the present loan_amount
    present_amount = loan.loan_amount
    
    if request.method == 'POST':
        loan.loan_amount = request.POST.get('amount')
        loan.loan_purpose = request.POST.get('purpose')
        loan.loan_type = request.POST.get('loan_type')
        loan.interest_rate = request.POST.get('interest_rate')
        loan.payment_period_months = request.POST.get('term_months')
        loan.status = request.POST.get('status')
        loan.save()
        
        # check the difference between present amount and new amount
        new_amount = Decimal(loan.loan_amount)
        amount_diff = new_amount - present_amount
        # update bank amount accordingly
        bank_amount = BankAmount.objects.first()
        if bank_amount:
            bank_amount.amount -= amount_diff
            bank_amount.save()
        return redirect('loans')
    
    context = {
        'loan': loan,
    }
    return render(request, 'app/loan_application_update.html', context)

# def loan_repayment(request, loan_id):
#     offices = Office.objects.all()
#     loan = get_object_or_404(LoanApplication, id=loan_id)
    
#     repayments = LoanRepayment.objects.filter(
#         loan_application=loan
#     ).order_by('-id')
    
#     total_paid = sum(
#         (repayment.repayment_amount for repayment in repayments),
#         Decimal('0.00')
#     )
    
#     # Get the office for this loan
#     office = None
#     if loan.office:
#         try:
#             office = Office.objects.get(name=loan.office)
#         except Office.DoesNotExist:
#             pass

#     if request.method == 'POST':
#         repayment_amount = Decimal(request.POST.get('repayment_amount'))
#         transaction_method = request.POST.get('transaction_method')
#         processed_by = request.user

#         try:
#             # Prevent overpayment
#             if repayment_amount > loan.repayment_amount_remaining:
#                 repayment_amount = loan.repayment_amount_remaining

#             # Get branch office from processed_by's office allocation
#             branch_office = processed_by.office_allocation
#             if not branch_office:
#                 messages.error(request, 'You are not allocated to any office/branch.')
#                 return redirect('loan_repayment', loan_id=loan_id)

#             # Get latest branch balance
#             branch_balance = BranchBalance.objects.filter(branch=branch_office).last()

#             if not branch_balance:
#                 # Create initial balance record with repayment amount
#                 BranchBalance.objects.create(
#                     branch=branch_office,
#                     office_balance=repayment_amount if transaction_method == 'cash' else Decimal('0.00'),
#                     bank_balance=repayment_amount if transaction_method != 'cash' else Decimal('0.00'),
#                     updated_by=processed_by,
#                 )
#             else:
#                 # Add repayment to correct balance based on transaction method
#                 if transaction_method == 'cash':
#                     BranchBalance.objects.create(
#                         branch=branch_office,
#                         office_balance=branch_balance.office_balance + repayment_amount,
#                         bank_balance=branch_balance.bank_balance,
#                         updated_by=processed_by,
#                     )
#                 else:
#                     BranchBalance.objects.create(
#                         branch=branch_office,
#                         office_balance=branch_balance.office_balance,
#                         bank_balance=branch_balance.bank_balance + repayment_amount,
#                         updated_by=processed_by,
#                     )

#             # Create repayment record
#             LoanRepayment.objects.create(
#                 loan_application=loan,
#                 repayment_amount=repayment_amount,
#                 transaction_method=transaction_method,
#                 processed_by=processed_by,
#             )

#             # Update loan remaining balance and status
#             loan.repayment_amount_remaining -= repayment_amount
#             if loan.repayment_amount_remaining <= 0:
#                 loan.status = 'Paid'
#             loan.save()

#             messages.success(request, f'Repayment of {repayment_amount} processed successfully.')
#             return redirect('loans')

#         except Exception as e:
#             messages.error(request, f'Error processing repayment: {str(e)}')

#     context = {
#         'offices': offices,
#         'loan': loan,
#         'repayments': repayments,
#         'total_paid': total_paid,
#         'office': office,
#     }
    
#     return render(request, 'app/loan_repayment.html', context)




def check_topup_eligibility(request, loan_id):
    """AJAX endpoint to check if a loan is eligible for topup"""
    try:
        loan = get_object_or_404(LoanApplication, id=loan_id)
        
        # Check eligibility
        eligible = loan.status == 'Approved'
        message = ''
        
        if not eligible:
            message = f"Loan is {loan.status}. Only Approved loans are eligible for topup."
        
        return JsonResponse({
            'eligible': eligible,
            'message': message,
            'outstanding_balance': float(loan.repayment_amount_remaining) if eligible else 0,
            'loan_id': loan.id,
            'client_name': str(loan.client),
            'current_amount': float(loan.loan_amount)
        })
    except LoanApplication.DoesNotExist:
        return JsonResponse({'eligible': False, 'message': 'Loan not found'})
    except Exception as e:
        return JsonResponse({'eligible': False, 'message': str(e)})

def loan_topup(request, loan_id):
    """Handle loan topup submission"""
    if request.method == 'POST':
        original_loan = get_object_or_404(LoanApplication, id=loan_id)

        try:
            with transaction.atomic():
                topup_amount          = Decimal(request.POST.get('topup_amount', '0').replace(',', '').strip())
                interest_rate         = Decimal(request.POST.get('interest_rate', original_loan.interest_rate))
                transaction_method    = request.POST.get('transaction_method')
                payment_period_months = request.POST.get('term_months')
                first_repayment_date = request.POST.get('application_date')
                processed_by          = request.user
                

                topup_date_str = request.POST.get('topup_date')
                payment_month = (
                    datetime.datetime.strptime(topup_date_str, '%Y-%m-%d').date()
                    if topup_date_str
                    else date.today()
                )
                # topup_date = date.today()
                topup_date_input = request.POST.get('topup_date', '')
                topup_date = datetime.datetime.strptime(topup_date_input, '%Y-%m-%d').date()

                # ── Branch office ─────────────────────────────────────────
                branch_office = get_selected_office(request)
                if not branch_office:
                    messages.error(request, 'You are not allocated to any office/branch.')
                    return redirect('process_loan_partb', pk=original_loan.client.id)

                # ── Outstanding balance (true remaining after all prior payments) ─
                original_loan = LoanApplication.objects.get(pk=loan_id)   # fresh, not stale
                outstanding   = original_loan.repayment_amount_remaining

                # ── Validate topup amount > outstanding ───────────────────
                if topup_amount <= outstanding:
                    messages.error(
                        request,
                        f"Topup amount ({topup_amount:,.0f}) must be greater than "
                        f"outstanding balance ({outstanding:,.0f})"
                    )
                    return redirect('process_loan_partb', pk=original_loan.client.id)

                # ── Net disbursement = what client actually receives ───────
                net_disbursement = topup_amount - outstanding
                new_loan_amount  = topup_amount

                # ── Branch balance check ──────────────────────────────────
                branch_balance = BranchBalance.objects.filter(branch=branch_office).last()
                if not branch_balance:
                    messages.error(request, f'No balance record found for {branch_office.name}.')
                    return redirect('process_loan_partb', pk=original_loan.client.id)

                if transaction_method == 'cash':
                    if branch_balance.office_balance < net_disbursement:
                        messages.error(
                            request,
                            f'Insufficient cash balance. Available: {branch_balance.office_balance:,.0f}/='
                        )
                        return redirect('process_loan_partb', pk=original_loan.client.id)
                    BranchBalance.objects.create(
                        branch=branch_office,
                        office_balance=branch_balance.office_balance - net_disbursement,
                        bank_balance=branch_balance.bank_balance,
                        updated_by=processed_by,
                    )
                else:  # bank transfer
                    if branch_balance.bank_balance < net_disbursement:
                        messages.error(
                            request,
                            f'Insufficient bank balance. Available: {branch_balance.bank_balance:,.0f}/='
                        )
                        return redirect('process_loan_partb', pk=original_loan.client.id)
                    BranchBalance.objects.create(
                        branch=branch_office,
                        office_balance=branch_balance.office_balance,
                        bank_balance=branch_balance.bank_balance - net_disbursement,
                        updated_by=processed_by,
                    )

                # ── Create new loan application for the top-up ────────────
                new_loan = LoanApplication.objects.create(
                    client=original_loan.client,
                    loan_amount=new_loan_amount,
                    loan_purpose=f"Topup of loan #{original_loan.id} - {original_loan.loan_purpose}",
                    loan_type=original_loan.loan_type,
                    interest_rate=interest_rate,
                    payment_period_months=payment_period_months,
                    application_date=topup_date,
                    first_repayment_date = first_repayment_date,
                    status='Approved',
                    processed_by=processed_by,
                    office=branch_office.name,
                    transaction_method=transaction_method,
                )

                new_loan.loan_purpose = (
                    f"Topup of loan #{original_loan.id} - "
                    f"Original total: {topup_amount:,.0f}/=, "
                    f"Outstanding: {outstanding:,.0f}/=, "
                    f"Net: {net_disbursement:,.0f}/= - "
                    f"{original_loan.loan_purpose}"
                )
                new_loan.save(update_fields=['loan_purpose'])

                # ── Create LoanTopup record linked to original loan ────────
                LoanTopup.objects.create(
                    loan_application=original_loan,
                    topup_amount=topup_amount,
                    old_balance_cleared=outstanding,
                    interest_rate=interest_rate,
                    transaction_method=transaction_method,
                    processed_by=processed_by,
                    topup_date=topup_date,
                    payment_month = payment_month,
                )

                # ── Record final clearance repayment on original loan ─────
                # This ensures repayment history is complete — capturing exactly
                # what remained after all prior payments, not the full loan amount.
                if outstanding > Decimal('0.00'):
                    LoanRepayment.objects.create(
                        loan_application=original_loan,
                        repayment_amount=outstanding,
                        repayment_date=topup_date,
                        transaction_method=transaction_method,
                        processed_by=processed_by,
                    )

                # ── Close original loan and zero out remaining balance ─────
                original_loan.status = 'Closed'
                original_loan.repayment_amount_remaining = Decimal('0.00')
                original_loan.save()

                messages.success(
                    request,
                    f"✅ Topup completed successfully!\n"
                    f"• New Loan #{new_loan.id}: TZS {new_loan_amount:,.0f}/=\n"
                    f"• Client Receives: TZS {net_disbursement:,.0f}/=\n"
                    f"• Original Loan #{original_loan.id} closed."
                )

                return redirect('clients')
                # return redirect('loan_receipt', loan_id=new_loan.id)

        except Exception as e:
            messages.error(request, f"Error processing topup: {str(e)}")
            return redirect('process_loan_partb', pk=original_loan.client.id)

    return redirect('loans')



def view_receipt(request, repayment_id):
    repayment = get_object_or_404(LoanRepayment, id=repayment_id)
    loan = repayment.loan_application
    
    # Get office details
    office = None
    if loan.office:
        try:
            office = Office.objects.get(name=loan.office)
        except Office.DoesNotExist:
            pass
    
    # Prepare context for receipt template
    context = {
        'repayment': repayment,
        'loan': loan,
        'office': office,
        'company_name': 'YOUR MICROFINANCE COMPANY',  # Replace with your company name
        'receipt_date': repayment.repayment_date,
        'receipt_no': f'RCP-{repayment.id:06d}',
    }
    
    return render(request, 'app/receipt_view.html', context)

def salary(request):
    offices = Office.objects.all()
    salaries = Salary.objects.all().order_by('-id')
    
    # Calculate summary data
    total_salary = salaries.aggregate(total=Sum('amount'))['total'] or 0
    current_month = datetime.datetime.now().replace(day=1)
    monthly_salary = salaries.filter(
        salary_for_month__month=current_month.month,
        salary_for_month__year=current_month.year
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Get last payment
    last_payment = salaries.first()
    
    # Get unique months for filter
    months = Salary.objects.dates('salary_for_month', 'month', order='DESC')
    month_choices = [date.strftime('%B %Y') for date in months]
    
    # Get unique years for filter
    years = Salary.objects.dates('salary_for_month', 'year', order='DESC')
    year_choices = [date.strftime('%Y') for date in years]
    
    # Get all employees for the add form
    employees = CustomUser.objects.all()  # Adjust filter as needed
    # employees = CustomUser.objects.filter(is_staff=False, is_active=True)  # Adjust filter as needed
    
    context = {
        'salaries': salaries,
        'total_salary': total_salary,
        'monthly_salary': monthly_salary,
        'last_payment': last_payment,
        'months': month_choices,
        'years': year_choices,
        'employees': employees,
        'offices': offices,
    }
    return render(request, 'app/salary.html', context)

def salary_add(request):
    if request.method == 'POST':
        employee_id = request.POST.get('employee')
        amount = request.POST.get('amount')
        salary_for_month = request.POST.get('salary_for_month') + "-01"
        fund_source_id = request.POST.get('fund_source')
        transaction_method = request.POST.get('transaction_method')
        processed_by = request.user

        try:
            amount_decimal = Decimal(amount)
            employee = get_object_or_404(CustomUser, id=employee_id)
            branch_office = get_object_or_404(Office, id=fund_source_id)

            # Get latest branch balance
            branch_balance = BranchBalance.objects.filter(branch=branch_office).last()

            if not branch_balance:
                messages.error(request, f'No balance record found for {branch_office.name}.')
                return redirect('salary')

            # Deduct based on transaction method
            if transaction_method == 'cash':
                if branch_balance.office_balance < amount_decimal:
                    messages.error(request, f'Insufficient cash balance. Available: {branch_balance.office_balance}')
                    return redirect('salary')

                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance - amount_decimal,
                    bank_balance=branch_balance.bank_balance,
                    updated_by=processed_by,
                )
            else:
                if branch_balance.bank_balance < amount_decimal:
                    messages.error(request, f'Insufficient bank balance. Available: {branch_balance.bank_balance}')
                    return redirect('salary')

                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance,
                    bank_balance=branch_balance.bank_balance - amount_decimal,
                    updated_by=processed_by,
                )

            # ✅ Correct fund_source usage here
            Salary.objects.create(
                employee=employee,
                amount=amount_decimal,
                salary_for_month=salary_for_month,
                transaction_method=transaction_method,
                fund_source=branch_office,  # <-- FIXED
                processed_by=processed_by,
            )

            messages.success(
                request,
                f'Salary of {amount_decimal} processed for {employee.get_full_name()} successfully.'
            )

        except Exception as e:
            messages.error(request, f'Error processing salary: {str(e)}')

    return redirect('salary')

# def loan_report(request):
#     # Get date filters from request
#     default_start_date = timezone.now().date() - timedelta(days=28)
#     default_end_date = timezone.now().date()
    
#     start_date_str = request.GET.get('start_date', default_start_date.strftime('%Y-%m-%d'))
#     end_date_str = request.GET.get('end_date', default_end_date.strftime('%Y-%m-%d'))
    
#     # Convert to date objects
#     try:
#         start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
#         end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
#     except:
#         start_date = default_start_date
#         end_date = default_end_date
    
#     # Get all offices
#     offices = Office.objects.all()
    
#     # Prepare office-wise data
#     office_data = []
#     total_loans = 0
#     total_amount_loaned = Decimal('0.00')
#     total_amount_owed = Decimal('0.00')
#     total_amount_due = Decimal('0.00')
    
#     for office in offices:
#         # Get users in this office
#         office_users = CustomUser.objects.filter(office_allocation=office)
        
#         if office_users.exists():
#             # Get loans processed by users from this office within date range
#             loans = LoanApplication.objects.filter(
#                 processed_by__in=office_users,
#                 application_date__range=[start_date, end_date]
#             )
            
#             # Calculate totals
#             loan_count = loans.count()
#             amount_loaned = loans.aggregate(total=Sum('loan_amount'))['total'] or Decimal('0.00')
#             amount_owed = loans.aggregate(total=Sum('total_repayment_amount'))['total'] or Decimal('0.00')
#             amount_due = loans.aggregate(total=Sum('repayment_amount_remaining'))['total'] or Decimal('0.00')
            
#             # Calculate repayment performance
#             repaid_amount = amount_owed - amount_due
            
#             office_data.append({
#                 'office': office,
#                 'loan_count': loan_count,
#                 'amount_loaned': amount_loaned,
#                 'amount_owed': amount_owed,
#                 'amount_due': amount_due,
#                 'repaid_amount': repaid_amount,
#                 'repaid_percentage': (repaid_amount / amount_owed * 100) if amount_owed > 0 else 0
#             })
            
#             # Update grand totals
#             total_loans += loan_count
#             total_amount_loaned += amount_loaned
#             total_amount_owed += amount_owed
#             total_amount_due += amount_due
    
#     # Get loan details for the table
#     loan_details = LoanApplication.objects.filter(
#         application_date__range=[start_date, end_date]
#     ).select_related('client', 'processed_by', 'processed_by__office_allocation')
    
#     # Calculate overall repayment percentage
#     total_repaid = total_amount_owed - total_amount_due
#     overall_repayment_percentage = (total_repaid / total_amount_owed * 100) if total_amount_owed > 0 else 0
    
#     # Get status breakdown
#     status_breakdown = loan_details.values('status').annotate(
#         count=Count('id'),
#         total_amount=Sum('loan_amount')
#     )
    
#     # Get recent loans for the "Recent Loans" card
#     recent_loans = loan_details.order_by('-application_date')[:5]
    
#     context = {
#         'office_data': office_data,
#         'loan_details': loan_details,
#         'total_loans': total_loans,
#         'total_amount_loaned': total_amount_loaned,
#         'total_amount_owed': total_amount_owed,
#         'total_amount_due': total_amount_due,
#         'total_repaid': total_repaid,
#         'overall_repayment_percentage': overall_repayment_percentage,
#         'status_breakdown': status_breakdown,
#         'recent_loans': recent_loans,
#         'start_date': start_date,
#         'end_date': end_date,
#         'start_date_str': start_date_str,
#         'end_date_str': end_date_str,
#         'default_start_date': default_start_date.strftime('%Y-%m-%d'),
#         'default_end_date': default_end_date.strftime('%Y-%m-%d'),
#         'offices': offices,
#     }
    
#     return render(request, 'app/loan_report.html', context)

def expense(request):
    base_ctx        = get_base_context(request)
    selected_office = base_ctx['selected_office']

    expenses_qs = Expense.objects.all().order_by('-id')
    if selected_office:
        expenses_qs = expenses_qs.filter(office=selected_office.name)

    expense_categories = ExpenseCategory.objects.all().order_by('-id')

    context = {
        **base_ctx,
        'expenses':          expenses_qs,
        'expense_categories': expense_categories,
    }
    return render(request, 'app/expense.html', context)

def expense_add(request):
    if request.method == 'POST':
        try:
            description = request.POST.get('description')
            amount = request.POST.get('amount')
            transaction_type_id = request.POST.get('transaction_type')
            expense_account = request.POST.get('expense_account')
            payment_method = request.POST.get('payment_method')
            transaction_date = request.POST.get('transaction_date')
            attachment = request.FILES.get('attachment')

            # Convert amount safely
            amount = Decimal(amount) if amount else Decimal('0.00')

            # Convert transaction date
            if transaction_date:
                transaction_date = datetime.datetime.strptime(transaction_date, "%Y-%m-%d").date()
            else:
                transaction_date = None

            # Get category object
            transaction_type = get_object_or_404(ExpenseCategory, id=transaction_type_id)

            # Get branch office from user's office allocation
            branch_office = get_selected_office(request)
            if not branch_office:
                messages.error(request, 'You are not allocated to any office/branch.')
                return redirect('expense')

            # Get latest branch balance
            branch_balance = BranchBalance.objects.filter(branch=branch_office).last()
            if not branch_balance:
                messages.error(request, f'No balance record found for {branch_office.name}.')
                return redirect('expense')

            # Check sufficient balance & deduct based on payment method
            if payment_method == 'cash':
                if branch_balance.office_balance < amount:
                    messages.error(request, f'Insufficient cash balance. Available: {branch_balance.office_balance}')
                    return redirect('expense')
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance - amount,
                    bank_balance=branch_balance.bank_balance,
                    updated_by=request.user,
                )
            else:
                if branch_balance.bank_balance < amount:
                    messages.error(request, f'Insufficient bank balance. Available: {branch_balance.bank_balance}')
                    return redirect('expense')
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance,
                    bank_balance=branch_balance.bank_balance - amount,
                    updated_by=request.user,
                )

            # Create expense record
            Expense.objects.create(
                description=description,
                amount=amount,
                recorded_by=request.user,
                office=branch_office.name,
                transaction_type=transaction_type,
                expense_account=expense_account,
                payment_method=payment_method,
                transaction_date=transaction_date,
                attachment=attachment
            )

            messages.success(request, "Expense recorded successfully.")
            return redirect('expense')

        except Exception as e:
            messages.error(request, f"Error adding expense: {str(e)}")

    return render(request, 'app/expense.html')


def expense_report(request):
    offices = Office.objects.all()
    default_start_date = timezone.now().date() - timedelta(days=28)
    default_end_date = timezone.now().date()

    start_date_str = request.GET.get('start_date', default_start_date.strftime('%Y-%m-%d'))
    end_date_str = request.GET.get('end_date', default_end_date.strftime('%Y-%m-%d'))
    office_filter = request.GET.get('office', '')
    recorded_by_filter = request.GET.get('recorded_by', '')

    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        start_date = default_start_date
        end_date = default_end_date
        start_date_str = start_date.strftime('%Y-%m-%d')
        end_date_str = end_date.strftime('%Y-%m-%d')

    # Filter on transaction_date (user-entered), fall back to expense_date for nulls
    expenses = Expense.objects.filter(
        transaction_date__range=[start_date, end_date]  # ← was expense_date
    ).select_related('recorded_by', 'transaction_type').order_by('-transaction_date')

    if office_filter:
        expenses = expenses.filter(office=office_filter)  # ← exact match, not icontains

    if recorded_by_filter:
        expenses = expenses.filter(recorded_by__username=recorded_by_filter)  # ← exact match

    total_expenses = expenses.count()
    total_amount = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    avg_expense = expenses.aggregate(avg=Avg('amount'))['avg'] or Decimal('0.00')

    unique_offices = Expense.objects.exclude(
        office__isnull=True
    ).exclude(office='').values_list('office', flat=True).distinct().order_by('office')

    unique_users = CustomUser.objects.filter(
        recorded_expenses__in=expenses
    ).distinct().order_by('username')

    highest_expense = expenses.order_by('-amount').first()
    lowest_expense = expenses.order_by('amount').first()

    context = {
        **get_base_context(request),
        'start_date':          start_date,
        'end_date':            end_date,
        'start_date_str':      start_date_str,
        'end_date_str':        end_date_str,
        'office_filter':       office_filter,
        'recorded_by_filter':  recorded_by_filter,
        'expenses':            expenses,
        'total_expenses':      total_expenses,
        'total_amount':        total_amount,
        'avg_expense':         avg_expense,
        'unique_offices':      unique_offices,
        'unique_users':        unique_users,
        'highest_expense':     highest_expense,
        'lowest_expense':      lowest_expense,
        'offices':             offices,

        # ── Add these to match the template ──
        'branch_name':         office_filter.upper() if office_filter else 'ALL BRANCHES',
        'date_from_display':   start_date.strftime('%d %b %Y'),
        'date_to_display':     end_date.strftime('%d %b %Y'),
        'grand_total':         total_amount,
        'rows': [
            {
                'date':        e.transaction_date or e.expense_date,
                'receipt_no':  '—',
                'category':    e.transaction_type.name if e.transaction_type else '—',
                'description': e.description or '',
                'amount':      e.amount,
                'attachment':  e.attachment.url if e.attachment else None,
                'hide_date':   False,  # handled below
            }
            for e in expenses
        ],
    }

    return render(request, 'app/expenses_report.html', context)





# ==========================================================================
from django.http import JsonResponse
def loan_history(request):
    clients = Client.objects.all().order_by('firstname', 'lastname')
    offices = Office.objects.all()
    
    selected_client = None
    loans_data = []
    company_office = Office.objects.first()
    
    if request.method == 'POST' or request.GET.get('client_id'):
        client_id = request.POST.get('client_id') or request.GET.get('client_id')
        selected_client = get_object_or_404(Client, id=client_id)
        
        # Get all loans for this client
        loans = LoanApplication.objects.filter(
            client=selected_client
        ).order_by('-application_date')
        
        for index, loan in enumerate(loans, 1):
            # Calculate total paid for this loan
            total_paid = LoanRepayment.objects.filter(
                loan_application=loan
            ).aggregate(total=Sum('repayment_amount'))['total'] or Decimal('0.00')
            
            # Calculate outstanding balance
            outstanding = loan.total_repayment_amount - total_paid if loan.total_repayment_amount else loan.loan_amount - total_paid
            
            # Determine status
            if outstanding <= 0:
                status = 'Completed'
                status_color = 'success'
            elif loan.status == 'Overdue':
                status = 'Overdue'
                status_color = 'danger'
            elif loan.status == 'Approved':
                status = 'Active'
                status_color = 'primary'
            else:
                status = loan.status
                status_color = 'warning'
            
            # Get repayment schedule
            repayment_schedule = generate_repayment_schedule(loan)
            
            loans_data.append({
                'sn': index,
                'loan': loan,
                'loan_id': f'LON-{loan.id:06d}',
                'loan_type': loan.loan_type,
                'loan_amount': loan.loan_amount,
                'interest_amount': loan.interest_amount or Decimal('0.00'),
                'penalty_amount': calculate_penalty(loan),  # You'll need to implement this
                'total_amount': loan.total_repayment_amount or loan.loan_amount,
                'paid_amount': total_paid,
                'outstanding_balance': outstanding,
                'status': status,
                'status_color': status_color,
                'repayment_schedule': repayment_schedule,
                'application_date': loan.application_date,
                'payment_period_months': loan.payment_period_months,
                'monthly_installment': loan.monthly_installment or Decimal('0.00'),
            })
    
    context = {
        'clients': clients,
        'offices': offices,
        'selected_client': selected_client,
        'loans_data': loans_data,
        'company_office': company_office,
        'total_loans': len(loans_data),
        'total_loan_amount': sum(loan['loan_amount'] for loan in loans_data),
        'total_paid_amount': sum(loan['paid_amount'] for loan in loans_data),
        'total_outstanding': sum(loan['outstanding_balance'] for loan in loans_data),
        'active_loans': sum(1 for loan in loans_data if loan['status'] == 'Active'),
        'completed_loans': sum(1 for loan in loans_data if loan['status'] == 'Completed'),
    }
    
    return render(request, 'app/loan_history.html', context)

def generate_repayment_schedule(loan):
    """Generate repayment schedule for a loan"""
    schedule = []
    
    if not loan.first_repayment_date:
        return schedule
    
    current_date = loan.first_repayment_date
    monthly_payment = loan.monthly_installment or (loan.total_repayment_amount / loan.payment_period_months)
    remaining_balance = loan.total_repayment_amount or loan.loan_amount
    paid_amount = LoanRepayment.objects.filter(
        loan_application=loan
    ).aggregate(total=Sum('repayment_amount'))['total'] or Decimal('0.00')
    
    # Get all repayments
    repayments = LoanRepayment.objects.filter(
        loan_application=loan
    ).order_by('repayment_date')
    
    repayment_dict = {}
    for repayment in repayments:
        date_key = repayment.repayment_date.strftime('%Y-%m')
        if date_key not in repayment_dict:
            repayment_dict[date_key] = Decimal('0.00')
        repayment_dict[date_key] += repayment.repayment_amount
    
    for i in range(loan.payment_period_months):
        due_date = current_date + relativedelta(months=i)
        date_key = due_date.strftime('%Y-%m')
        
        # Calculate payment received for this period
        payment_received = repayment_dict.get(date_key, Decimal('0.00'))
        
        # Determine if paid
        is_paid = payment_received >= monthly_payment
        paid_amount_this_month = min(payment_received, monthly_payment)
        
        schedule.append({
            'installment_no': i + 1,
            'due_date': due_date,
            'due_amount': monthly_payment,
            'paid_amount': paid_amount_this_month,
            'remaining': monthly_payment - paid_amount_this_month,
            'status': 'Paid' if is_paid else 'Pending',
            'status_color': 'success' if is_paid else 'warning',
            'payment_date': next((r.repayment_date for r in repayments 
                                if r.repayment_date.strftime('%Y-%m') == date_key), None)
        })
    
    return schedule

def calculate_penalty(loan):
    """Calculate penalty for overdue loans"""
    # Implement your penalty calculation logic here
    # This is a simple example
    from datetime import date
    from dateutil.relativedelta import relativedelta
    
    if loan.status != 'Overdue' or not loan.first_repayment_date:
        return Decimal('0.00')
    
    # Simple penalty: 5% of monthly installment per month overdue
    today = date.today()
    months_overdue = 0
    current_date = loan.first_repayment_date
    
    while current_date < today:
        if current_date < today:
            months_overdue += 1
        current_date += relativedelta(months=1)
    
    penalty_rate = Decimal('0.05')  # 5% per month
    monthly_installment = loan.monthly_installment or Decimal('0.00')
    penalty = monthly_installment * penalty_rate * months_overdue
    
    return penalty

def get_repayment_schedule_ajax(request, loan_id):
    """AJAX endpoint to get repayment schedule"""
    loan = get_object_or_404(LoanApplication, id=loan_id)
    schedule = generate_repayment_schedule(loan)
    
    schedule_data = []
    for item in schedule:
        schedule_data.append({
            'installment_no': item['installment_no'],
            'due_date': item['due_date'].strftime('%d/%m/%Y'),
            'due_amount': float(item['due_amount']),
            'paid_amount': float(item['paid_amount']),
            'remaining': float(item['remaining']),
            'status': item['status'],
            'status_color': item['status_color'],
            'payment_date': item['payment_date'].strftime('%d/%m/%Y') if item['payment_date'] else '-'
        })
    
    return JsonResponse({
        'success': True,
        'loan_id': f'LON-{loan.id:06d}',
        'client_name': f"{loan.client.firstname} {loan.client.lastname}",
        'loan_type': loan.loan_type,
        'monthly_installment': float(loan.monthly_installment or 0),
        'schedule': schedule_data
    })
# ==========================================================================
from collections import defaultdict
def transaction_statement(request):
    offices = Office.objects.all()
    company_office = Office.objects.first()
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    transaction_type = request.GET.get('transaction_type', 'all')
    office_filter = request.GET.get('office')
    
    # Initialize transactions list
    transactions = []
    
    # Get all loan repayments (Credits - money received)
    repayments = LoanRepayment.objects.all().select_related(
        'loan_application__client', 'processed_by'
    )
    
    # Get all loan disbursements (Debits - money given out)
    loans = LoanApplication.objects.all().select_related(
        'client', 'processed_by'
    )
    
    # Get all office transactions
    office_transactions = OfficeTransaction.objects.all().select_related(
        'office_from', 'office_to'
    )
    
    # Apply date filters
    if start_date:
        start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        repayments = repayments.filter(repayment_date__gte=start_date_obj)
        loans = loans.filter(application_date__gte=start_date_obj)
        office_transactions = office_transactions.filter(transaction_date__gte=start_date_obj)
    
    if end_date:
        end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        repayments = repayments.filter(repayment_date__lte=end_date_obj)
        loans = loans.filter(application_date__lte=end_date_obj)
        office_transactions = office_transactions.filter(transaction_date__lte=end_date_obj)
    
    # Apply office filter
    if office_filter and office_filter != 'all':
        repayments = repayments.filter(loan_application__office=office_filter)
        loans = loans.filter(office=office_filter)
        office_transactions = office_transactions.filter(
            Q(office_from_id=office_filter) | Q(office_to_id=office_filter)
        )
    
    # Apply transaction type filter
    if transaction_type == 'all' or transaction_type == 'repayments':
        # Add loan repayments (CREDIT)
        for repayment in repayments:
            transactions.append({
                'date': repayment.repayment_date,
                'receipt_no': f'RCP-{repayment.id:06d}',
                'name': f"{repayment.loan_application.client.firstname} {repayment.loan_application.client.lastname}",
                'description': f'Loan Repayment - {repayment.loan_application.loan_type}',
                'credit': repayment.repayment_amount,
                'debit': Decimal('0.00'),
                'processed_by': repayment.processed_by.get_full_name() or repayment.processed_by.username,
                'type': 'repayment',
                'category': 'Loan Repayment',
                'reference': f'LON-{repayment.loan_application.id:06d}'
            })
    
    if transaction_type == 'all' or transaction_type == 'disbursements':
        # Add loan disbursements (DEBIT)
        for loan in loans:
            transactions.append({
                'date': loan.application_date,
                'receipt_no': f'LON-{loan.id:06d}',
                'name': f"{loan.client.firstname} {loan.client.lastname}",
                'description': f'Loan Disbursement - {loan.loan_type}',
                'credit': Decimal('0.00'),
                'debit': loan.loan_amount,
                'processed_by': loan.processed_by.get_full_name() or loan.processed_by.username,
                'type': 'disbursement',
                'category': 'Loan Disbursement',
                'reference': f'CLT-{loan.client.id:06d}'
            })
    
    if transaction_type == 'all' or transaction_type == 'transfers':
        # Add office transfers
        for trans in office_transactions:
            transactions.append({
                'date': trans.transaction_date,
                'receipt_no': f'TRF-{trans.id:06d}',
                'name': f"{trans.office_from.name} → {trans.office_to.name}",
                'description': f'Office Transfer - {trans.transaction_type}',
                'credit': trans.amount if trans.office_to else Decimal('0.00'),
                'debit': trans.amount if trans.office_from else Decimal('0.00'),
                'processed_by': 'System',
                'type': 'transfer',
                'category': 'Office Transfer',
                'reference': trans.transaction_type
            })
    
    # Sort all transactions by date (newest first)
    transactions.sort(key=lambda x: x['date'], reverse=True)
    
    # Calculate totals
    total_credit = sum(t['credit'] for t in transactions)
    total_debit = sum(t['debit'] for t in transactions)
    net_balance = total_credit - total_debit
    
    # Get unique offices for filter
    offices_list = Office.objects.all()
    
    context = {
        'offices': offices,
        'company_office': company_office,
        'transactions': transactions,
        'total_credit': total_credit,
        'total_debit': total_debit,
        'net_balance': net_balance,
        'total_transactions': len(transactions),
        'start_date': start_date,
        'end_date': end_date,
        'transaction_type': transaction_type,
        'selected_office': office_filter,
        'offices_list': offices_list,
    }
    
    return render(request, 'app/transaction_statement.html', context)

# ==========================================================================
def loan_collection_statement2(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    office_filter = request.GET.get('office')
    
    # Get all loan repayments with related data
    repayments = LoanRepayment.objects.all().select_related(
        'loan_application',
        'loan_application__client',
        'processed_by'
    ).order_by('repayment_date')
    
    # Apply date filters
    if start_date:
        start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        repayments = repayments.filter(repayment_date__gte=start_date_obj)
    if end_date:
        end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        repayments = repayments.filter(repayment_date__lte=end_date_obj)
    
    # ← ADD: office filter (office lives on LoanApplication)
    if office_filter:
        repayments = repayments.filter(loan_application__office=office_filter)

    # ← ADD: unique offices for dropdown
    offices = LoanApplication.objects.values_list('office', flat=True) \
        .exclude(office__isnull=True).exclude(office='') \
        .distinct().order_by('office')
        
    # Calculate collection data
    collection_data = []
    running_balance = Decimal('0.00')
    
    for index, repayment in enumerate(repayments, 1):
        # Get loan details
        loan = repayment.loan_application
        
        # Calculate interest portion (assuming equal distribution)
        monthly_interest_rate = (loan.interest_rate / Decimal('100')) / Decimal('12')
        total_interest = loan.loan_amount * monthly_interest_rate * loan.payment_period_months
        interest_per_payment = total_interest / loan.payment_period_months if loan.payment_period_months > 0 else 0
        
        # Calculate principal portion
        principal_per_payment = (loan.loan_amount / loan.payment_period_months) if loan.payment_period_months > 0 else 0
        
        # For simplicity in this example, we'll use these calculations
        # You can adjust based on your actual business logic
        interest_amount = repayment.repayment_amount * Decimal('0.2')  # 20% as interest
        principal_amount = repayment.repayment_amount - interest_amount
        
        # Update running balance (optional)
        running_balance += repayment.repayment_amount
        
        collection_data.append({
            'sn': index,
            'date': repayment.repayment_date,
            'receipt_no': f'RCP-{repayment.id:06d}',
            'name': f"{loan.client.firstname} {loan.client.lastname}",
            'description': f"Loan Repayment - {loan.loan_type} (LON-{loan.id:06d})",
            'rate': loan.interest_rate,
            'principal': principal_amount,
            'interest': interest_amount,
            'total': repayment.repayment_amount,
            'created_by': repayment.processed_by.get_full_name() or repayment.processed_by.username,
            'loan_id': loan.id,
            'client_id': loan.client.id
        })
    
    # Calculate totals
    total_principal = sum(item['principal'] for item in collection_data)
    total_interest = sum(item['interest'] for item in collection_data)
    total_collected = sum(item['total'] for item in collection_data)
    
    context = {
        **get_base_context(request),
        'collection_data': collection_data,
        'total_principal': total_principal,
        'total_interest': total_interest,
        'total_collected': total_collected,
        'total_transactions': len(collection_data),
        'start_date': start_date,
        'end_date': end_date,
        'office_filter': office_filter,   # ← ADD
        'offices': offices,               # ← ADD
    }
    
    return render(request, 'app/loan_collection_statement.html', context)

# ===========================================================================================

def bank_cash_transaction_statement(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    office_id = request.GET.get('office_id')
    source_filter = request.GET.get('source')  # bank or cash
    destination_filter = request.GET.get('destination')  # bank or cash
    
    # Get all bank/cash transactions with related data
    transactions = BankCashTransaction.objects.all().select_related(
        'office_from'
    ).order_by('-transaction_date', '-id')  # Order by transaction_date (most recent first)
    
    # Apply date filters
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        transactions = transactions.filter(transaction_date__gte=start_date_obj)
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        transactions = transactions.filter(transaction_date__lte=end_date_obj)
    
    # Apply office filter
    if office_id:
        transactions = transactions.filter(office_from_id=office_id)
    
    # Apply source/destination filters
    if source_filter:
        transactions = transactions.filter(source=source_filter)
    if destination_filter:
        transactions = transactions.filter(destination=destination_filter)
    
    # Calculate transaction data
    transaction_data = []
    running_balance = Decimal('0.00')
    
    # Track bank and cash balances separately
    bank_balance = Decimal('0.00')
    cash_balance = Decimal('0.00')
    
    for index, transaction in enumerate(transactions, 1):
        # Get office details
        office = transaction.office_from
        
        # Determine the effect on balances
        # If source is bank, decrease bank balance; if source is cash, decrease cash balance
        # If destination is bank, increase bank balance; if destination is cash, increase cash balance
        
        # For running balance calculation (net effect on total funds)
        if transaction.source == 'bank' and transaction.destination == 'cash':
            # Bank to Cash: Bank decreases, Cash increases (total funds unchanged)
            bank_effect = -transaction.amount
            cash_effect = +transaction.amount
            transaction_type = 'Bank to Cash Transfer'
        elif transaction.source == 'cash' and transaction.destination == 'bank':
            # Cash to Bank: Cash decreases, Bank increases (total funds unchanged)
            bank_effect = +transaction.amount
            cash_effect = -transaction.amount
            transaction_type = 'Cash to Bank Transfer'
        elif transaction.source == 'bank' and transaction.destination == 'bank':
            # Bank to Bank: No net effect (internal transfer)
            bank_effect = Decimal('0.00')
            cash_effect = Decimal('0.00')
            transaction_type = 'Bank Transfer'
        elif transaction.source == 'cash' and transaction.destination == 'cash':
            # Cash to Cash: No net effect (internal transfer)
            bank_effect = Decimal('0.00')
            cash_effect = Decimal('0.00')
            transaction_type = 'Cash Transfer'
        else:
            # Default case (shouldn't happen with valid data)
            bank_effect = Decimal('0.00')
            cash_effect = Decimal('0.00')
            transaction_type = f"{transaction.source.title()} to {transaction.destination.title()}"
        
        # Update running balances
        running_balance += (bank_effect + cash_effect)  # Total funds change
        bank_balance += bank_effect
        cash_balance += cash_effect
        
        transaction_data.append({
            'sn': index,
            'date': transaction.transaction_date,
            'time': transaction.transaction_date.strftime('%H:%M:%S') if hasattr(transaction.transaction_date, 'strftime') else '00:00:00',
            'transaction_id': transaction.id,
            'receipt_no': f'BCT-{transaction.id:06d}',
            'office': {
                'id': office.id,
                'name': office.name,
                'region': office.region if hasattr(office, 'region') else ''
            },
            'source': transaction.source,
            'source_display': transaction.source.title(),
            'destination': transaction.destination,
            'destination_display': transaction.destination.title(),
            'transaction_type': transaction_type,
            'amount': transaction.amount,
            'bank_effect': bank_effect,
            'cash_effect': cash_effect,
            'net_effect': bank_effect + cash_effect,
            'running_balance': running_balance,
            'bank_balance': bank_balance,
            'cash_balance': cash_balance,
            'description': f"{transaction.source.title()} to {transaction.destination.title()} - {office.name}",
            'created_at': transaction.transaction_date,
        })
    
    # Calculate summaries
    total_transactions = len(transaction_data)
    
    # Calculate totals by transaction type
    bank_to_cash_total = sum(
        t['amount'] for t in transaction_data 
        if t['source'] == 'bank' and t['destination'] == 'cash'
    )
    cash_to_bank_total = sum(
        t['amount'] for t in transaction_data 
        if t['source'] == 'cash' and t['destination'] == 'bank'
    )
    
    # Calculate totals by source/destination
    total_bank_out = sum(
        t['amount'] for t in transaction_data 
        if t['source'] == 'bank'
    )
    total_cash_out = sum(
        t['amount'] for t in transaction_data 
        if t['source'] == 'cash'
    )
    total_bank_in = sum(
        t['amount'] for t in transaction_data 
        if t['destination'] == 'bank'
    )
    total_cash_in = sum(
        t['amount'] for t in transaction_data 
        if t['destination'] == 'cash'
    )
    
    # Net totals
    net_bank_flow = total_bank_in - total_bank_out
    net_cash_flow = total_cash_in - total_cash_out
    total_transaction_amount = sum(t['amount'] for t in transaction_data)
    
    # Group by office
    office_summary = {}
    for transaction in transaction_data:
        office_name = transaction['office']['name']
        if office_name not in office_summary:
            office_summary[office_name] = {
                'office_id': transaction['office']['id'],
                'office_name': office_name,
                'region': transaction['office']['region'],
                'total_transactions': 0,
                'bank_to_cash': Decimal('0.00'),
                'cash_to_bank': Decimal('0.00'),
                'bank_out': Decimal('0.00'),
                'cash_out': Decimal('0.00'),
                'bank_in': Decimal('0.00'),
                'cash_in': Decimal('0.00'),
                'total_amount': Decimal('0.00'),
            }
        
        office_summary[office_name]['total_transactions'] += 1
        office_summary[office_name]['total_amount'] += transaction['amount']
        
        if transaction['source'] == 'bank' and transaction['destination'] == 'cash':
            office_summary[office_name]['bank_to_cash'] += transaction['amount']
        elif transaction['source'] == 'cash' and transaction['destination'] == 'bank':
            office_summary[office_name]['cash_to_bank'] += transaction['amount']
        
        if transaction['source'] == 'bank':
            office_summary[office_name]['bank_out'] += transaction['amount']
        elif transaction['source'] == 'cash':
            office_summary[office_name]['cash_out'] += transaction['amount']
        
        if transaction['destination'] == 'bank':
            office_summary[office_name]['bank_in'] += transaction['amount']
        elif transaction['destination'] == 'cash':
            office_summary[office_name]['cash_in'] += transaction['amount']
    
    # Get all offices for filter dropdown
    offices = Office.objects.all()
    
    # Get date range for display
    if transactions.exists():
        # Since we're ordering by -transaction_date, first() is newest, last() is oldest
        first_date = transactions.last().transaction_date if transactions.last() else None  # Oldest
        last_date = transactions.first().transaction_date if transactions.first() else None   # Newest
    else:
        first_date = None
        last_date = None
    
    context = {
        'transaction_data': transaction_data,
        'office_summary': office_summary,
        'offices': offices,
        'total_transactions': total_transactions,
        'total_bank_to_cash': bank_to_cash_total,
        'total_cash_to_bank': cash_to_bank_total,
        'total_bank_out': total_bank_out,
        'total_cash_out': total_cash_out,
        'total_bank_in': total_bank_in,
        'total_cash_in': total_cash_in,
        'net_bank_flow': net_bank_flow,
        'net_cash_flow': net_cash_flow,
        'total_transaction_amount': total_transaction_amount,
        'start_date': start_date,
        'end_date': end_date,
        'office_id': office_id,
        'source_filter': source_filter,
        'destination_filter': destination_filter,
        'first_date': first_date,
        'last_date': last_date,
        'bank_balance': bank_balance,
        'cash_balance': cash_balance,
        'total_balance': bank_balance + cash_balance,
    }
    
    return render(request, 'app/bank_cash_transaction_statement.html', context)



import calendar
import datetime
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from django.shortcuts import render
from .models import LoanApplication, Office

def branches_loan_report(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status_filter = request.GET.get('status')
    loan_type_filter = request.GET.get('loan_type')
    search_query = request.GET.get('search')
    office_filter = request.GET.get('office')
    region_filter = request.GET.get('region')
    
    # Base queryset with related data
    loans = LoanApplication.objects.all().select_related(
        'client', 
        'processed_by'
    ).order_by('-application_date')
    
    # Apply filters
    if start_date:
        start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        loans = loans.filter(application_date__gte=start_date_obj)
    if end_date:
        end_date_obj = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
        loans = loans.filter(application_date__lte=end_date_obj)
    
    if status_filter:
        loans = loans.filter(status=status_filter)
    
    if loan_type_filter:
        loans = loans.filter(loan_type=loan_type_filter)
    
    if office_filter:
        loans = loans.filter(office=office_filter)
    
    if region_filter:
        loans = loans.filter(client__region=region_filter)
    
    if search_query:
        loans = loans.filter(
            Q(client__firstname__icontains=search_query) |
            Q(client__lastname__icontains=search_query) |
            Q(client__phonenumber__icontains=search_query) |
            Q(client__checkno__icontains=search_query) |
            Q(client__employmentcardno__icontains=search_query)
        )
    
    # Calculate loan data with repayment information
    loan_data = []
    
    # Summary totals
    total_loan_amount = Decimal('0.00')
    total_paid_amount = Decimal('0.00')
    total_outstanding = Decimal('0.00')
    total_interest = Decimal('0.00')
    
    for index, loan in enumerate(loans, 1):
        # Get client information
        client = loan.client
        
        # Calculate total paid amount from repayments
        paid_amount = loan.repayments.aggregate(
            total=Coalesce(Sum('repayment_amount'), Decimal('0.00'))
        )['total']
        
        # Calculate outstanding amount
        # Use repayment_amount_remaining if available, otherwise calculate
        if hasattr(loan, 'repayment_amount_remaining') and loan.repayment_amount_remaining:
            outstanding = loan.repayment_amount_remaining
        else:
            outstanding = loan.total_repayment_amount - paid_amount if loan.total_repayment_amount else loan.loan_amount - paid_amount
        
        # Calculate payment progress
        if loan.total_repayment_amount and loan.total_repayment_amount > 0:
            payment_percentage = (paid_amount / loan.total_repayment_amount) * 100
        else:
            payment_percentage = 0
        
        # Determine loan status based on payment
        if loan.status.lower() == 'approved':
            if outstanding <= 0:
                current_status = 'Completed'
                status_color = 'success'
            elif paid_amount > 0:
                current_status = 'Active'
                status_color = 'primary'
            else:
                current_status = 'Approved'
                status_color = 'info'
        else:
            current_status = loan.status
            status_color = 'warning' if loan.status == 'Pending' else 'secondary'
        
        # Calculate next payment due date (if applicable)
        next_payment_date = None
        if loan.status.lower() == 'approved' and outstanding > 0 and loan.first_repayment_date:
            # Calculate how many payments have been made
            payments_made = loan.repayments.count()
            if payments_made < loan.payment_period_months:
                # Next payment date = first_repayment_date + (payments_made * 1 month)
                next_payment_date = loan.first_repayment_date + relativedelta(months=payments_made)
        
        loan_data.append({
            'sn': index,
            'loan_id': loan.id,
            'application_date': loan.application_date,
            'client': {
                'id': client.id,
                'full_name': f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
                'firstname': client.firstname,
                'middlename': client.middlename,
                'lastname': client.lastname,
                'phonenumber': client.phonenumber or 'N/A',
                'region': client.region or 'N/A',
                'district': client.district or 'N/A',
                'checkno': client.checkno or 'N/A',
                'employmentcardno': client.employmentcardno or 'N/A',
                'employername': client.employername or 'N/A',
            },
            'loan_type': loan.loan_type,
            'loan_amount': loan.loan_amount,
            'interest_rate': loan.interest_rate,
            'interest_amount': loan.interest_amount or 0,
            'total_interest': loan.total_interest_amount or 0,
            'total_repayment_amount': loan.total_repayment_amount or loan.loan_amount,
            'payment_period_months': loan.payment_period_months,
            'monthly_installment': loan.monthly_installment or 0,
            'paid_amount': paid_amount,
            'outstanding': outstanding,
            'payment_percentage': round(payment_percentage, 2),
            'status': current_status,
            'status_color': status_color,
            'original_status': loan.status,
            'office': loan.office or 'N/A',
            'processed_by': loan.processed_by.get_full_name() or loan.processed_by.username,
            'first_repayment_date': loan.first_repayment_date,
            'next_payment_date': next_payment_date,
            'last_repayment_date': loan.repayments.order_by('-repayment_date').first().repayment_date if loan.repayments.exists() else None,
            'repayments_count': loan.repayments.count(),
        })
        
        # Update totals
        total_loan_amount += loan.loan_amount
        total_paid_amount += paid_amount
        total_outstanding += outstanding
        total_interest += (loan.total_interest_amount or 0)
    
    # Get unique values for filter dropdowns
    loan_types = LoanApplication.objects.values_list('loan_type', flat=True).distinct().order_by('loan_type')
    offices = LoanApplication.objects.values_list('office', flat=True).exclude(office__isnull=True).exclude(office='').distinct().order_by('office')
    regions = Client.objects.values_list('region', flat=True).exclude(region__isnull=True).exclude(region='').distinct().order_by('region')
    statuses = ['Pending', 'Approved', 'Rejected', 'Completed', 'Active']
    
    # Calculate summary statistics
    summary = {
        'total_loans': len(loan_data),
        'total_loan_amount': total_loan_amount,
        'total_paid_amount': total_paid_amount,
        'total_outstanding': total_outstanding,
        'total_interest': total_interest,
        'average_loan_amount': total_loan_amount / len(loan_data) if loan_data else 0,
        'collection_rate': (total_paid_amount / total_loan_amount * 100) if total_loan_amount > 0 else 0,
        'active_loans': sum(1 for loan in loan_data if loan['status'] in ['Active', 'Approved']),
        'completed_loans': sum(1 for loan in loan_data if loan['status'] == 'Completed'),
        'pending_loans': sum(1 for loan in loan_data if loan['status'] == 'Pending'),
        'overdue_loans': sum(1 for loan in loan_data if loan['status'] == 'Active' and loan['next_payment_date'] and loan['next_payment_date'] < datetime.datetime.now().date()),
    }
    
    # Group by loan type for chart data
    loan_type_summary = {}
    for loan in loan_data:
        loan_type = loan['loan_type']
        if loan_type not in loan_type_summary:
            loan_type_summary[loan_type] = {
                'count': 0,
                'total_amount': Decimal('0.00'),
                'paid_amount': Decimal('0.00'),
                'outstanding': Decimal('0.00'),
            }
        loan_type_summary[loan_type]['count'] += 1
        loan_type_summary[loan_type]['total_amount'] += loan['loan_amount']
        loan_type_summary[loan_type]['paid_amount'] += loan['paid_amount']
        loan_type_summary[loan_type]['outstanding'] += loan['outstanding']
    
    context = {
        **get_base_context(request),
        'loan_data': loan_data,
        'summary': summary,
        'loan_type_summary': loan_type_summary,
        'loan_types': loan_types,
        'offices': offices,
        'regions': regions,
        'statuses': statuses,
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'status': status_filter,
            'loan_type': loan_type_filter,
            'search': search_query,
            'office': office_filter,
            'region': region_filter,
        },
        'today': timezone.now().date(),
    }
    
    return render(request, 'app/branches_loan_report.html', context)


from django.db.models import Sum, Q, F, DecimalField
def expired_loans_report(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    loan_type_filter = request.GET.get('loan_type')
    search_query = request.GET.get('search')
    office_filter = request.GET.get('office')
    region_filter = request.GET.get('region')
    days_overdue_filter = request.GET.get('days_overdue')
    
    # Current date for calculations
    
    today = timezone.now().date()
    
    # Base queryset - only approved loans that are not fully paid
    loans = LoanApplication.objects.filter(
        status__iexact='approved',
        total_repayment_amount__gt=F('repayment_amount_remaining')
    ).select_related(
        'client', 
        'processed_by'
    ).order_by('-application_date')
    
    # Calculate overdue loans
    loan_data = []
    
    # Summary totals
    total_expired_loans = 0
    total_expired_amount = Decimal('0.00')
    total_overdue_amount = Decimal('0.00')
    total_paid_on_expired = Decimal('0.00')
    total_interest_on_expired = Decimal('0.00')
    
    # Overdue categories
    overdue_categories = {
        '0-30': {'count': 0, 'amount': Decimal('0.00')},
        '31-60': {'count': 0, 'amount': Decimal('0.00')},
        '61-90': {'count': 0, 'amount': Decimal('0.00')},
        '91-180': {'count': 0, 'amount': Decimal('0.00')},
        '180+': {'count': 0, 'amount': Decimal('0.00')},
    }
    
    for index, loan in enumerate(loans, 1):
        # Get client information
        client = loan.client
        
        # Calculate total paid amount from repayments
        paid_amount = loan.repayments.aggregate(
            total=Coalesce(Sum('repayment_amount'), Decimal('0.00'))
        )['total']
        
        # Calculate outstanding amount
        outstanding = loan.repayment_amount_remaining if loan.repayment_amount_remaining else (loan.total_repayment_amount - paid_amount)
        
        # Calculate expected payments up to today
        expected_payments = 0
        if loan.first_repayment_date:
            # Calculate how many payments should have been made by today
            months_since_first = relativedelta(today, loan.first_repayment_date).months
            years_since_first = relativedelta(today, loan.first_repayment_date).years
            total_months_since_first = months_since_first + (years_since_first * 12)
            
            # Expected payments = min(months_since_first + 1, payment_period_months)
            expected_payments = min(total_months_since_first + 1, loan.payment_period_months)
            
            # Calculate expected amount to be paid by now
            expected_amount = expected_payments * (loan.monthly_installment or 0)
        else:
            expected_amount = 0
        
        # Calculate overdue amount
        overdue_amount = max(expected_amount - paid_amount, Decimal('0.00'))
        
        # Calculate days overdue
        days_overdue = 0
        next_payment_date = None
        last_payment_date = None
        
        if loan.first_repayment_date and outstanding > 0:
            # Get last payment date
            last_repayment = loan.repayments.order_by('-repayment_date').first()
            if last_repayment:
                last_payment_date = last_repayment.repayment_date
                
                # Calculate next payment due date
                payments_made = loan.repayments.count()
                if payments_made < loan.payment_period_months:
                    next_payment_date = loan.first_repayment_date + relativedelta(months=payments_made)
                    
                    # Calculate days overdue
                    if next_payment_date < today:
                        days_overdue = (today - next_payment_date).days
            else:
                # No payments made yet
                if loan.first_repayment_date < today:
                    days_overdue = (today - loan.first_repayment_date).days
                    next_payment_date = loan.first_repayment_date
        
        # Only include loans that are actually overdue
        if days_overdue > 0 or overdue_amount > 0:
            total_expired_loans += 1
            total_expired_amount += loan.loan_amount
            total_overdue_amount += overdue_amount
            total_paid_on_expired += paid_amount
            total_interest_on_expired += (loan.total_interest_amount or 0)
            
            # Categorize by days overdue
            if days_overdue <= 30:
                overdue_categories['0-30']['count'] += 1
                overdue_categories['0-30']['amount'] += overdue_amount
            elif days_overdue <= 60:
                overdue_categories['31-60']['count'] += 1
                overdue_categories['31-60']['amount'] += overdue_amount
            elif days_overdue <= 90:
                overdue_categories['61-90']['count'] += 1
                overdue_categories['61-90']['amount'] += overdue_amount
            elif days_overdue <= 180:
                overdue_categories['91-180']['count'] += 1
                overdue_categories['91-180']['amount'] += overdue_amount
            else:
                overdue_categories['180+']['count'] += 1
                overdue_categories['180+']['amount'] += overdue_amount
            
            # Calculate payment progress
            if loan.total_repayment_amount and loan.total_repayment_amount > 0:
                payment_percentage = (paid_amount / loan.total_repayment_amount) * 100
            else:
                payment_percentage = 0
            
            # Determine severity based on days overdue
            if days_overdue <= 30:
                severity = 'Mild'
                severity_color = 'warning'
            elif days_overdue <= 60:
                severity = 'Moderate'
                severity_color = 'orange'
            elif days_overdue <= 90:
                severity = 'Serious'
                severity_color = 'danger'
            else:
                severity = 'Critical'
                severity_color = 'dark'
            
            loan_data.append({
                'sn': len(loan_data) + 1,
                'loan_id': loan.id,
                'application_date': loan.application_date,
                'client': {
                    'id': client.id,
                    'full_name': f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
                    'firstname': client.firstname,
                    'middlename': client.middlename,
                    'lastname': client.lastname,
                    'phonenumber': client.phonenumber or 'N/A',
                    'region': client.region or 'N/A',
                    'district': client.district or 'N/A',
                    'checkno': client.checkno or 'N/A',
                    'employmentcardno': client.employmentcardno or 'N/A',
                    'employername': client.employername or 'N/A',
                },
                'loan_type': loan.loan_type,
                'loan_amount': loan.loan_amount,
                'interest_rate': loan.interest_rate,
                'total_interest': loan.total_interest_amount or 0,
                'total_repayment_amount': loan.total_repayment_amount or loan.loan_amount,
                'monthly_installment': loan.monthly_installment or 0,
                'paid_amount': paid_amount,
                'outstanding': outstanding,
                'overdue_amount': overdue_amount,
                'expected_amount': expected_amount,
                'payment_percentage': round(payment_percentage, 2),
                'days_overdue': days_overdue,
                'severity': severity,
                'severity_color': severity_color,
                'office': loan.office or 'N/A',
                'processed_by': loan.processed_by.get_full_name() or loan.processed_by.username,
                'first_repayment_date': loan.first_repayment_date,
                'next_payment_date': next_payment_date,
                'last_payment_date': last_payment_date,
                'repayments_count': loan.repayments.count(),
                'expected_payments': expected_payments,
                'actual_payments': loan.repayments.count(),
            })
    
    # Apply additional filters after calculation
    if days_overdue_filter:
        if days_overdue_filter == '0-30':
            loan_data = [l for l in loan_data if l['days_overdue'] <= 30]
        elif days_overdue_filter == '31-60':
            loan_data = [l for l in loan_data if 31 <= l['days_overdue'] <= 60]
        elif days_overdue_filter == '61-90':
            loan_data = [l for l in loan_data if 61 <= l['days_overdue'] <= 90]
        elif days_overdue_filter == '91-180':
            loan_data = [l for l in loan_data if 91 <= l['days_overdue'] <= 180]
        elif days_overdue_filter == '180+':
            loan_data = [l for l in loan_data if l['days_overdue'] > 180]
    
    if loan_type_filter:
        loan_data = [l for l in loan_data if l['loan_type'] == loan_type_filter]
    
    if office_filter:
        loan_data = [l for l in loan_data if l['office'] == office_filter]
    
    if region_filter:
        loan_data = [l for l in loan_data if l['client']['region'] == region_filter]
    
    if search_query:
        search_query = search_query.lower()
        loan_data = [l for l in loan_data if 
                    search_query in l['client']['full_name'].lower() or
                    search_query in l['client']['phonenumber'].lower() or
                    search_query in l['client']['checkno'].lower()]
    
    if start_date:
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        loan_data = [l for l in loan_data if l['application_date'] >= start_date_obj]
    
    if end_date:
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        loan_data = [l for l in loan_data if l['application_date'] <= end_date_obj]
    
    # Update totals after filtering
    total_expired_loans = len(loan_data)
    total_expired_amount = sum(l['loan_amount'] for l in loan_data)
    total_overdue_amount = sum(l['overdue_amount'] for l in loan_data)
    total_paid_on_expired = sum(l['paid_amount'] for l in loan_data)
    total_interest_on_expired = sum(l['total_interest'] for l in loan_data)
    
    # Recalculate categories after filtering
    overdue_categories = {
        '0-30': {'count': 0, 'amount': Decimal('0.00')},
        '31-60': {'count': 0, 'amount': Decimal('0.00')},
        '61-90': {'count': 0, 'amount': Decimal('0.00')},
        '91-180': {'count': 0, 'amount': Decimal('0.00')},
        '180+': {'count': 0, 'amount': Decimal('0.00')},
    }
    
    for loan in loan_data:
        days = loan['days_overdue']
        amount = loan['overdue_amount']
        
        if days <= 30:
            overdue_categories['0-30']['count'] += 1
            overdue_categories['0-30']['amount'] += amount
        elif days <= 60:
            overdue_categories['31-60']['count'] += 1
            overdue_categories['31-60']['amount'] += amount
        elif days <= 90:
            overdue_categories['61-90']['count'] += 1
            overdue_categories['61-90']['amount'] += amount
        elif days <= 180:
            overdue_categories['91-180']['count'] += 1
            overdue_categories['91-180']['amount'] += amount
        else:
            overdue_categories['180+']['count'] += 1
            overdue_categories['180+']['amount'] += amount
    
    # Get unique values for filter dropdowns
    loan_types = LoanApplication.objects.values_list('loan_type', flat=True).distinct().order_by('loan_type')
    offices = LoanApplication.objects.values_list('office', flat=True).exclude(office__isnull=True).exclude(office='').distinct().order_by('office')
    regions = Client.objects.values_list('region', flat=True).exclude(region__isnull=True).exclude(region='').distinct().order_by('region')
    
    # Calculate summary statistics
    summary = {
        'total_expired_loans': total_expired_loans,
        'total_expired_amount': total_expired_amount,
        'total_overdue_amount': total_overdue_amount,
        'total_paid_on_expired': total_paid_on_expired,
        'total_interest_on_expired': total_interest_on_expired,
        'average_overdue': total_overdue_amount / total_expired_loans if total_expired_loans > 0 else 0,
        'average_days': sum(l['days_overdue'] for l in loan_data) / total_expired_loans if total_expired_loans > 0 else 0,
        'max_days': max((l['days_overdue'] for l in loan_data), default=0),
        'collection_rate_on_expired': (total_paid_on_expired / (total_paid_on_expired + total_overdue_amount) * 100) if (total_paid_on_expired + total_overdue_amount) > 0 else 0,
    }
    
    context = {
        'loan_data': loan_data,
        'summary': summary,
        'overdue_categories': overdue_categories,
        'loan_types': loan_types,
        'offices': offices,
        'regions': regions,
        'filters': {
            'start_date': start_date,
            'end_date': end_date,
            'loan_type': loan_type_filter,
            'search': search_query,
            'office': office_filter,
            'region': region_filter,
            'days_overdue': days_overdue_filter,
        },
        'today': today,
    }
    
    return render(request, 'app/expired_loans_report.html', context)


import datetime
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from django.shortcuts import render
from .models import LoanApplication, Office


# def loans_owed_report(request):
#     """
#     Loans Owed Report view.
#     Groups active loans by their first_repayment_date month.
#     Always shows exactly 12 months per group (from the group's start month).
#     Filters: month (first repayment month), year, office.
#     Builds hama_rows: clients with any outstanding/partial month (for HAMA table) for previously year meaning last year of 12 month back from now.
#     """

#     today = datetime.date.today()

#     # ---- Filter params -------------------------------------------------
#     selected_month  = request.GET.get('month', '')
#     selected_year   = request.GET.get('year', str(today.year))
#     selected_office = request.GET.get('office', '')

#     # ---- Base queryset -------------------------------------------------
#     loans = (
#         LoanApplication.objects
#         .filter(repayment_amount_remaining__gt=0)
#         .select_related('client')
#         .prefetch_related('repayments')
#         .order_by('first_repayment_date', 'client__firstname')
#     )

#     if selected_year:
#         try:
#             loans = loans.filter(first_repayment_date__year=int(selected_year))
#         except ValueError:
#             pass

#     if selected_month:
#         try:
#             loans = loans.filter(first_repayment_date__month=int(selected_month))
#         except ValueError:
#             pass

#     if selected_office:
#         loans = loans.filter(office=selected_office)

#     # ---- Offices & year range ------------------------------------------
#     offices = Office.objects.all().order_by('name')

#     loan_offices = (
#         LoanApplication.objects
#         .exclude(office__isnull=True).exclude(office='')
#         .values_list('office', flat=True)
#         .distinct().order_by('office')
#     )

#     earliest = LoanApplication.objects.order_by('first_repayment_date').first()
#     start_year = (
#         earliest.first_repayment_date.year
#         if earliest and earliest.first_repayment_date
#         else today.year
#     )
#     year_range = list(range(start_year, today.year + 2))

#     # ---- Build per-loan schedule data ----------------------------------
#     groups = defaultdict(list)

#     for loan in loans:
#         if not loan.first_repayment_date:
#             continue

#         start = loan.first_repayment_date.replace(day=1)

#         months_elapsed = (
#             (today.year  - start.year)  * 12 +
#             (today.month - start.month) + 1
#         )
#         total_months = max(
#             loan.payment_period_months or 12,
#             12,
#             months_elapsed
#         )

#         repayments_by_month = defaultdict(float)
#         for r in loan.repayments.all():
#             key = (r.repayment_date.year, r.repayment_date.month)
#             repayments_by_month[key] += float(r.repayment_amount)

#         schedule = []
#         for i in range(total_months):
#             month_date    = start + relativedelta(months=i)
#             paid_in_month = repayments_by_month.get(
#                 (month_date.year, month_date.month), 0.0
#             )
#             expected    = float(loan.monthly_installment or 0)
#             in_schedule = i < (loan.payment_period_months or 12)
#             out         = max(expected - paid_in_month, 0) if in_schedule else 0.0
#             is_fully_paid = paid_in_month >= expected and expected > 0 and in_schedule

#             schedule.append({
#                 'month':         month_date.strftime('%b-%Y'),
#                 'month_date':    month_date,
#                 'paid':          paid_in_month,
#                 'out':           out,
#                 'expected':      expected if in_schedule else 0.0,
#                 'is_fully_paid': is_fully_paid,
#             })

#         total_paid = (
#             float(loan.total_repayment_amount or 0) -
#             float(loan.repayment_amount_remaining or 0)
#         )

#         groups[start].append({
#             'loan':          loan,
#             'client':        loan.client,
#             'schedule':      schedule,
#             'total_paid':    max(total_paid, 0),
#             'loaned_amount': float(loan.loan_amount or 0),
#             'total_amount':  float(loan.total_repayment_amount or 0),
#             'balance':       float(loan.repayment_amount_remaining or 0),
#         })

#     # ---- Sort & assemble group_data ------------------------------------
#     group_data = []

#     for start_month, loan_list in sorted(groups.items()):

#         # Master column list
#         seen = set()
#         all_months = []
#         for item in loan_list:
#             for entry in item['schedule']:
#                 if entry['month'] not in seen:
#                     seen.add(entry['month'])
#                     all_months.append({'label': entry['month'], 'date': entry['month_date']})
#         all_months.sort(key=lambda x: x['date'])

#         # Pad to at least 12 months
#         while len(all_months) < 12:
#             next_date = all_months[-1]['date'] + relativedelta(months=1) if all_months else start_month
#             all_months.append({'label': next_date.strftime('%b-%Y'), 'date': next_date})

#         # Align schedules
#         for item in loan_list:
#             sched_map = {e['month']: e for e in item['schedule']}
#             item['aligned_schedule'] = [
#                 sched_map.get(m['label'], {
#                     'month': m['label'], 'month_date': m['date'],
#                     'paid': 0.0, 'out': 0.0, 'expected': 0.0, 'is_fully_paid': False,
#                 })
#                 for m in all_months
#             ]

#         # Per-column totals + detail rows
#         month_totals = []
#         for col_idx, m in enumerate(all_months):
#             detail_rows = []
#             for item in loan_list:
#                 entry = item['aligned_schedule'][col_idx]
#                 detail_rows.append({
#                     'client_name':   '{} {}'.format(item['client'].firstname, item['client'].lastname),
#                     'loan_id':       item['loan'].id,
#                     'expected':      entry['expected'],
#                     'paid':          entry['paid'],
#                     'out':           entry['out'],
#                     'is_fully_paid': entry['is_fully_paid'],
#                     'has_schedule':  entry['expected'] > 0,
#                     'is_partial':    entry['paid'] > 0 and not entry['is_fully_paid'] and entry['expected'] > 0,
#                 })
#             month_totals.append({
#                 'month':       m['label'],
#                 'total_paid':  sum(r['paid'] for r in detail_rows),
#                 'total_out':   sum(r['out']  for r in detail_rows),
#                 'detail_rows': detail_rows,
#             })

#         # ---- HAMA rows: focus on the selected month from the top filter,
#         # falling back to today's month if no filter is selected
#         if selected_month and selected_year:
#             month_int = int(selected_month)
#             year_int  = int(selected_year)
#             current_month_label = '{}-{}'.format(
#                 calendar.month_abbr[month_int], year_int
#             )
#         else:
#             current_month_label = today.strftime('%b-%Y')

#         current_col_idx = next(
#             (i for i, m in enumerate(all_months) if m['label'] == current_month_label),
#             None
#         )

#         hama_rows = []
#         for item in loan_list:
#             # Skip if current month not in this group's columns
#             if current_col_idx is None:
#                 continue
#             entry = item['aligned_schedule'][current_col_idx]
#             # Only include clients who have NOT fully paid this month
#             if entry['is_fully_paid'] or entry['expected'] == 0:
#                 continue

#             # Build month_cells: only current month has data, all others are empty
#             month_cells = []
#             for col_idx, col_entry in enumerate(item['aligned_schedule']):
#                 if col_idx == current_col_idx:
#                     if col_entry['paid'] > 0:
#                         # Partial payment
#                         month_cells.append({
#                             'type':    'partial',
#                             'paid':    col_entry['paid'],
#                             'out':     col_entry['out'],
#                             'expected': col_entry['expected'],
#                         })
#                     else:
#                         # Completely unpaid
#                         month_cells.append({
#                             'type':    'out',
#                             'paid':    0,
#                             'out':     col_entry['out'],
#                             'expected': col_entry['expected'],
#                         })
#                 else:
#                     month_cells.append({'type': 'empty', 'paid': 0, 'out': 0, 'expected': 0})

#             client = item['client']
#             name_parts = [client.firstname]
#             if getattr(client, 'middlename', ''):
#                 name_parts.append(client.middlename[:1] + '.')
#             name_parts.append(client.lastname)

#             hama_rows.append({
#                 'client_name':   ' '.join(name_parts),
#                 'check_no':      getattr(client, 'checkno', '') or '-',
#                 'loan_type':     'mafinga-{}'.format(item['loan'].id),
#                 'month_cells':   month_cells,
#                 'current_month': current_month_label,
#                 'total_paid':    item['total_paid'],
#                 'loaned_amount': item['loaned_amount'],
#                 'interest':      item['total_amount'] - item['loaned_amount'],
#                 'total_amount':  item['total_amount'],
#                 'balance':       item['balance'],
#             })

#         # HAMA footer: totals only for the current month column
#         hama_col_totals = []
#         for ci in range(len(all_months)):
#             if ci == current_col_idx and hama_rows:
#                 col_paid = sum(
#                     item['aligned_schedule'][ci]['paid']
#                     for item in loan_list
#                     if not item['aligned_schedule'][ci]['is_fully_paid']
#                     and item['aligned_schedule'][ci]['expected'] > 0
#                 )
#                 col_out = sum(
#                     item['aligned_schedule'][ci]['out']
#                     for item in loan_list
#                     if not item['aligned_schedule'][ci]['is_fully_paid']
#                     and item['aligned_schedule'][ci]['expected'] > 0
#                 )
#                 hama_col_totals.append({'paid': col_paid, 'out': col_out})
#             else:
#                 hama_col_totals.append({'paid': 0, 'out': 0})

#         group_data.append({
#             'start_month':         start_month.strftime('%B %Y').upper(),
#             'start_month_raw':     start_month,
#             'months':              all_months,
#             'loans':               loan_list,
#             'month_totals':        month_totals,
#             'hama_rows':           hama_rows,
#             'hama_col_totals':     hama_col_totals,
#             'group_grand_paid':    sum(i['total_paid']    for i in loan_list),
#             'group_grand_loaned':  sum(i['loaned_amount'] for i in loan_list),
#             'group_grand_total':   sum(i['total_amount']  for i in loan_list),
#             'group_grand_balance': sum(i['balance']       for i in loan_list),
#         })

#     month_names = [
#         (1,  'January'),  (2,  'February'), (3,  'March'),
#         (4,  'April'),    (5,  'May'),      (6,  'June'),
#         (7,  'July'),     (8,  'August'),   (9,  'September'),
#         (10, 'October'), (11, 'November'), (12, 'December'),
#     ]

#     context = {
#         'group_data':      group_data,
#         'today':           today,
#         'offices':         offices,
#         'loan_offices':    loan_offices,
#         'year_range':      year_range,
#         'month_names':     month_names,
#         'selected_month':  selected_month,
#         'selected_year':   selected_year,
#         'selected_office': selected_office,
#         'total_loans':     sum(len(g['loans']) for g in group_data),
#     }
#     return render(request, 'app/loans_owed_report.html', context)


from django.views.decorators.http import require_POST

@require_POST
def loans_owed_approve(request):
    """
    Approve a fully-paid loan (single button) or a batch of checked loans.
    Only loans with repayment_amount_remaining == 0 are approved.
    After approval is_approved=True — they vanish from the owed report.
    """
    single   = request.POST.get('single_loan_id')
    loan_ids = request.POST.getlist('loan_ids')

    ids_to_approve = [single] if single else loan_ids

    if ids_to_approve:
        approved = LoanApplication.objects.filter(
            id__in=ids_to_approve,
            repayment_amount_remaining=0,   # safety: only fully-paid
        ).update(is_approved=True)

        if approved:
            messages.success(request, f'{approved} loan(s) approved and removed from report.')
        else:
            messages.warning(request, 'Selected loan(s) are not fully paid yet.')
    else:
        messages.error(request, 'No loans selected.')

    return redirect(request.META.get('HTTP_REFERER', 'loans_owed_report'))


def transfer_staff(request):
    staff_members = CustomUser.objects.all().order_by('-id')
    context = {
        **get_base_context(request),
        'staff_members': staff_members,
        
    }
    return render(request, 'app/transfer_staff.html', context)

def process_transfer_staff(request):
    if request.method == 'POST':
        staff_id = request.POST.get('staff_id')
        new_office_id = request.POST.get('new_office')
        
        try:
            staff_member = CustomUser.objects.get(id=staff_id)
            new_office = Office.objects.get(id=new_office_id)
            
            # Check if staff is active before transferring
            # if staff_member.status != 'active':
            #     messages.error(request, f'Cannot transfer {staff_member.get_full_name()} because they are not active.')
            #     return redirect('transfer_staff')
            
            # Update the staff member's office
            staff_member.office_allocation = new_office
            staff_member.save()
            
            messages.success(request, f'{staff_member.get_full_name()} successfully transferred to {new_office.name}.')
            
        except CustomUser.DoesNotExist:
            messages.error(request, 'Staff member not found.')
        except Office.DoesNotExist:
            messages.error(request, 'Office not found.')
        except Exception as e:
            messages.error(request, f'Error transferring staff: {str(e)}')
            
    return redirect('transfer_staff')



def block_user(request):
    """View to display all users for blocking/unblocking"""
    users = CustomUser.objects.all().order_by('-id')
    context = {
        'users': users,
        **get_base_context(request),
    }
    return render(request, 'app/block_user.html', context)

def process_block_user(request):
    """Process block/unblock user action"""
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')  # 'block' or 'unblock'
        
        try:
            user = CustomUser.objects.get(id=user_id)
            
            # Prevent blocking yourself
            if user.id == request.user.id:
                messages.error(request, 'You cannot block/unblock your own account.')
                return redirect('block_user')
            
            if action == 'block':
                user.is_active = False
                user.save()
                messages.success(request, f'{user.get_full_name()} has been blocked successfully.')
            elif action == 'unblock':
                user.is_active = True
                user.save()
                messages.success(request, f'{user.get_full_name()} has been unblocked successfully.')
            else:
                messages.error(request, 'Invalid action.')
                
        except CustomUser.DoesNotExist:
            messages.error(request, 'User not found.')
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
            
    return redirect('block_user')


def blocked_staff_list(request):
    """View to display only blocked staff members (read-only)"""
    blocked_staff = CustomUser.objects.filter(is_active=False).order_by('-id')
    context = {
        'blocked_staff': blocked_staff,
        **get_base_context(request),
        'total_blocked': blocked_staff.count(),
    }
    return render(request, 'app/blocked_staff_list.html', context)



def loans_issued(request):
    loans = LoanApplication.objects.all().order_by('-id')
    clients = Client.objects.all()
    context = {
        'loans': loans,
        'clients': clients,
    }
    return render(request, 'app/loans_issued.html', context)

def completed_loans(request):
    """View to display all completed/fully repaid loans"""
    
    # Get all loan applications that are fully repaid
    # A loan is completed when repayment_amount_remaining <= 0
    completed_loans = LoanApplication.objects.filter(
        repayment_amount_remaining__lte=0
    ).select_related('client', 'processed_by').order_by('-updated_at')
    
    # Alternative approach: Get loans where total repayments >= total_repayment_amount
    # This is more accurate but requires annotation
    # completed_loans = LoanApplication.objects.annotate(
    #     total_repaid=Sum('repayments__repayment_amount')
    # ).filter(
    #     total_repaid__gte=F('total_repayment_amount')
    # ).select_related('client', 'processed_by').order_by('-updated_at')
    
    offices = Office.objects.all()
    
    # Calculate summary statistics
    total_completed = completed_loans.count()
    total_amount_repaid = completed_loans.aggregate(total=Sum('total_repayment_amount'))['total'] or 0
    total_interest_earned = completed_loans.aggregate(total=Sum('total_interest_amount'))['total'] or 0
    
    context = {
        'completed_loans': completed_loans,
        'offices': offices,
        'total_completed': total_completed,
        'total_amount_repaid': total_amount_repaid,
        'total_interest_earned': total_interest_earned,
    }
    return render(request, 'app/completed_loans.html', context)


def nyongeza(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    nyongeza_qs = Nyongeza.objects.all().order_by('-id')
    if filter_office:
        nyongeza_qs = nyongeza_qs.filter(Office=filter_office)

    totals     = nyongeza_qs.aggregate(
        total_bank=Sum('amount', filter=Q(deposit_method='bank')),
        total_cash=Sum('amount', filter=Q(deposit_method='cash')),
        total_all=Sum('amount'),
    )
    total_bank = totals['total_bank'] or Decimal('0.00')
    total_cash = totals['total_cash'] or Decimal('0.00')
    total_all  = totals['total_all']  or Decimal('0.00')

    dates = nyongeza_qs.dates('date', 'day', order='DESC')
    users = CustomUser.objects.filter(
        recorded_nyongezas__in=nyongeza_qs
    ).distinct()

    return render(request, 'app/nyongeza.html', {
        **base_ctx,
        'nyongeza_data': nyongeza_qs,
        'total_bank':    total_bank,
        'total_cash':    total_cash,
        'total_all':     total_all,
        'dates':         dates,
        'users':         users,
    })


def nyongeza_add(request):
    if request.method == 'POST':
        amount         = request.POST.get('amount')
        description    = request.POST.get('description')
        deposit_method = request.POST.get('deposit_method')

        try:
            amount_decimal = Decimal(amount)

            # Use selected office, not profile allocation
            branch_office = get_selected_office(request)
            if not branch_office:
                messages.error(request, 'Please select a branch office first.')
                return redirect('nyongeza')

            branch_balance = BranchBalance.objects.filter(branch=branch_office).last()

            if not branch_balance:
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=amount_decimal if deposit_method == 'cash' else Decimal('0.00'),
                    bank_balance=amount_decimal if deposit_method != 'cash' else Decimal('0.00'),
                    updated_by=request.user,
                )
            else:
                if deposit_method == 'cash':
                    BranchBalance.objects.create(
                        branch=branch_office,
                        office_balance=branch_balance.office_balance + amount_decimal,
                        bank_balance=branch_balance.bank_balance,
                        updated_by=request.user,
                    )
                else:
                    BranchBalance.objects.create(
                        branch=branch_office,
                        office_balance=branch_balance.office_balance,
                        bank_balance=branch_balance.bank_balance + amount_decimal,
                        updated_by=request.user,
                    )

            Nyongeza.objects.create(
                amount=amount_decimal,
                description=description,
                deposit_method=deposit_method,
                recorded_by=request.user,
                Office=branch_office,
            )
            messages.success(request, 'Nyongeza added successfully.')

        except Exception as e:
            messages.error(request, f'Error adding Nyongeza: {str(e)}')

    return redirect('nyongeza')


def loan_topup_add(request):
    if request.method == 'POST':
        loan_id = request.POST.get('loan_id')
        topup_amount = request.POST.get('topup_amount')
        
        try:
            loan = LoanApplication.objects.get(id=loan_id)
            topup_decimal = Decimal(topup_amount)
            
            # Update the loan amount and total repayment amount
            loan.loan_amount += topup_decimal
            loan.total_repayment_amount += topup_decimal  # Assuming interest is not added on top-up
            loan.save()
            
            messages.success(request, f'Loan {loan.id} topped up successfully by {topup_decimal}.')
            
        except LoanApplication.DoesNotExist:
            messages.error(request, 'Loan not found.')
        except Exception as e:
            messages.error(request, f'Error topping up loan: {str(e)}')
    
    return redirect('loans_issued')

import calendar

def financial_statement(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    branch = request.GET.get('branch')

    # Check if dates are specified
    dates_specified = start_date and end_date

    # Initialize variables
    opening_stock = Decimal('0.00')
    nyongeza_before = Decimal('0.00')
    total_repayments_before = Decimal('0.00')
    total_expenses_before = Decimal('0.00')
    total_loans_before = Decimal('0.00')
    salaries_before = Decimal('0.00')
    total_income_before = Decimal('0.00')
    total_expenses_before_calc = Decimal('0.00')

    # Get the selected office object if branch is specified
    selected_office = None
    if branch:
        try:
            selected_office = Office.objects.get(name=branch)
        except Office.DoesNotExist:
            selected_office = None

    # Set date range for current period display
    if dates_specified:
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

        # ============ OPENING STOCK CALCULATION (Before start_date) ============
        # Base querysets for before period
        nyongeza_before_qs = Nyongeza.objects.filter(date__lt=start_date)
        loan_repayments_before_qs = LoanRepayment.objects.filter(repayment_date__lt=start_date)
        expenses_before_qs = Expense.objects.filter(expense_date__lt=start_date)
        loan_applications_before_qs = LoanApplication.objects.filter(
            application_date__lt=start_date, status='Approved'
        )
        salaries_before_qs = Salary.objects.filter(salary_for_month__lt=start_date)

        # Apply branch filter to ALL before period querysets if branch is selected
        if selected_office:
            nyongeza_before_qs = nyongeza_before_qs.filter(Office=selected_office)
            loan_repayments_before_qs = loan_repayments_before_qs.filter(loan_application__office=selected_office)
            expenses_before_qs = expenses_before_qs.filter(office=selected_office)
            loan_applications_before_qs = loan_applications_before_qs.filter(office=selected_office)
            salaries_before_qs = salaries_before_qs.filter(fund_source=selected_office)  # FIX: Use fund_source for Salary

        # Calculate before period totals with branch filter applied
        nyongeza_before = nyongeza_before_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        total_repayments_before = loan_repayments_before_qs.aggregate(total=Sum('repayment_amount'))['total'] or Decimal('0.00')
        total_expenses_before = expenses_before_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        total_loans_before = loan_applications_before_qs.aggregate(total=Sum('loan_amount'))['total'] or Decimal('0.00')
        salaries_before = salaries_before_qs.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        total_income_before = nyongeza_before + total_repayments_before
        total_expenses_before_calc = total_expenses_before + total_loans_before + salaries_before
        opening_stock = total_income_before - total_expenses_before_calc

        # Create filters for current period
        loan_repayments_filter = Q(repayment_date__gte=start_date) & Q(repayment_date__lte=end_date)
        nyongeza_filter = Q(date__gte=start_date) & Q(date__lte=end_date)
        expenses_filter = Q(expense_date__gte=start_date) & Q(expense_date__lte=end_date)
        loan_applications_filter = (
            Q(application_date__gte=start_date) &
            Q(application_date__lte=end_date) &
            Q(status='Approved')
        )
        salaries_filter = Q(salary_for_month__gte=start_date) & Q(salary_for_month__lte=end_date)

    else:
        # When no dates specified, start from the beginning
        opening_stock = Decimal('0.00')
        
        # Find the earliest date across all models for this branch (or all branches)
        first_nyongeza = Nyongeza.objects.all()
        first_repayment = LoanRepayment.objects.all()
        first_expense = Expense.objects.all()
        first_loan = LoanApplication.objects.filter(status='Approved')
        first_salary = Salary.objects.all()
        
        # Apply branch filter to find earliest dates
        if selected_office:
            first_nyongeza = first_nyongeza.filter(Office=selected_office)
            first_repayment = first_repayment.filter(loan_application__office=selected_office)
            first_expense = first_expense.filter(office=selected_office)
            first_loan = first_loan.filter(office=selected_office)
            first_salary = first_salary.filter(fund_source=selected_office)  # FIX: Use fund_source for Salary
        
        first_nyongeza = first_nyongeza.order_by('date').first()
        first_repayment = first_repayment.order_by('repayment_date').first()
        first_expense = first_expense.order_by('expense_date').first()
        first_loan = first_loan.order_by('application_date').first()
        first_salary = first_salary.order_by('salary_for_month').first()

        all_first_dates = []
        if first_nyongeza:
            all_first_dates.append(first_nyongeza.date)
        if first_repayment:
            all_first_dates.append(first_repayment.repayment_date)
        if first_expense:
            all_first_dates.append(first_expense.expense_date)
        if first_loan:
            all_first_dates.append(first_loan.application_date)
        if first_salary:
            all_first_dates.append(first_salary.salary_for_month)

        start_date = min(all_first_dates) if all_first_dates else timezone.now().date()
        end_date = timezone.now().date()

        # Initialize empty filters for current period
        loan_repayments_filter = Q()
        nyongeza_filter = Q()
        expenses_filter = Q()
        loan_applications_filter = Q(status='Approved')
        salaries_filter = Q()

    # Apply branch filter to current period filters
    if selected_office:
        loan_repayments_filter &= Q(loan_application__office=selected_office)
        nyongeza_filter &= Q(Office=selected_office)
        expenses_filter &= Q(office=selected_office)
        loan_applications_filter &= Q(office=selected_office)
        salaries_filter &= Q(fund_source=selected_office)  # FIX: Use fund_source for Salary

    # ============ CURRENT PERIOD CALCULATIONS ============
    # Get current period data with all filters applied
    loan_repayments = LoanRepayment.objects.filter(loan_repayments_filter)
    total_mapato = loan_repayments.aggregate(total=Sum('repayment_amount'))['total'] or Decimal('0.00')

    nyongeza = Nyongeza.objects.filter(nyongeza_filter)
    total_nyongeza = nyongeza.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_income = total_mapato + total_nyongeza

    expenses = Expense.objects.filter(expenses_filter)
    total_expenses_current = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    loan_applications = LoanApplication.objects.filter(loan_applications_filter)
    total_loans_current = loan_applications.aggregate(total=Sum('loan_amount'))['total'] or Decimal('0.00')

    salaries = Salary.objects.filter(salaries_filter)  # Now properly filtered by fund_source
    total_salaries_current = salaries.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_outflow_current = total_expenses_current + total_loans_current + total_salaries_current

    # Calculate closing balance - ALL COMPONENTS ARE NOW BRANCH-SPECIFIC
    closing_balance = opening_stock + total_income - total_outflow_current

    # ============ VERIFICATION: Calculate expected closing balance from actual cash ============
    # This should match closing_balance if everything is calculated correctly
    if selected_office:
        # Get the latest branch balance for verification
        latest_balance = BranchBalance.objects.filter(
            branch=selected_office
        ).order_by('-last_updated').first()
        
        if latest_balance:
            expected_cash = latest_balance.bank_balance + latest_balance.office_balance
            
            # You can add this to context for debugging if needed
            # if closing_balance != expected_cash:
            #     print(f"Warning: Calculated closing balance ({closing_balance}) doesn't match actual cash ({expected_cash})")

    # ============ CASH IN BANK & CASH IN OFFICE ============
    if selected_office:
        # Get the single latest record for the selected branch
        branch_balance_qs = BranchBalance.objects.filter(
            branch=selected_office
        ).order_by('-last_updated')

        latest_branch_balance = branch_balance_qs.first()

        if latest_branch_balance:
            cash_in_bank = latest_branch_balance.bank_balance
            cash_in_office = latest_branch_balance.office_balance
        else:
            cash_in_bank = Decimal('0.00')
            cash_in_office = Decimal('0.00')

        branch_balances_detail = []
        if latest_branch_balance:
            branch_balances_detail.append({
                'branch_name': latest_branch_balance.branch.name,
                'office_balance': latest_branch_balance.office_balance,
                'bank_balance': latest_branch_balance.bank_balance,
                'last_updated': latest_branch_balance.last_updated,
            })
    else:
        # Aggregate across all branches
        all_branches = Office.objects.all()
        cash_in_bank = Decimal('0.00')
        cash_in_office = Decimal('0.00')
        branch_balances_detail = []

        for office in all_branches:
            latest = BranchBalance.objects.filter(
                branch=office
            ).order_by('-last_updated').first()

            if latest:
                cash_in_bank += latest.bank_balance
                cash_in_office += latest.office_balance
                branch_balances_detail.append({
                    'branch_name': office.name,
                    'office_balance': latest.office_balance,
                    'bank_balance': latest.bank_balance,
                    'last_updated': latest.last_updated,
                })

    total_cash = cash_in_bank + cash_in_office

    # Group expenses by category (only for filtered branch)
    expenses_by_category = []
    categories = ExpenseCategory.objects.all()
    for category in categories:
        category_expenses = expenses.filter(transaction_type=category)
        category_total = category_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        if category_total > 0:
            expenses_by_category.append({
                'category': category.name,
                'amount': category_total,
                'count': category_expenses.count()
            })

    offices = Office.objects.all()
    month_options = generate_month_options()

    context = {
        # Filter info
        'start_date': start_date,
        'end_date': end_date,
        'dates_specified': dates_specified,
        'selected_branch': branch,
        'showing_all_data': not dates_specified,

        # Opening Stock (BRANCH-SPECIFIC)
        'opening_stock': opening_stock,
        'opening_stock_explanation': (
            'Starting from zero (viewing all transactions)'
            if not dates_specified
            else f'Net position for {branch if branch else "all branches"} before {start_date.strftime("%B %d, %Y")}'
        ),

        # Before period breakdown (BRANCH-SPECIFIC)
        'nyongeza_before': nyongeza_before,
        'repayments_before': total_repayments_before,
        'expenses_before': total_expenses_before,
        'loans_before': total_loans_before,
        'salaries_before': salaries_before,
        'total_income_before': total_income_before,
        'total_expenses_before': total_expenses_before_calc,

        # Current period income (BRANCH-SPECIFIC)
        'total_mapato': total_mapato,
        'mapato_count': loan_repayments.count(),
        'mapato_list': loan_repayments.order_by('-repayment_date')[:10],

        'total_nyongeza': total_nyongeza,
        'nyongeza_count': nyongeza.count(),
        'nyongeza_list': nyongeza.order_by('-date')[:10],

        'total_income': total_income,

        # Current period expenses (BRANCH-SPECIFIC)
        'total_expenses_current': total_expenses_current,
        'total_loans_current': total_loans_current,
        'total_salaries_current': total_salaries_current,  # Now branch-specific
        'total_outflow_current': total_outflow_current,
        'expense_count': expenses.count(),
        'loan_count': loan_applications.count(),
        'salary_count': salaries.count(),  # Now branch-specific

        'expenses_by_category': expenses_by_category,
        'recent_expenses': expenses.order_by('-expense_date')[:10],
        'recent_loans': loan_applications.order_by('-application_date')[:10],
        'recent_salaries': salaries.order_by('-salary_for_month')[:10],  # Now branch-specific

        # Closing Balance (CONSISTENT - ALL BRANCH-SPECIFIC)
        'closing_balance': closing_balance,

        # Cash in Bank & Cash in Office
        'cash_in_bank': cash_in_bank,
        'cash_in_office': cash_in_office,
        'total_cash': total_cash,
        'branch_balances_detail': branch_balances_detail,

        # Filter options
        'offices': offices,
        'month_options': month_options,

        # Helper
        'current_date': timezone.now().date(),
        'total_income_display': total_income,
        'total_outflow_display': total_outflow_current,
        
        # Add selected office for template use
        'selected_office': selected_office,
    }

    return render(request, 'app/financial_statement.html', context)


def generate_month_options():
    """Generate month options for the last 12 months"""
    month_options = []
    for i in range(11, -1, -1):
        date = timezone.now().date() - timedelta(days=30*i)
        month_options.append({
            'value': date.strftime('%Y-%m'),
            'label': date.strftime('%B %Y')
        })
    return month_options



def hq_financial_statement(request):
    """
    HQ Financial Statement View — shows all branches as columns.
    Rows: Opening Balance, MAPATO, NYONGEZA, HAZINA (subtotal),
          FOMU, MATUMIZI OFISINI, MATUMIZI-KITUO, MATUMIZI-MKURUGENZI,
          MAKATO BENKI (subtotal), BALANCE OFISINI, BALANCE BENKI, Closing Balance.
    """

    # ── Filter parameters ────────────────────────────────────────────────────
    start_date_str = request.GET.get('start_date')
    end_date_str   = request.GET.get('end_date')
    dates_specified = bool(start_date_str and end_date_str)

    if dates_specified:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(end_date_str,   '%Y-%m-%d').date()
    else:
        end_date   = timezone.now().date()
        start_date = end_date.replace(day=1)          # default: current month

    # ── Helpers ──────────────────────────────────────────────────────────────
    def _sum(qs, field='amount'):
        return qs.aggregate(total=Sum(field))['total'] or Decimal('0.00')

    def _office_filter(office_name):
        """Return Q objects scoped to a branch name (or '' for HQ/unassigned)."""
        return office_name  # we pass the office name string directly

    # ── All branches ─────────────────────────────────────────────────────────
    offices = list(Office.objects.all().order_by('name'))
    office_names = [o.name for o in offices]

    # ── Per-branch calculations ───────────────────────────────────────────────
    branch_data = {}   # { office_name: { row_key: Decimal } }

    for office in offices:
        oname = office.name
        d = {}

        # ---------- Opening Balance (everything BEFORE start_date) ----------
        rep_before = _sum(
            LoanRepayment.objects.filter(
                repayment_date__lt=start_date,
                loan_application__office=oname
            ),
            'repayment_amount'
        )
        nyong_before = _sum(
            Nyongeza.objects.filter(date__lt=start_date),
            'amount'
        )
        exp_before = _sum(
            Expense.objects.filter(expense_date__lt=start_date, office=oname)
        )
        loans_before = _sum(
            LoanApplication.objects.filter(
                application_date__lt=start_date,
                status='Approved',
                office=oname
            ),
            'loan_amount'
        )
        sal_before = _sum(
            Salary.objects.filter(salary_for_month__lt=start_date)
        )
        # Bank cash withdrawals/deposits before period (MAKATO BENKI equivalent)
        bank_before = _sum(
            BankCashTransaction.objects.filter(
                office_from=office,
                transaction_date__lt=start_date
            )
        )

        income_before   = rep_before + nyong_before
        expense_before  = exp_before + loans_before + sal_before + bank_before
        d['opening_balance'] = income_before - expense_before

        # ---------- Current period filters ----------
        if dates_specified:
            date_q_rep   = Q(repayment_date__gte=start_date, repayment_date__lte=end_date)
            date_q_nyong = Q(date__gte=start_date, date__lte=end_date)
            date_q_exp   = Q(expense_date__gte=start_date, expense_date__lte=end_date)
            date_q_loan  = Q(application_date__gte=start_date, application_date__lte=end_date)
            date_q_sal   = Q(salary_for_month__gte=start_date, salary_for_month__lte=end_date)
            date_q_bank  = Q(transaction_date__gte=start_date, transaction_date__lte=end_date)
        else:
            date_q_rep   = Q()
            date_q_nyong = Q()
            date_q_exp   = Q()
            date_q_loan  = Q()
            date_q_sal   = Q()
            date_q_bank  = Q()

        # ---------- MAPATO (loan repayments) ----------
        d['mapato'] = _sum(
            LoanRepayment.objects.filter(date_q_rep, loan_application__office=oname),
            'repayment_amount'
        )

        # ---------- NYONGEZA ----------
        d['nyongeza'] = _sum(Nyongeza.objects.filter(date_q_nyong))

        # ---------- HAZINA (Opening + Mapato + Nyongeza) ----------
        d['hazina'] = d['opening_balance'] + d['mapato'] + d['nyongeza']

        # ---------- FOMU (loan disbursements) ----------
        d['fomu'] = _sum(
            LoanApplication.objects.filter(
                date_q_loan, status='Approved', office=oname
            ),
            'loan_amount'
        )

        # ---------- MATUMIZI OFISINI (office running expenses) ----------
        d['matumizi_ofisini'] = _sum(
            Expense.objects.filter(
                date_q_exp, office=oname,
                transaction_type__name__icontains='ofis'
            )
        )

        # ---------- MATUMIZI-KITUO (branch/centre expenses) ----------
        d['matumizi_kituo'] = _sum(
            Expense.objects.filter(
                date_q_exp, office=oname,
                transaction_type__name__icontains='kituo'
            )
        )

        # ---------- MATUMIZI-MKURUGENZI (HQ/director expenses) ----------
        d['matumizi_mkurugenzi'] = _sum(
            Expense.objects.filter(
                date_q_exp, office=oname,
                transaction_type__name__icontains='mkurugenzi'
            )
        )

        # ---------- MAKATO BENKI (bank deductions / cash-bank transfers) ----------
        d['makato_benki'] = _sum(
            BankCashTransaction.objects.filter(date_q_bank, office_from=office)
        )

        # ---------- Total Outflow ----------
        total_outflow = (
            d['fomu'] + d['matumizi_ofisini'] +
            d['matumizi_kituo'] + d['matumizi_mkurugenzi'] +
            d['makato_benki']
        )

        # ---------- BALANCE OFISINI (cash at office) ----------
        #   Hazina minus loans disbursed and non-bank expenses
        d['balance_ofisini'] = (
            d['hazina']
            - d['fomu']
            - d['matumizi_ofisini']
            - d['matumizi_kituo']
            - d['matumizi_mkurugenzi']
        )

        # ---------- BALANCE BENKI (bank balance) ----------
        d['balance_benki'] = d['hazina'] - d['makato_benki'] - d['fomu']

        # ---------- Closing Balance ----------
        d['closing_balance'] = d['hazina'] - total_outflow

        branch_data[oname] = d

    # ── Totals column ────────────────────────────────────────────────────────
    row_keys = [
        'opening_balance', 'mapato', 'nyongeza', 'hazina',
        'fomu', 'matumizi_ofisini', 'matumizi_kituo', 'matumizi_mkurugenzi',
        'makato_benki', 'balance_ofisini', 'balance_benki', 'closing_balance',
    ]
    totals = {}
    for key in row_keys:
        totals[key] = sum(branch_data[o][key] for o in office_names)

    # ── Month options ────────────────────────────────────────────────────────
    month_options = []
    for i in range(11, -1, -1):
        d_ = timezone.now().date() - timedelta(days=30 * i)
        month_options.append({
            'value': d_.strftime('%Y-%m'),
            'label': d_.strftime('%B %Y'),
        })

    context = {
        'start_date':      start_date,
        'end_date':        end_date,
        'dates_specified': dates_specified,
        'offices':         offices,
        'office_names':    office_names,
        'branch_data':     branch_data,
        'totals':          totals,
        'month_options':   month_options,
        'current_date':    timezone.now().date(),
    }

    return render(request, 'app/hq_financial_statement.html', context)

def fomu_mkopo_wa_dharula(request):
    return render(request, 'app/fomu_mkopo_wa_dharula.html')

def fomu_mkopo_wa_dharula(request, loan_id=None):
    """
    View for loan application form - can be used for both new and existing loans
    If loan_id is provided, it loads existing loan data
    """
    loan = None
    client = None
    
    # If loan_id is provided, get that specific loan
    if loan_id:
        loan = get_object_or_404(LoanApplication, id=loan_id)
        client = loan.client
    
    context = {
        'loan': loan,
        'client': client,
        'current_user': request.user,
        'today': timezone.now().date(),
    }
    return render(request, 'app/fomu_mkopo_wa_dharula.html', context)

def bank_transfer_expenses2(request):
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    office_filter = request.GET.get('office')

    transactions = HQTransaction.objects.select_related(
        'from_branch', 'to_branch', 'processed_by'
    ).order_by('transaction_date', 'id')

    if date_from:
        try:
            transactions = transactions.filter(
                transaction_date__gte=datetime.datetime.strptime(date_from, "%Y-%m-%d").date()
            )
        except ValueError:
            pass

    if date_to:
        try:
            transactions = transactions.filter(
                transaction_date__lte=datetime.datetime.strptime(date_to, "%Y-%m-%d").date()
            )
        except ValueError:
            pass

    if office_filter:
        transactions = transactions.filter(
            models.Q(from_branch__name=office_filter) | models.Q(to_branch__name=office_filter)
        )

    # Offices from Office model via HQTransaction branches
    offices = Office.objects.filter(
        models.Q(hq_transactions_from__isnull=False) | models.Q(hq_transactions_to__isnull=False)
    ).values_list('name', flat=True).distinct().order_by('name')

    grand_total = transactions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    transactions_with_receipt = [
        {'transaction': txn, 'receipt_number': str(txn.id).zfill(6)}
        for txn in transactions
    ]

    context = {
        **get_base_context(request),
        'transactions_with_receipt': transactions_with_receipt,
        'grand_total': grand_total,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'offices': offices,
        'office_filter': office_filter or '',
        'total_count': len(transactions_with_receipt),
    }

    return render(request, 'app/bank_transfer_expenses.html', context)






def office_transaction_page(request):
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/office_transaction_page.html', context)



def bank_charges(request):
    base_ctx        = get_base_context(request)
    selected_office = base_ctx['selected_office']
    context = {
        **get_base_context(request),
        'selected_office': selected_office,
    }
    return render(request, 'app/bank_charges.html', context)

def bank_charge_add(request):
    if request.method == 'POST':
        try:
            description = request.POST.get('description')
            amount = request.POST.get('amount')
            payment_method = request.POST.get('payment_method')
            transaction_date = request.POST.get('transaction_date')
            attachment = request.FILES.get('attachment')

            # Convert amount safely
            amount = Decimal(amount) if amount else Decimal('0.00')

            # Convert transaction date safely
            if transaction_date:
                transaction_date = timezone.datetime.strptime(
                    transaction_date, "%Y-%m-%d"
                ).date()
            else:
                transaction_date = None

            # Get branch office from logged-in user
            branch_office = get_selected_office(request)
            if not branch_office:
                messages.error(request, 'You are not allocated to any office/branch.')
                return redirect('bank_charges')

            # Get latest branch balance
            branch_balance = BranchBalance.objects.filter(branch=branch_office).last()
            if not branch_balance:
                messages.error(request, f'No balance record found for {branch_office.name}.')
                return redirect('bank_charges')

            # Deduct based on payment method
            if payment_method == 'cash':
                if branch_balance.office_balance < amount:
                    messages.error(
                        request,
                        f'Insufficient cash balance. Available: {branch_balance.office_balance}'
                    )
                    return redirect('bank_charges')

                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance - amount,
                    bank_balance=branch_balance.bank_balance,
                    updated_by=request.user,
                )
            else:
                if branch_balance.bank_balance < amount:
                    messages.error(
                        request,
                        f'Insufficient bank balance. Available: {branch_balance.bank_balance}'
                    )
                    return redirect('bank_charges')

                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance,
                    bank_balance=branch_balance.bank_balance - amount,
                    updated_by=request.user,
                )

            # ✅ Create BankCharge record (MATCHES MODEL)
            BankCharge.objects.create(
                description=description,
                amount=amount,
                recorded_by=request.user,
                office=branch_office.name,
                payment_method=payment_method,
                transaction_date=transaction_date,
                attachment=attachment,
            )

            messages.success(request, "Bank charge recorded successfully.")
            return redirect('bank_charges')

        except Exception as e:
            messages.error(request, f"Error adding bank charge: {str(e)}")

    return render(request, 'app/bank_charges.html')


def process_loan(request, pk):
    """
    PART A — Show editable client details.
    GET  : render the form pre-filled with client data.
    (Edits are saved when the user clicks NEXT, via process_loan_partb POST.)
    """
    client = get_object_or_404(Client, pk=pk)

    # Calculate age from date_of_birth for the Age dropdown
    client_age = None
    if client.date_of_birth:
        today = datetime.date.today()
        client_age = today.year - client.date_of_birth.year - (
            (today.month, today.day) < (client.date_of_birth.month, client.date_of_birth.day)
        )

    age_choices = list(range(18, 76))  # 18–75

    return render(request, 'app/process_loan_parta.html', {
        'client':      client,
        'client_age':  client_age,
        'age_choices': age_choices,
        **get_base_context(request),
    })
    


def process_loan_partb(request, pk):
    """
    Called by Part A form (POST).
    1. Saves any edits made to client fields in Part A.
    2. Looks up existing active loans by type (Maendeleo / Dharura) with
       an outstanding balance > 0, and passes them to the template
       separately so the UI can enforce the correct rules:
         - Maendeleo loan exists  → selecting Maendeleo forces a top-up
         - Dharura loan exists    → selecting Dharura is blocked entirely
         - Maendeleo loan exists  → selecting Dharura is allowed as a new loan
         - No balance (paid off)  → treated as no active loan for that type
    """
    client = get_object_or_404(Client, pk=pk)

    if request.method == 'POST':
        # ── Save Part A edits to client ──────────────────────────────
        client.firstname           = request.POST.get('firstname',           client.firstname).strip()
        client.middlename          = request.POST.get('middlename',          client.middlename  or '').strip()
        client.lastname            = request.POST.get('lastname',            client.lastname).strip()
        client.phonenumber         = request.POST.get('phonenumber',         client.phonenumber or '').strip()
        client.employername        = request.POST.get('employername',        client.employername or '').strip()
        client.mtaa                = request.POST.get('mtaa',                client.mtaa        or '').strip()
        client.kata                = request.POST.get('kata',                client.kata        or '').strip()
        client.wilaya              = request.POST.get('wilaya',              client.wilaya      or '').strip()
        client.idara               = request.POST.get('idara',               client.idara       or '').strip()
        client.checkno             = request.POST.get('checkno',             client.checkno     or '').strip()
        client.bank_account_number = request.POST.get('bank_account_number', client.bank_account_number or '').strip()
        client.bank_name           = request.POST.get('bank_name',           client.bank_name   or '').strip()

        if hasattr(client, 'gender'):
            client.gender = request.POST.get('gender', getattr(client, 'gender', '') or '').strip()

        client.save()

    # ── Helper: fetch active loan of a given type with balance > 0 ───
    def get_active_loan(loan_type):
        loan = LoanApplication.objects.filter(
            client=client,
            status__in=['Approved', 'Pending'],
            loan_type=loan_type,
        ).order_by('-created_at').first()

        # Treat as "no active loan" if the balance is zero or missing
        if loan and (loan.repayment_amount_remaining or 0) > 0:
            return loan
        return None

    existing_maendeleo = get_active_loan('Maendeleo')
    existing_dharura   = get_active_loan('Dharura')

    return render(request, 'app/process_loan_partb.html', {
        'client':             client,
        'existing_maendeleo': existing_maendeleo,
        'existing_dharura':   existing_dharura,
        **get_base_context(request),
    })


def loan_application(request):
    if request.method == 'POST':
        client_id             = request.POST.get('client')
        loan_amount           = request.POST.get('amount', '').replace(',', '').strip()
        loan_purpose          = request.POST.get('purpose', '')
        loan_type             = request.POST.get('loan_type')
        interest_rate         = request.POST.get('interest_rate')
        payment_period_months = request.POST.get('term_months')
        application_date_str  = request.POST.get('application_date')  # e.g. "2026-03-08"
        transaction_method    = request.POST.get('transaction_method', 'cash')
        processed_by          = request.user
        topup_date            = request.POST.get('topup_date', '')

        # ── Validate required fields ──────────────────────────────────
        missing = []
        if not client_id:             missing.append('client')
        if not loan_amount:           missing.append('amount')
        if not loan_type:             missing.append('loan type')
        if not interest_rate:         missing.append('interest rate')
        if not payment_period_months: missing.append('period (term months)')
        if not application_date_str:  missing.append('application date')

        if missing:
            messages.error(
                request,
                f"Missing required fields: {', '.join(missing)}. Please fill in all fields."
            )
            if client_id:
                return redirect('process_loan_partb', pk=int(client_id))
            return redirect('client_list')

        try:
            loan_amount_decimal = Decimal(loan_amount)

            # ── Convert date string → date object ─────────────────────
            # model.save() calls self.application_date.day — needs a real date object
            application_date = datetime.datetime.strptime(application_date_str, '%Y-%m-%d').date()

            client = get_object_or_404(Client, id=client_id)

            # ── Check for existing pending loan of same type ──────────
            existing_loan = LoanApplication.objects.filter(
                client=client,
                loan_type=loan_type,
                status='Pending'
            ).first()
            if existing_loan:
                messages.error(
                    request,
                    f"Client {client} already has a pending {loan_type} loan. "
                    "Cannot apply for a new one until the existing one is processed."
                )
                return redirect('process_loan_partb', pk=client.id)

            # ── Dharura-specific rules ────────────────────────────────
            if loan_type == 'Dharura':

                # 1. Maximum loan amount cap
                DHARURA_MAX = Decimal('200000')
                if loan_amount_decimal > DHARURA_MAX:
                    messages.error(
                        request,
                        f'Dharura loan cannot exceed TZS {DHARURA_MAX:,.0f}/=. '
                        f'Requested amount: TZS {loan_amount_decimal:,.0f}/='
                    )
                    return redirect('process_loan_partb', pk=client.id)

                # 2. Branch-wide active Dharura loan cap (max 10)
                # We need branch_office here, so fetch it early for this check
                branch_office_check = get_selected_office(request)
                if branch_office_check:
                    DHARURA_BRANCH_LIMIT = 10
                    active_dharura_count = LoanApplication.objects.filter(
                        office=branch_office_check.name,
                        loan_type='Dharura',
                        status__in=['Approved', 'Pending'],
                        repayment_amount_remaining__gt=0,
                    ).count()

                    if active_dharura_count >= DHARURA_BRANCH_LIMIT:
                        messages.error(
                            request,
                            f'Branch "{branch_office_check.name}" has reached the maximum of '
                            f'{DHARURA_BRANCH_LIMIT} active Dharura loans. '
                            'A new Dharura loan cannot be issued until an existing one is fully repaid.'
                        )
                        return redirect('process_loan_partb', pk=client.id)

            # ── Branch office & balance ───────────────────────────────
            branch_office = get_selected_office(request)
            if not branch_office:
                messages.error(request, 'No branch office found for your account.')
                return redirect('process_loan_partb', pk=client.id)

            branch_balance = BranchBalance.objects.filter(branch=branch_office).last()
            if not branch_balance:
                messages.error(request, f'No balance record found for {branch_office.name}.')
                return redirect('process_loan_partb', pk=client.id)

            # ── Deduct balance based on transaction method ────────────
            if transaction_method == 'cash':
                if branch_balance.office_balance < loan_amount_decimal:
                    messages.error(
                        request,
                        f'Insufficient cash balance. Available: {branch_balance.office_balance:,.0f}/='
                    )
                    return redirect('process_loan_partb', pk=client.id)
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance - loan_amount_decimal,
                    bank_balance=branch_balance.bank_balance,
                    updated_by=processed_by,
                )
            else:
                if branch_balance.bank_balance < loan_amount_decimal:
                    messages.error(
                        request,
                        f'Insufficient bank balance. Available: {branch_balance.bank_balance:,.0f}/='
                    )
                    return redirect('process_loan_partb', pk=client.id)
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=branch_balance.office_balance,
                    bank_balance=branch_balance.bank_balance - loan_amount_decimal,
                    updated_by=processed_by,
                )

            # ── Create loan application ───────────────────────────────
            new_loan = LoanApplication.objects.create(
                client=client,
                loan_amount=loan_amount_decimal,
                loan_purpose=loan_purpose,
                loan_type=loan_type,
                interest_rate=Decimal(interest_rate),
                payment_period_months=int(payment_period_months),
                application_date=application_date,
                processed_by=processed_by,
                office=branch_office.name,
                transaction_method=transaction_method,
            )

            messages.success(
                request,
                f'Loan application for {client} processed successfully.'
            )
            return redirect('clients')
            # return redirect('loan_receipt', loan_id=new_loan.id)

        except Exception as e:
            messages.error(request, f'Error processing loan application: {str(e)}')
            if client_id:
                return redirect('process_loan_partb', pk=int(client_id))

    return redirect('client_list')



def loan_payment_page(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    active_loans = LoanApplication.objects.select_related('client').filter(
        repayment_amount_remaining__gt=0,
        status='Approved'
    ).exclude(
        status='Paid'
    ).order_by('client__firstname', 'client__lastname')

    if filter_office:
        active_loans = active_loans.filter(office=filter_office.name)

    # Group loans by client — one entry per client, list their loans
    from collections import OrderedDict
    clients_map = OrderedDict()
    for loan in active_loans:
        cid = loan.client.id
        if cid not in clients_map:
            clients_map[cid] = {
                'client': loan.client,
                'loans': [],
            }
        clients_map[cid]['loans'].append(loan)

    # Build a flat list for the template: each entry = one client
    # primary_loan_id = first loan id (used if client has only 1 loan)
    client_loan_list = []
    for cid, data in clients_map.items():
        client_loan_list.append({
            'client':        data['client'],
            'loans':         data['loans'],
            'primary_loan_id': data['loans'][0].id,
            'has_multiple':  len(data['loans']) > 1,
        })

    return render(request, 'app/loan_payment_page.html', {
        **base_ctx,
        'client_loan_list': client_loan_list,
    })



def loan_payment_select(request):
    """
    Handles the POST from Page 1 (customer selection).
    Redirects to the repayment form for the selected loan.
    """
    if request.method == 'POST':
        loan_id = request.POST.get('loan_id')
        if not loan_id:
            messages.error(request, 'Please select a customer before proceeding.')
            return redirect('loan_payment_page')
        return redirect('loan_repayment', loan_id=int(loan_id))

    return redirect('loan_payment_page')



def loan_repayment(request, loan_id):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    loan       = get_object_or_404(LoanApplication, id=loan_id)
    repayments = LoanRepayment.objects.filter(
        loan_application=loan
    ).order_by('-id')

    total_paid = sum(
        (r.repayment_amount for r in repayments),
        Decimal('0.00')
    )

    office = None
    if loan.office:
        try:
            office = Office.objects.get(name=loan.office)
        except Office.DoesNotExist:
            pass

    # ── Build repayment month options ──────────────────────────────────
    repayment_months = []
    today       = datetime.date.today()
    start_date  = loan.first_repayment_date or loan.application_date
    current     = datetime.date(start_date.year, start_date.month, 1)
    end         = datetime.date(today.year, today.month, 1)
    month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    while current <= end:
        repayment_months.append({
            'value': current.strftime('%Y-%m'),
            'label': f"{month_names[current.month]}/{current.year}",
        })
        current = (
            datetime.date(current.year + 1, 1, 1)
            if current.month == 12
            else datetime.date(current.year, current.month + 1, 1)
        )
    repayment_months.reverse()

    if request.method == 'POST':
        repayment_amount_raw  = request.POST.get('repayment_amount', '').strip()
        transaction_method    = request.POST.get('transaction_method', 'cash')
        repayment_month_str   = request.POST.get('repayment_month', '').strip()
        # added below
        repayment_date = request.POST.get('transaction_date', '')
        # 
        processed_by          = request.user

        try:
            repayment_amount = Decimal(repayment_amount_raw)

            # Prevent overpayment — cap at remaining balance
            if repayment_amount > loan.repayment_amount_remaining:
                repayment_amount = loan.repayment_amount_remaining

            # Parse payment_month from "YYYY-MM" → date (first day of that month)
            payment_month = None
            if repayment_month_str:
                try:
                    payment_month = datetime.datetime.strptime(repayment_month_str, '%Y-%m').date()
                except ValueError:
                    payment_month = None

            # ── Branch office check ────────────────────────────────────
            branch_office = get_selected_office(request)
            if not branch_office:
                messages.error(request, 'Please select a branch office first.')
                return redirect('loan_repayment', loan_id=loan_id)

            # ── Update branch balance ──────────────────────────────────
            branch_balance = BranchBalance.objects.filter(branch=branch_office).last()

            if not branch_balance:
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=repayment_amount if transaction_method == 'cash' else Decimal('0.00'),
                    bank_balance=repayment_amount if transaction_method != 'cash' else Decimal('0.00'),
                    updated_by=processed_by,
                )
            else:
                BranchBalance.objects.create(
                    branch=branch_office,
                    office_balance=(
                        branch_balance.office_balance + repayment_amount
                        if transaction_method == 'cash'
                        else branch_balance.office_balance
                    ),
                    bank_balance=(
                        branch_balance.bank_balance + repayment_amount
                        if transaction_method != 'cash'
                        else branch_balance.bank_balance
                    ),
                    updated_by=processed_by,
                )

            # ── Create repayment record ────────────────────────────────
            # LoanRepayment.save() auto-decrements repayment_amount_remaining
            LoanRepayment.objects.create(
                loan_application=loan,
                repayment_amount=repayment_amount,
                transaction_method=transaction_method,
                payment_month=payment_month,
                # 
                repayment_date = repayment_date,
                # 
                processed_by=processed_by,
            )

            # ── Re-fetch loan to get updated balance from model.save() ─
            loan.refresh_from_db()
            if loan.repayment_amount_remaining <= 0:
                loan.status = 'Paid'
                loan.save(update_fields=['status'])

            messages.success(
                request,
                f'Repayment of TZS {repayment_amount:,.0f} processed successfully.'
            )
            return redirect('loan_payment_page')

        except Exception as e:
            messages.error(request, f'Error processing repayment: {str(e)}')

    context = {
        **base_ctx,
        'loan':             loan,
        'repayments':       repayments,
        'total_paid':       total_paid,
        'office':           office,
        'repayment_months': repayment_months,
    }
    return render(request, 'app/loan_repayment.html', context)




def bank_cash_transaction_add(request):
    """
    Handles: Transfer between Cash and Bank for the selected office.
    Both cash and bank balances live on the same BranchBalance record,
    so a single new BranchBalance entry with updated figures is all that's needed.
    """
    processed_by   = request.user
    office_from    = get_selected_office(request)
    branch_balance = BranchBalance.objects.filter(branch=office_from).last() if office_from else None

    if request.method == 'POST':
        try:
            source               = request.POST.get('source')
            destination          = request.POST.get('destination')
            amount               = Decimal(request.POST.get('amount'))
            transaction_date_str = request.POST.get('transaction_date', '')
            attachment           = request.FILES.get('attachment')   # ← new

            try:
                transaction_date = datetime.datetime.strptime(transaction_date_str, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Invalid transaction date.')
                return redirect('bank_cash_transaction')

            if not office_from:
                messages.error(request, 'Please select a branch office first.')
                return redirect('bank_cash_transaction')

            with transaction.atomic():
                branch_balance = BranchBalance.objects.select_for_update().filter(branch=office_from).last()
                if not branch_balance:
                    messages.error(request, f'No balance record found for {office_from.name}.')
                    return redirect('bank_cash_transaction')

                new_office_balance = branch_balance.office_balance
                new_bank_balance   = branch_balance.bank_balance

                if source == 'cash' and destination == 'bank':
                    if branch_balance.office_balance < amount:
                        messages.error(request, f'Insufficient cash. Available: TZS {branch_balance.office_balance:,.0f}')
                        return redirect('bank_cash_transaction')
                    new_office_balance -= amount
                    new_bank_balance   += amount
                    description = f'Cash to Bank transfer for {office_from.name}'

                elif source == 'bank' and destination == 'cash':
                    if branch_balance.bank_balance < amount:
                        messages.error(request, f'Insufficient bank balance. Available: TZS {branch_balance.bank_balance:,.0f}')
                        return redirect('bank_cash_transaction')
                    new_bank_balance   -= amount
                    new_office_balance += amount
                    description = f'Bank to Cash transfer for {office_from.name}'

                else:
                    messages.error(request, 'Invalid source/destination combination.')
                    return redirect('bank_cash_transaction')

                BranchBalance.objects.create(
                    branch=office_from,
                    office_balance=new_office_balance,
                    bank_balance=new_bank_balance,
                    updated_by=processed_by,
                )

                # ── Save BankCashTransaction with optional attachment ──────
                bank_cash_txn = BankCashTransaction.objects.create(
                    office_from=office_from,
                    source=source,
                    destination=destination,
                    amount=amount,
                )
                if attachment:
                    bank_cash_txn.attachment = attachment
                    bank_cash_txn.save(update_fields=['attachment'])

                # ── Save HQTransaction with same attachment ────────────────
                hq_txn = HQTransaction.objects.create(
                    from_branch=office_from,
                    to_branch=office_from,
                    amount=amount,
                    description=description,
                    transaction_date=transaction_date,
                    processed_by=processed_by,
                )
                if attachment:
                    # Re-open is not possible after first save consumed the stream,
                    # so we link the same file path from bank_cash_txn
                    hq_txn.attachment = bank_cash_txn.attachment
                    hq_txn.save(update_fields=['attachment'])

            messages.success(
                request,
                f'Transfer of TZS {amount:,.0f} from {source.title()} to {destination.title()} '
                f'completed for {office_from.name}.'
            )

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

        return redirect('bank_cash_transaction')

    return render(request, 'app/bank_cash_transaction_add.html', {
        **get_base_context(request),
        'office_from':    office_from,
        'branch_balance': branch_balance,
    })


def branches_transaction_add(request):
    """
    Handles: Transfer to another Branch.
    Deducts from sending office, adds to receiving office.
    office_from is always the currently selected office.
    """
    processed_by   = request.user
    office_from    = get_selected_office(request)
    branch_balance = BranchBalance.objects.filter(branch=office_from).last() if office_from else None

    if request.method == 'POST':
        office_to_id         = request.POST.get('office_to')
        transaction_type     = request.POST.get('transaction_type', 'branch')
        transaction_method   = request.POST.get('transaction_method', 'cash')
        amount               = Decimal(request.POST.get('amount'))
        transaction_date_str = request.POST.get('transaction_date', '')
        attachment           = request.FILES.get('attachment')   # ← new

        try:
            transaction_date = datetime.datetime.strptime(transaction_date_str, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Invalid transaction date.')
            return redirect('bank_cash_transaction')

        if not office_from:
            messages.error(request, 'Please select a branch office first.')
            return redirect('bank_cash_transaction')

        office_to = get_object_or_404(Office, id=office_to_id)

        if office_from == office_to:
            messages.error(request, 'Cannot transfer to the same office.')
            return redirect('bank_cash_transaction')

        with transaction.atomic():
            # ── Validate sender balance ───────────────────────────────────
            branch_balance_from = BranchBalance.objects.select_for_update().filter(branch=office_from).last()
            if not branch_balance_from:
                messages.error(request, f'No balance record found for {office_from.name}.')
                return redirect('bank_cash_transaction')

            if transaction_method == 'cash':
                if branch_balance_from.office_balance < amount:
                    messages.error(request, f'Insufficient cash. Available: TZS {branch_balance_from.office_balance:,.0f}')
                    return redirect('bank_cash_transaction')
            else:
                if branch_balance_from.bank_balance < amount:
                    messages.error(request, f'Insufficient bank balance. Available: TZS {branch_balance_from.bank_balance:,.0f}')
                    return redirect('bank_cash_transaction')

            # ── Deduct from sender ────────────────────────────────────────
            BranchBalance.objects.create(
                branch=office_from,
                office_balance=(
                    branch_balance_from.office_balance - amount
                    if transaction_method == 'cash'
                    else branch_balance_from.office_balance
                ),
                bank_balance=(
                    branch_balance_from.bank_balance - amount
                    if transaction_method == 'bank'
                    else branch_balance_from.bank_balance
                ),
                updated_by=processed_by,
            )

            # ── Add to receiver ───────────────────────────────────────────
            branch_balance_to = BranchBalance.objects.select_for_update().filter(branch=office_to).last()
            if branch_balance_to:
                BranchBalance.objects.create(
                    branch=office_to,
                    office_balance=(
                        branch_balance_to.office_balance + amount
                        if transaction_method == 'cash'
                        else branch_balance_to.office_balance
                    ),
                    bank_balance=(
                        branch_balance_to.bank_balance + amount
                        if transaction_method == 'bank'
                        else branch_balance_to.bank_balance
                    ),
                    updated_by=processed_by,
                )
            else:
                BranchBalance.objects.create(
                    branch=office_to,
                    office_balance=amount if transaction_method == 'cash' else Decimal('0.00'),
                    bank_balance=amount   if transaction_method == 'bank'  else Decimal('0.00'),
                    updated_by=processed_by,
                )

            # ── Record the inter-office transaction with attachment ────────
            office_txn = OfficeTransaction.objects.create(
                office_from=office_from,
                office_to=office_to,
                transaction_type=transaction_type,
                transaction_method=transaction_method,
                amount=amount,
                processed_by=processed_by,
            )
            if attachment:
                office_txn.attachment = attachment
                office_txn.save(update_fields=['attachment'])

            # ── Record in HQTransaction with same attachment path ─────────
            hq_txn = HQTransaction.objects.create(
                from_branch=office_from,
                to_branch=office_to,
                amount=amount,
                description=(
                    f'Branch transfer ({transaction_method.title()}) '
                    f'from {office_from.name} to {office_to.name}'
                ),
                transaction_date=transaction_date,
                processed_by=processed_by,
            )
            if attachment:
                hq_txn.attachment = office_txn.attachment   # reuse saved file path
                hq_txn.save(update_fields=['attachment'])

        messages.success(
            request,
            f'TZS {amount:,.0f} ({transaction_method.title()}) transferred from '
            f'{office_from.name} to {office_to.name} successfully.'
        )
        return redirect('bank_cash_transaction')

    return render(request, 'app/bank_cash_transaction_add.html', {
        **get_base_context(request),
        'office_from':    office_from,
        'branch_balance': branch_balance,
    })





def loan_receipt_select(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    clients = Client.objects.filter(
        loan_applications__repayments__isnull=False
    ).distinct().order_by('firstname', 'lastname')

    if filter_office:
        clients = clients.filter(
            loan_applications__office=filter_office.name
        ).distinct()

    return render(request, 'app/loan_receipt_select.html', {
        **base_ctx,
        'clients': clients,
    })



def loan_receipt_list(request):
    client_id = request.POST.get('client_id') or request.GET.get('client_id')

    if not client_id:
        messages.error(request, 'Please select a customer.')
        return redirect('loan_receipt_select')

    client        = get_object_or_404(Client, id=client_id)
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    all_clients = Client.objects.filter(
        loan_applications__repayments__isnull=False
    ).distinct().order_by('firstname', 'lastname')

    if filter_office:
        all_clients = all_clients.filter(
            loan_applications__office=filter_office.name
        ).distinct()

    repayments_qs = LoanRepayment.objects.filter(
        loan_application__client=client
    ).select_related(
        'loan_application', 'processed_by'
    ).order_by('created_at', 'id')

    if filter_office:
        repayments_qs = repayments_qs.filter(
            loan_application__office=filter_office.name
        )

    repayment_rows = []
    for r in repayments_qs:
        r.receipt_number = str(r.id).zfill(6)
        # Detect topup-clearance rows by payment_month being set differently
        # or by a description convention — adjust this logic to your actual pattern
        r.row_type = 'topup_clearance' if r.payment_month and r.loan_application.topups.filter(
            old_balance_cleared__gt=0
        ).exists() and r.repayment_amount == r.loan_application.topups.filter(
            old_balance_cleared__gt=0
        ).first().old_balance_cleared else 'repayment'
        repayment_rows.append(r)

    branch_name = '—'
    latest_loan = client.loan_applications.order_by('-created_at').first()
    if latest_loan and latest_loan.office:
        branch_name = latest_loan.office.upper()

    return render(request, 'app/loan_receipt_list.html', {
        **base_ctx,
        'client':         client,
        'all_clients':    all_clients,
        'repayment_rows': repayment_rows,
        'branch_name':    branch_name,
    })


from django.urls import reverse
@require_POST
def delete_repayment(request, repayment_id):
    """
    Reverse of loan_repayment():
    - loan_repayment() creates a new BranchBalance snapshot adding the amount
      to office_balance (cash) or bank_balance (non-cash).
    - delete_repayment() creates a new BranchBalance snapshot subtracting
      the same amount from the same field, then deletes the repayment record.
      LoanRepayment.delete() override already restores repayment_amount_remaining.
    """
    from django.db import transaction as db_transaction
    from django.db.models import F

    repayment = get_object_or_404(LoanRepayment, id=repayment_id)
    loan      = repayment.loan_application
    amount    = repayment.repayment_amount
    method    = repayment.transaction_method or 'cash'

    # Resolve the branch office (same way loan_repayment does via get_selected_office)
    branch_office = None
    if loan.office:
        try:
            branch_office = Office.objects.get(name=loan.office)
        except Office.DoesNotExist:
            pass

    try:
        with db_transaction.atomic():

            # ── 1. Create a new BranchBalance snapshot (reverse of repayment) ──
            if branch_office:
                last_bb = BranchBalance.objects.filter(
                    branch=branch_office
                ).order_by('-last_updated', '-id').first()

                if last_bb:
                    # Mirror the exact reverse of what loan_repayment() did:
                    # cash  → was added to office_balance → subtract from office_balance
                    # other → was added to bank_balance   → subtract from bank_balance
                    if method == 'cash':
                        new_office_bal = last_bb.office_balance + amount
                        new_bank_bal   = last_bb.bank_balance
                    else:
                        new_office_bal = last_bb.office_balance
                        new_bank_bal   = last_bb.bank_balance + amount

                    BranchBalance.objects.create(
                        branch         = branch_office,
                        office_balance = new_office_bal,
                        bank_balance   = new_bank_bal,
                        updated_by     = request.user,
                    )
                # If no BranchBalance row exists yet, nothing to reverse — skip silently

            # ── 2. Delete repayment ──────────────────────────────────────────
            # LoanRepayment.delete() override restores repayment_amount_remaining
            repayment.delete()

            # ── 3. If loan was marked Paid, reopen it ────────────────────────
            loan.refresh_from_db()
            if loan.status == 'Paid' and loan.repayment_amount_remaining > 0:
                loan.status = 'Approved'
                loan.save(update_fields=['status'])

        # return JsonResponse({'success': True, 'amount': str(amount)})
        client_id = loan.client_id
        return redirect(f"{reverse('loan_receipt_list')}?client_id={client_id}")
    

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

def loan_repayment_receipt(request, pk):
    repayment = get_object_or_404(
        LoanRepayment.objects.select_related(
            'loan_application__client',
            'processed_by'
        ),
        pk=pk
    )

    loan   = repayment.loan_application
    client = loan.client

    receipt_number = str(repayment.id).zfill(6)

    client_fullname = ' '.join(filter(None, [
        client.firstname,
        client.middlename,
        client.lastname,
    ])).upper()

    swahili_months = {
        1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR', 5: 'MEI', 6: 'JUN',
        7: 'JUL', 8: 'AUG', 9: 'SEP', 10: 'OKT', 11: 'NOV', 12: 'DEC',
    }

    # ── Use payment_month field (the month the client is paying for) ──
    # Fall back to repayment_date if payment_month was not set
    month_source = repayment.payment_month or repayment.repayment_date
    payment_month = (
        f"{swahili_months.get(month_source.month, month_source.month)}/{month_source.year}"
    )

    paid_up_to_this = LoanRepayment.objects.filter(
        loan_application=loan,
        id__lte=repayment.id
    ).aggregate(total=Sum('repayment_amount'))['total'] or Decimal('0.00')

    balance_after = max(
        (loan.total_repayment_amount or Decimal('0.00')) - paid_up_to_this,
        Decimal('0.00')
    )
    
    # keep payment_month as is, add this line below it
    payment_date = month_source.strftime("%d/%m/%Y")

    officer_user = repayment.processed_by
    officer_name = (
        officer_user.get_full_name().strip() or officer_user.username
    ).upper()

    return render(request, 'app/loan_repayment_receipt.html', {
        **get_base_context(request),
        'repayment':       repayment,
        'loan':            loan,
        'client':          client,
        'receipt_number':  receipt_number,
        'client_fullname': client_fullname,
        'payment_month':   payment_month,
        'balance_after':   balance_after,
        'officer_name':    officer_name,
        'payment_date':    payment_date,
    })


def loan_outstanding_select(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    clients = Client.objects.filter(
        loan_applications__isnull=False
    ).distinct().order_by('firstname', 'lastname')

    if filter_office:
        clients = clients.filter(
            loan_applications__office=filter_office.name
        ).distinct()

    return render(request, 'app/loan_outstanding_select.html', {
        **base_ctx,
        'clients': clients,
    })


def loan_outstanding_report(request):
    client_id = request.POST.get('client_id') or request.GET.get('client_id')

    if not client_id:
        return redirect('loan_outstanding_select')

    client        = get_object_or_404(Client, id=client_id)
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    all_clients = Client.objects.filter(
        loan_applications__isnull=False
    ).distinct().order_by('firstname', 'lastname')

    if filter_office:
        all_clients = all_clients.filter(
            loan_applications__office=filter_office.name
        ).distinct()

    loans = LoanApplication.objects.filter(
        client=client
    ).prefetch_related('repayments').order_by('created_at')

    if filter_office:
        loans = loans.filter(office=filter_office.name)

    branch_name = '—'
    latest_loan = loans.order_by('-created_at').first()
    if latest_loan and latest_loan.office:
        branch_name = latest_loan.office.upper()

    loan_rows = []
    for loan in loans:
        repayment_list = list(loan.repayments.all())
        paid_amount    = sum(r.repayment_amount for r in repayment_list)
        outstanding    = max(loan.repayment_amount_remaining, Decimal('0.00'))

        loan_rows.append({
            'loan':                loan,
            'loan_id_label':       f"{(loan.office or 'loan').lower()}-{loan.id}",
            'loan_amount':         loan.loan_amount or Decimal('0.00'),
            'interest_amount':     loan.total_interest_amount or Decimal('0.00'),
            'penalty_amount':      Decimal('0.00'),
            'total_amount':        loan.total_repayment_amount or Decimal('0.00'),
            'paid_amount':         paid_amount,
            'outstanding_balance': outstanding,
            'repayment_count':     len(repayment_list),   # ← NEW
            'is_approved':         loan.is_approved,       # ← NEW
        })

    totals = {
        'loan_amount':         sum(r['loan_amount']         for r in loan_rows),
        'interest_amount':     sum(r['interest_amount']     for r in loan_rows),
        'penalty_amount':      sum(r['penalty_amount']      for r in loan_rows),
        'total_amount':        sum(r['total_amount']        for r in loan_rows),
        'paid_amount':         sum(r['paid_amount']         for r in loan_rows),
        'outstanding_balance': sum(r['outstanding_balance'] for r in loan_rows),
    }

    return render(request, 'app/loan_outstanding_report.html', {
        **base_ctx,
        'client':      client,
        'all_clients': all_clients,
        'loan_rows':   loan_rows,
        'totals':      totals,
        'branch_name': branch_name,
    })

from django.db import transaction as db_transaction
@require_POST
def delete_loan(request, loan_id):
    loan = get_object_or_404(LoanApplication, id=loan_id)

    # Safety: only delete if no repayments exist
    if loan.repayments.exists():
        return JsonResponse(
            {'success': False, 'error': 'Cannot delete a loan that has repayments.'},
            status=400
        )

    client_id = loan.client_id

    try:
        with db_transaction.atomic():
            # Find the office by name (matching how it was stored on loan creation)
            branch_office = Office.objects.filter(name=loan.office).first()

            if branch_office:
                branch_balance = BranchBalance.objects.filter(branch=branch_office).last()

                if branch_balance:
                    loan_amount = loan.loan_amount
                    transaction_method = loan.transaction_method

                    if transaction_method == 'cash':
                        BranchBalance.objects.create(
                            branch=branch_office,
                            office_balance=branch_balance.office_balance + loan_amount,
                            bank_balance=branch_balance.bank_balance,
                            updated_by=request.user,
                        )
                    else:
                        BranchBalance.objects.create(
                            branch=branch_office,
                            office_balance=branch_balance.office_balance,
                            bank_balance=branch_balance.bank_balance + loan_amount,
                            updated_by=request.user,
                        )

            loan.delete()

        return JsonResponse({'success': True, 'client_id': client_id})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_POST  
def toggle_loan_approve(request, loan_id):
    loan = get_object_or_404(LoanApplication, id=loan_id)
    loan.is_approved = not loan.is_approved
    loan.save(update_fields=['is_approved'])
    return JsonResponse({
        'success': True,
        'is_approved': loan.is_approved,
    })
    

from decimal import Decimal, ROUND_CEILING

def loan_repayment_schedule(request, loan_id):
    loan     = get_object_or_404(LoanApplication, id=loan_id)
    client   = loan.client
    base_ctx = get_base_context(request)

    branch_name     = (loan.office or '—').upper()
    client_fullname = ' '.join(filter(None, [
        client.firstname, client.middlename, client.lastname
    ]))

    repayments = list(loan.repayments.all().order_by('repayment_date', 'id'))
    topups     = list(loan.topups.all())

    # ── Build paid_by_month from LoanRepayment ───────────────────────────
    paid_by_month = {}
    for r in repayments:
        if r.payment_month:
            key = (r.payment_month.year, r.payment_month.month)
            paid_by_month[key] = paid_by_month.get(key, Decimal('0.00')) + (r.repayment_amount or Decimal('0.00'))

    # ── Build topup_by_month from LoanTopup ──────────────────────────────
    topup_by_month = {}
    for t in topups:
        if t.payment_month:
            key = (t.payment_month.year, t.payment_month.month)
            topup_by_month[key] = topup_by_month.get(key, Decimal('0.00')) + (t.old_balance_cleared or Decimal('0.00'))

    # ── Rounding: ceil to nearest 1,000 ──────────────────────────────────
    def ceil_1000(val):
        return (val / Decimal('1000')).to_integral_value(rounding=ROUND_CEILING) * Decimal('1000')

    # ── Schedule parameters ──────────────────────────────────────────────
    periods      = loan.payment_period_months or 1
    loan_amount  = loan.loan_amount or Decimal('0.00')
    total_int    = loan.total_interest_amount or Decimal('0.00')
    total_return = loan_amount + total_int

    std_monthly   = ceil_1000(total_return / periods)
    std_principal = ceil_1000(loan_amount  / periods)
    std_interest  = std_monthly - std_principal

    last_principal = loan_amount - (std_principal * (periods - 1))
    last_interest  = total_int   - (std_interest  * (periods - 1))
    last_monthly   = last_principal + last_interest

    month_names = {
        1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
        7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
    }

    start = loan.first_repayment_date or loan.application_date

    # ── Running totals ───────────────────────────────────────────────────
    total_principal = Decimal('0.00')
    total_interest  = Decimal('0.00')
    total_penalty   = Decimal('0.00')
    total_total     = Decimal('0.00')
    total_paid_sum  = Decimal('0.00')

    schedule_rows    = []
    scheduled_months = set()

    # ── Pass 1: build raw scheduled rows ─────────────────────────────────
    scheduled_raw   = []
    running_balance = total_return

    for i in range(periods):
        year  = start.year  + (start.month - 1 + i) // 12
        month = (start.month - 1 + i) % 12 + 1
        scheduled_months.add((year, month))

        month_label = f"{month_names[month]}/{year}"
        is_last     = (i == periods - 1)

        p = last_principal if is_last else std_principal
        n = last_interest  if is_last else std_interest
        t = last_monthly   if is_last else std_monthly

        running_balance -= t
        if abs(running_balance) < Decimal('0.50'):
            running_balance = Decimal('0.00')

        paid = paid_by_month.get((year, month), Decimal('0.00'))

        scheduled_raw.append({
            'month_label':  month_label,
            'principal':    p,
            'interest':     n,
            'penalty':      Decimal('0.00'),
            'total':        t,
            'paid':         paid,
            'loan_balance': running_balance,
        })

        total_principal += p
        total_interest  += n
        total_total     += t
        total_paid_sum  += paid

    # ── Pass 2: pool all payments, apply top-down oldest debt first ───────
    remaining_pool = sum(r['paid'] for r in scheduled_raw)

    for row in scheduled_raw:
        t = row['total']

        if remaining_pool >= t:
            remaining_pool -= t
            outstanding = Decimal('0.00')
        elif remaining_pool > 0:
            outstanding    = t - remaining_pool
            remaining_pool = Decimal('0.00')
        else:
            outstanding = t

        schedule_rows.append({
            'month_label':  row['month_label'],
            'principal':    row['principal'],
            'interest':     row['interest'],
            'penalty':      row['penalty'],
            'total':        row['total'],
            'paid':         row['paid'],
            'outstanding':  outstanding,
            'loan_balance': row['loan_balance'],
            'extra':        False,
            'is_topup':     False,
        })

    final_outstanding = sum(
        (r['outstanding'] for r in schedule_rows if not r['extra']),
        Decimal('0.00')
    )

    # ── Extra rows: repayments outside schedule ───────────────────────────
    extra_repayment_months = sorted(
        [(y, m) for (y, m) in paid_by_month if (y, m) not in scheduled_months]
    )
    for (year, month) in extra_repayment_months:
        paid        = paid_by_month[(year, month)]
        month_label = f"{month_names[month]}/{year}"
        schedule_rows.append({
            'month_label':  month_label,
            'principal':    Decimal('0.00'),
            'interest':     Decimal('0.00'),
            'penalty':      Decimal('0.00'),
            'total':        Decimal('0.00'),
            'paid':         paid,
            'outstanding':  Decimal('0.00'),
            'loan_balance': Decimal('0.00'),
            'extra':        True,
            'is_topup':     False,
        })
        total_paid_sum += paid

    # ── Extra rows: topups outside schedule ──────────────────────────────
    extra_topup_months = sorted(
        [(y, m) for (y, m) in topup_by_month if (y, m) not in scheduled_months]
    )
    for (year, month) in extra_topup_months:
        paid        = topup_by_month[(year, month)]
        month_label = f"{month_names[month]}/{year}"
        schedule_rows.append({
            'month_label':  month_label,
            'principal':    Decimal('0.00'),
            'interest':     Decimal('0.00'),
            'penalty':      Decimal('0.00'),
            'total':        Decimal('0.00'),
            'paid':         paid,
            'outstanding':  Decimal('0.00'),
            'loan_balance': Decimal('0.00'),
            'extra':        True,
            'is_topup':     True,
        })
        total_paid_sum += paid

    # ── Sort: scheduled first, extras at bottom ───────────────────────────
    scheduled_rows = [r for r in schedule_rows if not r['extra']]
    extra_rows     = sorted(
        [r for r in schedule_rows if r['extra']],
        key=lambda r: r['month_label']
    )
    schedule_rows = scheduled_rows + extra_rows

    totals = {
        'principal':   total_principal,
        'interest':    total_interest,
        'penalty':     total_penalty,
        'total':       total_total,
        'paid':        total_paid_sum,
        'outstanding': final_outstanding,
    }

    return render(request, 'app/loan_repayment_schedule.html', {
        **base_ctx,
        'loan':            loan,
        'client':          client,
        'client_fullname': client_fullname,
        'branch_name':     branch_name,
        'schedule_rows':   schedule_rows,
        'totals':          totals,
        'total_return':    total_return,
    })   
 
    

def customer_statement_select(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    clients = Client.objects.filter(
        loan_applications__isnull=False
    ).distinct().order_by('firstname', 'lastname')

    if filter_office:
        clients = clients.filter(
            loan_applications__office=filter_office.name
        ).distinct()

    return render(request, 'app/customer_statement_select.html', {
        **base_ctx,
        'clients': clients,
    })



def customer_statement(request):
    client_id = request.POST.get('client_id') or request.GET.get('client_id')

    if not client_id:
        return redirect('customer_statement_select')

    client        = get_object_or_404(Client, id=client_id)
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    all_clients = Client.objects.filter(
        loan_applications__isnull=False
    ).distinct().order_by('firstname', 'lastname')

    if filter_office:
        all_clients = all_clients.filter(
            loan_applications__office=filter_office.name
        ).distinct()

    loans = LoanApplication.objects.filter(
        client=client
    ).prefetch_related('repayments', 'topups').order_by('application_date', 'id')

    if filter_office:
        loans = loans.filter(office=filter_office.name)

    branch_name = '—'
    latest_loan = loans.order_by('-created_at').first()
    if latest_loan and latest_loan.office:
        branch_name = latest_loan.office.upper()

    # ── Build loan blocks ─────────────────────────────────────────────
    # Each block = one LoanApplication with all its rows pre-computed.
    # Blocks are ordered oldest loan first → newest loan last.
    # Within a block: disbursement → repayments/topups in date order.
    # Top-up events within a block:
    #   same-date repayments → clearance → new topup disbursement
    # The global running balance flows across ALL blocks so the
    # "Balance" column is always the client's total outstanding.

    loan_blocks    = []   # list of {'loan': obj, 'rows': [...]}
    global_balance = Decimal('0.00')

    for loan in loans:
        loan_id_label   = f"{(loan.office or 'loan').lower()}-{loan.id}"
        total_repayable = loan.total_repayment_amount or Decimal('0.00')

        block_rows = []

        # ── Repayments grouped by payment_month ───────────────────────
        # Sorted and keyed on payment_month (DateField).
        # Records with payment_month=None are skipped to avoid
        # TypeError when sorting mixed None/date values.
        repayments_by_date = {}
        for r in loan.repayments.all().order_by('payment_month', 'id'):
            if r.payment_month is None:
                continue
            repayments_by_date.setdefault(r.payment_month, []).append(r)

        # ── Topups grouped by topup_date ──────────────────────────────
        topups_by_date = {}
        for t in loan.topups.all().order_by('topup_date', 'id'):
            if t.topup_date is None:
                continue
            topups_by_date.setdefault(t.topup_date, []).append(t)

        # All activity dates for this loan, sorted ascending.
        # Repayments contribute their payment_month; topups their topup_date.
        all_dates = sorted(
            date for date in
            set(list(repayments_by_date.keys()) + list(topups_by_date.keys()))
            if date is not None
        )

        # ── 1. Original disbursement ──────────────────────────────────
        global_balance += total_repayable
        block_rows.append({
            'type':          'disbursement',
            'date':          loan.application_date,
            'receipt_no':    str(loan.id).zfill(6),
            'description':   'Loan Taken and interest',
            'loan_type':     loan.loan_type,
            'loan_id_label': loan_id_label,
            'loan_amount':   total_repayable,
            'paid_amount':   Decimal('0.00'),
            'balance':       global_balance,
            'record_id':     loan.id,
        })

        # ── 2. Activity (repayments + topups) in date order ───────────
        # On a topup date the last repayment is omitted because it
        # represents the same transaction as the clearance row — showing
        # both would double-count the payment.
        for day in all_dates:
            day_repayments = repayments_by_date.get(day, [])
            day_topups     = topups_by_date.get(day, [])

            if day_topups:
                # Show all repayments except the last one; the clearance
                # row below replaces it.
                repayments_to_show = day_repayments[:-1]
            else:
                repayments_to_show = day_repayments

            # ── Repayments for this date (filtered) ───────────────────
            for r in repayments_to_show:
                global_balance = max(
                    global_balance - r.repayment_amount,
                    Decimal('0.00')
                )
                block_rows.append({
                    'type':          'repayment',
                    'date':          r.repayment_date,   # display still uses repayment_date
                    'receipt_no':    str(r.id).zfill(6),
                    'description':   'Loan payment',
                    'loan_type':     loan.loan_type,
                    'loan_id_label': loan_id_label,
                    'loan_amount':   Decimal('0.00'),
                    'paid_amount':   r.repayment_amount,
                    'balance':       global_balance,
                    'record_id':     r.id,
                })

            # ── Topup events: clearance only ──────────────────────────
            # The topup disbursement is NOT shown here — it appears as
            # the next loan block's "Loan Taken and interest" row,
            # keeping the statement clean and non-duplicated.
            for topup in day_topups:
                old_balance = topup.old_balance_cleared or global_balance or Decimal('0.00')

                if old_balance > Decimal('0.00'):
                    global_balance = Decimal('0.00')
                    block_rows.append({
                        'type':          'topup_clearance',
                        'date':          topup.topup_date,
                        'receipt_no':    str(topup.id).zfill(6),
                        'description':   'Clearance loan balance for top-up',
                        'loan_type':     loan.loan_type,
                        'loan_id_label': loan_id_label,
                        'loan_amount':   Decimal('0.00'),
                        'paid_amount':   old_balance,
                        'balance':       global_balance,
                        'record_id':     topup.id,
                    })

        loan_blocks.append({
            'loan':          loan,
            'rows':          block_rows,
            'loan_id_label': loan_id_label,
        })

    return render(request, 'app/customer_statement.html', {
        **base_ctx,
        'client':      client,
        'all_clients': all_clients,
        'branch_name': branch_name,
        'loan_blocks': loan_blocks,
    })

    




def edit_repayment_2(request, repayment_id, record_type='repayment'):
    """
    Edit page for a LoanRepayment (loan payment) or a LoanTopup (clearance).

    URL patterns:
        path('repayment/edit/<int:repayment_id>/', edit_repayment, {'record_type': 'repayment'}, name='edit_repayment'),
        path('topup/edit/<int:repayment_id>/',    edit_repayment, {'record_type': 'topup'},      name='edit_topup'),
    """
    if record_type == 'topup':
        record = get_object_or_404(LoanTopup, id=repayment_id)
        client = record.loan_application.client
    else:
        record = get_object_or_404(LoanRepayment, id=repayment_id)
        client = record.loan_application.client

    # All loans for this client (for the Loan ID dropdown)
    client_loans = LoanApplication.objects.filter(
        client=client
    ).order_by('application_date', 'id')

    receipt_no = str(record.id).zfill(6)

    if record_type == 'topup':
        loan_id_label = (
            f"{(record.loan_application.office or 'loan').lower()}"
            f"-{record.loan_application.id}"
        )
        current_loan_id    = record.loan_application.id
        amount             = record.old_balance_cleared
        record_date        = record.topup_date
        transaction_method = record.transaction_method or ''
        title = f"SWARP PAYMENT RECEIPT-{receipt_no} [LOAN ID: {loan_id_label}]"
    else:
        loan_id_label = (
            f"{(record.loan_application.office or 'loan').lower()}"
            f"-{record.loan_application.id}"
        )
        current_loan_id    = record.loan_application.id
        amount             = record.repayment_amount
        record_date        = record.repayment_date
        transaction_method = record.transaction_method or ''
        title = f"SWARP PAYMENT RECEIPT-{receipt_no} [LOAN ID: {loan_id_label}]"

    if request.method == 'POST':
        new_loan_id            = request.POST.get('loan_id')
        new_amount_str         = request.POST.get('amount', '').replace(',', '').strip()
        new_date_str           = request.POST.get('date', '')
        # transaction_method field removed from form — keep existing value

        errors = []
        if not new_loan_id:    errors.append('Loan ID is required.')
        if not new_amount_str: errors.append('Amount is required.')
        if not new_date_str:   errors.append('Date is required.')

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            try:
                new_amount = Decimal(new_amount_str)
                new_date   = datetime.datetime.strptime(new_date_str, '%Y-%m-%d').date()
                new_loan   = get_object_or_404(LoanApplication, id=new_loan_id, client=client)

                if record_type == 'topup':
                    old_amount = record.old_balance_cleared or Decimal('0')

                    # Restore old cleared amount to previous loan's remaining balance
                    LoanApplication.objects.filter(pk=record.loan_application_id).update(
                        repayment_amount_remaining=models.F('repayment_amount_remaining') + old_amount
                    )

                    record.loan_application    = new_loan
                    record.old_balance_cleared = new_amount
                    record.topup_date          = new_date
                    # transaction_method unchanged
                    record.save(update_fields=[
                        'loan_application', 'old_balance_cleared',
                        'topup_date'
                    ])

                    # Deduct new cleared amount from new loan's remaining balance
                    LoanApplication.objects.filter(pk=new_loan.id).update(
                        repayment_amount_remaining=models.F('repayment_amount_remaining') - new_amount
                    )

                else:
                    old_amount = record.repayment_amount

                    # Restore old repayment to previous loan's remaining balance
                    LoanApplication.objects.filter(pk=record.loan_application_id).update(
                        repayment_amount_remaining=models.F('repayment_amount_remaining') + old_amount
                    )

                    record.loan_application   = new_loan
                    record.repayment_amount   = new_amount
                    record.repayment_date     = new_date
                    # transaction_method unchanged
                    record.save(update_fields=[
                        'loan_application', 'repayment_amount',
                        'repayment_date'
                    ])

                    # Deduct new repayment from new loan's remaining balance
                    LoanApplication.objects.filter(pk=new_loan.id).update(
                        repayment_amount_remaining=models.F('repayment_amount_remaining') - new_amount
                    )

                messages.success(request, 'Record updated successfully.')
                return redirect(f"{reverse('customer_statement')}?client_id={client.id}")

            except Exception as e:
                messages.error(request, f'Error updating record: {str(e)}')

    context = {
        'title':              title,
        'record':             record,
        'record_type':        record_type,
        'receipt_no':         receipt_no,
        'client':             client,
        'client_loans':       client_loans,
        'current_loan_id':    current_loan_id,
        'amount':             amount,
        'record_date':        record_date,
        'transaction_method': transaction_method,
        'loan_id_label':      loan_id_label,
    }
    return render(request, 'app/edit_repayment_2.html', context)

  

# ── Helper ─────────────────────────────────────────────────────────────
def _officer_name(user):
    """Return 'Firstname Lastname Firstname Lastname' style, matching screenshot."""
    if not user:
        return '—'
    full = user.get_full_name().strip()
    return f"{full}" if full else user.username


# ══════════════════════════════════════════════════════════════════════
#  PAGE 1 — Date range filter
# ══════════════════════════════════════════════════════════════════════
def branch_transaction_statement(request):
    return render(request, 'app/branch_txn_filter.html', {
        **get_base_context(request),
    })



def branch_transaction_statement_report(request):
    if request.method != 'POST':
        return redirect('branch_transaction_statement')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name if selected_office else 'All Branches'
    office_name     = filter_office.name if filter_office else None

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('branch_transaction_statement')

    date_from_display = date_from.strftime('%d/%m/%Y')
    date_to_display   = date_to.strftime('%d/%m/%Y')

    import pytz
    from django.utils import timezone as dj_timezone
    from datetime import timezone as py_timezone

    # ── Extend date_to to include the full end day ────────────────────
    date_from_dt = datetime.datetime.combine(date_from, datetime.time.min)  # 00:00:00
    date_to_dt   = datetime.datetime.combine(date_to,   datetime.time.max)  # 23:59:59.999999

    _epoch = datetime.datetime(2000, 1, 1, tzinfo=py_timezone.utc)

    def _aware(dt):
        if dt is None:
            return _epoch
        if dj_timezone.is_naive(dt):
            return dj_timezone.make_aware(dt, pytz.UTC)
        return dt

    raw = []

    # ── 1. Loan Repayments (CREDIT) ───────────────────────────────────
    repayments = LoanRepayment.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('loan_application__client', 'processed_by').order_by('created_at', 'id')

    if office_name:
        repayments = repayments.filter(loan_application__office=office_name)

    for r in repayments:
        c = r.loan_application.client
        raw.append({
            'created_at':       _aware(r.created_at),
            'sort_sub':         0,
            'record_type':      'repayment',
            'record_id':        r.id,
            'date':             r.created_at,
            'receipt_no':       str(r.id).zfill(6),
            'name':             f"{c.firstname} {c.middlename or ''} {c.lastname}".strip(),
            'description':      'Loan payment',
            'description_bold': False,
            'is_expense':       False,
            'credit':           r.repayment_amount,
            'debit':            None,
            'processed_by':     _officer_name(r.processed_by),
            'deletable':        True,
        })

    # ── 2. Loan Top-ups ───────────────────────────────────────────────
    topups = LoanTopup.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('loan_application__client', 'processed_by').order_by('created_at', 'id')

    if office_name:
        topups = topups.filter(loan_application__office=office_name)

    topup_loan_ids = set()
    for t in topups:
        topup_loan_ids.add(t.loan_application_id)
        c           = t.loan_application.client
        old_balance = t.old_balance_cleared or t.loan_application.repayment_amount_remaining
        client_name = f"{c.firstname} {c.middlename or ''} {c.lastname}".strip()
        officer     = _officer_name(t.processed_by)
        ts          = _aware(t.created_at)

        if old_balance and old_balance > 0:
            raw.append({
                'created_at':       ts,
                'sort_sub':         1,
                'record_type':      'topup',
                'record_id':        t.id,
                'date':             t.created_at,
                'receipt_no':       str(t.id).zfill(6),
                'name':             client_name,
                'description':      'Clearance loan balance for top-up',
                'description_bold': False,
                'is_expense':       False,
                'credit':           old_balance,
                'debit':            None,
                'processed_by':     officer,
                'deletable':        True,
            })

        raw.append({
            'created_at':       ts,
            'sort_sub':         2,
            'record_type':      'topup',
            'record_id':        t.id,
            'date':             t.created_at,
            'receipt_no':       str(t.id).zfill(6),
            'name':             client_name,
            'description':      'Loan amount deposited to customer',
            'description_bold': False,
            'is_expense':       False,
            'credit':           None,
            'debit':            t.topup_amount,
            'processed_by':     officer,
            'deletable':        False,
        })

    # ── 3. New Loan Disbursements (DEBIT) ─────────────────────────────
    disbursements = LoanApplication.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('client', 'processed_by').order_by('created_at', 'id')

    if office_name:
        disbursements = disbursements.filter(office=office_name)

    for loan in disbursements:
        if loan.id in topup_loan_ids:
            continue
        c = loan.client
        raw.append({
            'created_at':       _aware(loan.created_at),
            'sort_sub':         0,
            'record_type':      'loan',
            'record_id':        loan.id,
            'date':             loan.created_at,
            'receipt_no':       str(loan.id).zfill(6),
            'name':             f"{c.firstname} {c.middlename or ''} {c.lastname}".strip(),
            'description':      'Loan amount deposited to customer',
            'description_bold': False,
            'is_expense':       False,
            'credit':           None,
            'debit':            loan.loan_amount,
            'processed_by':     _officer_name(loan.processed_by),
            'deletable':        not loan.repayments.exists() and not loan.topups.exists(),
        })

    # ── 4. Expenses (DEBIT) ───────────────────────────────────────────
    expenses = Expense.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('recorded_by', 'transaction_type').order_by('created_at', 'id')

    if office_name:
        expenses = expenses.filter(office=office_name)

    for exp in expenses:
        cat = exp.transaction_type.name if exp.transaction_type else 'Expense'
        raw.append({
            'created_at':       _aware(exp.created_at),
            'sort_sub':         0,
            'record_type':      'expense',
            'record_id':        exp.id,
            'date':             exp.created_at,
            'receipt_no':       str(exp.id).zfill(6),
            'name':             _officer_name(exp.recorded_by),
            'description':      f"{cat} [{exp.description}],",
            'description_bold': True,
            'is_expense':       True,
            'credit':           None,
            'debit':            exp.amount,
            'processed_by':     _officer_name(exp.recorded_by),
            'attachment':       exp.attachment.url if exp.attachment else None,
            'deletable':        True,
        })

    # ── 5. Salaries (DEBIT) ───────────────────────────────────────────
    salaries = Salary.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('employee', 'processed_by').order_by('created_at', 'id')

    if filter_office:
        salaries = salaries.filter(fund_source=filter_office)
    elif office_name:
        salaries = salaries.filter(fund_source__name=office_name)

    for sal in salaries:
        raw.append({
            'created_at':       _aware(sal.created_at),
            'sort_sub':         0,
            'record_type':      'salary',
            'record_id':        sal.id,
            'date':             sal.created_at,
            'receipt_no':       str(sal.id).zfill(6),
            'name':             _officer_name(sal.employee),
            'description':      'Salary payment',
            'description_bold': False,
            'is_expense':       False,
            'credit':           None,
            'debit':            sal.amount,
            'processed_by':     _officer_name(sal.processed_by),
            'deletable':        True,
        })

    # ── 6. Bank Charges (DEBIT) ───────────────────────────────────────
    bank_charges = BankCharge.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('recorded_by').order_by('created_at', 'id')

    if office_name:
        bank_charges = bank_charges.filter(office=office_name)

    for bc in bank_charges:
        raw.append({
            'created_at':       _aware(bc.created_at),
            'sort_sub':         0,
            'record_type':      'bankcharge',
            'record_id':        bc.id,
            'date':             bc.created_at,
            'receipt_no':       str(bc.id).zfill(6),
            'name':             _officer_name(bc.recorded_by),
            'description':      f"Bank Charge [{bc.description}]",
            'description_bold': True,
            'is_expense':       True,
            'credit':           None,
            'debit':            bc.amount,
            'processed_by':     _officer_name(bc.recorded_by),
            'deletable':        True,
        })

    # ── 7. HQ Transfers ───────────────────────────────────────────────
    hq_transactions = HQTransaction.objects.filter(
        created_at__gte=date_from_dt,
        created_at__lte=date_to_dt,
    ).select_related('from_branch', 'to_branch', 'processed_by').order_by('processed_at', 'id')

    if office_name:
        hq_transactions = hq_transactions.filter(
            Q(from_branch__name=office_name) | Q(to_branch__name=office_name)
        )

    for hq in hq_transactions:
        officer = _officer_name(hq.processed_by)
        ts      = _aware(hq.processed_at)

        if office_name:
            is_received = (hq.to_branch and hq.to_branch.name == office_name)
            is_sent     = (hq.from_branch and hq.from_branch.name == office_name)
        else:
            is_received = True
            is_sent     = False

        if is_received:
            from_label = hq.from_branch.name if hq.from_branch else 'HQ'
            raw.append({
                'created_at':       ts,
                'sort_sub':         1,
                'record_type':      'hq',
                'record_id':        hq.id,
                'date':             hq.created_at,
                'receipt_no':       str(hq.id).zfill(6),
                'name':             from_label,
                'description':      f"Transfer received from {from_label}",
                'description_bold': False,
                'is_expense':       False,
                'credit':           hq.amount,
                'debit':            None,
                'processed_by':     officer,
                'attachment':       hq.attachment.url if getattr(hq, 'attachment', None) else None,
                'deletable':        True,
            })

        if is_sent:
            to_label = hq.to_branch.name if hq.to_branch else 'HQ'
            raw.append({
                'created_at':       ts,
                'sort_sub':         2,
                'record_type':      'hq',
                'record_id':        hq.id,
                'date':             hq.created_at,
                'receipt_no':       str(hq.id).zfill(6),
                'name':             to_label,
                'description':      f"Transfer sent to {to_label}",
                'description_bold': False,
                'is_expense':       True,
                'credit':           None,
                'debit':            hq.amount,
                'processed_by':     officer,
                'attachment':       hq.attachment.url if getattr(hq, 'attachment', None) else None,
                'deletable':        False,
            })

    # ── Sort ──────────────────────────────────────────────────────────
    raw.sort(key=lambda e: (e['date'], e['created_at'], e['sort_sub']))

    # ── Annotate hide_date ────────────────────────────────────────────
    prev_date = None
    for entry in raw:
        entry_date = entry['date'].date() if hasattr(entry['date'], 'date') else entry['date']
        entry['hide_date'] = (entry_date == prev_date)
        prev_date = entry_date

    grand_credit = sum(e['credit'] or Decimal('0.00') for e in raw)
    grand_debit  = sum(e['debit']  or Decimal('0.00') for e in raw)

    return render(request, 'app/branch_txn_report.html', {
        **base_ctx,
        'rows':              raw,
        'grand_credit':      grand_credit,
        'grand_debit':       grand_debit,
        'date_from_display': date_from_display,
        'date_to_display':   date_to_display,
        'branch_name':       branch_name,
    })
 
@require_POST
def delete_transaction(request, record_type, record_id):
    """
    Delete any transaction row from the branch transaction statement and
    reverse its effect on the relevant fund balances.
 
    record_type values (match what is set in the template):
        repayment   → LoanRepayment
        topup       → LoanTopup
        loan        → LoanApplication
        expense     → Expense
        salary      → Salary
        bankcharge  → BankCharge
        hq          → HQTransaction
    """
    try:
        with db_transaction.atomic():
 
            # ── 1. Loan Repayment ─────────────────────────────────────
            if record_type == 'repayment':
                record = get_object_or_404(LoanRepayment, id=record_id)
                # LoanRepayment.delete() already restores repayment_amount_remaining
                # via its own delete() override — just call it.
                record.delete()
 
            # ── 2. Loan Top-up ────────────────────────────────────────
            elif record_type == 'topup':
                record = get_object_or_404(LoanTopup, id=record_id)
 
                # Reverse: the topup disbursed money OUT of the branch balance.
                # Find the branch office and reverse the BranchBalance.
                office_name = record.loan_application.office
                branch_office = Office.objects.filter(name=office_name).first()
                if branch_office:
                    latest = BranchBalance.objects.filter(
                        branch=branch_office
                    ).order_by('-last_updated').first()
                    if latest:
                        method = record.transaction_method or 'cash'
                        if method == 'bank':
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance,
                                bank_balance=latest.bank_balance + record.topup_amount,
                                updated_by=request.user,
                            )
                        else:
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance + record.topup_amount,
                                bank_balance=latest.bank_balance,
                                updated_by=request.user,
                            )
 
                # Restore the loan's remaining balance back to what it was
                # before the topup (old_balance_cleared was zeroed on topup).
                LoanApplication.objects.filter(
                    pk=record.loan_application_id
                ).update(
                    repayment_amount_remaining=models.F('repayment_amount_remaining')
                    + record.old_balance_cleared
                )
 
                record.delete()
 
            # ── 3. Loan Disbursement ──────────────────────────────────
            elif record_type == 'loan':
                record = get_object_or_404(LoanApplication, id=record_id)
 
                # Safety: only delete if no repayments or topups exist
                if record.repayments.exists():
                    return JsonResponse(
                        {'success': False,
                         'error': 'Cannot delete a loan that already has repayments.'},
                        status=400
                    )
                if record.topups.exists():
                    return JsonResponse(
                        {'success': False,
                         'error': 'Cannot delete a loan that already has top-ups.'},
                        status=400
                    )
 
                # Reverse the BranchBalance deduction made when loan was issued
                office_name   = record.office
                branch_office = Office.objects.filter(name=office_name).first()
                if branch_office:
                    latest = BranchBalance.objects.filter(
                        branch=branch_office
                    ).order_by('-last_updated').first()
                    if latest:
                        method = record.transaction_method or 'cash'
                        if method == 'bank':
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance,
                                bank_balance=latest.bank_balance + record.loan_amount,
                                updated_by=request.user,
                            )
                        else:
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance + record.loan_amount,
                                bank_balance=latest.bank_balance,
                                updated_by=request.user,
                            )
 
                record.delete()
 
            # ── 4. Expense ────────────────────────────────────────────
            elif record_type == 'expense':
                record = get_object_or_404(Expense, id=record_id)
 
                # Expense debited the branch balance — restore it
                office_name   = record.office
                branch_office = Office.objects.filter(name=office_name).first()
                if branch_office:
                    latest = BranchBalance.objects.filter(
                        branch=branch_office
                    ).order_by('-last_updated').first()
                    if latest:
                        method = record.payment_method or 'cash'
                        if method == 'bank':
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance,
                                bank_balance=latest.bank_balance + record.amount,
                                updated_by=request.user,
                            )
                        else:
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance + record.amount,
                                bank_balance=latest.bank_balance,
                                updated_by=request.user,
                            )
 
                record.delete()
 
            # ── 5. Salary ─────────────────────────────────────────────
            elif record_type == 'salary':
                record = get_object_or_404(Salary, id=record_id)
 
                # Salary debited the fund_source branch — restore it
                branch_office = record.fund_source
                if branch_office:
                    latest = BranchBalance.objects.filter(
                        branch=branch_office
                    ).order_by('-last_updated').first()
                    if latest:
                        method = record.transaction_method or 'bank'
                        if method == 'bank':
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance,
                                bank_balance=latest.bank_balance + record.amount,
                                updated_by=request.user,
                            )
                        else:
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance + record.amount,
                                bank_balance=latest.bank_balance,
                                updated_by=request.user,
                            )
 
                record.delete()
 
            # ── 6. Bank Charge ────────────────────────────────────────
            elif record_type == 'bankcharge':
                record = get_object_or_404(BankCharge, id=record_id)
 
                # BankCharge debited the branch balance — restore it
                office_name   = record.office
                branch_office = Office.objects.filter(name=office_name).first()
                if branch_office:
                    latest = BranchBalance.objects.filter(
                        branch=branch_office
                    ).order_by('-last_updated').first()
                    if latest:
                        method = record.payment_method or 'bank'
                        if method == 'bank':
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance,
                                bank_balance=latest.bank_balance + record.amount,
                                updated_by=request.user,
                            )
                        else:
                            BranchBalance.objects.create(
                                branch=branch_office,
                                office_balance=latest.office_balance + record.amount,
                                bank_balance=latest.bank_balance,
                                updated_by=request.user,
                            )
 
                record.delete()
 
            # ── 7. HQ Transaction ─────────────────────────────────────
            elif record_type == 'hq':
                record = get_object_or_404(HQTransaction, id=record_id)
 
                # Reverse: money was sent FROM from_branch TO to_branch.
                # Restore from_branch balance and deduct from to_branch balance.
                if record.from_branch:
                    latest = BranchBalance.objects.filter(
                        branch=record.from_branch
                    ).order_by('-last_updated').first()
                    if latest:
                        BranchBalance.objects.create(
                            branch=record.from_branch,
                            office_balance=latest.office_balance,
                            bank_balance=latest.bank_balance + record.amount,
                            updated_by=request.user,
                        )
 
                if record.to_branch:
                    latest = BranchBalance.objects.filter(
                        branch=record.to_branch
                    ).order_by('-last_updated').first()
                    if latest:
                        BranchBalance.objects.create(
                            branch=record.to_branch,
                            office_balance=latest.office_balance,
                            bank_balance=latest.bank_balance - record.amount,
                            updated_by=request.user,
                        )
 
                record.delete()
 
            else:
                return JsonResponse(
                    {'success': False, 'error': f'Unknown record type: {record_type}'},
                    status=400
                )
 
        return JsonResponse({'success': True})
 
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500) 



   
    
# ── Helper ─────────────────────────────────────────────────────────────
def _collector_name(user):
    """Return full name duplicated, e.g. 'Anjera Jeremia Anjeran'."""
    if not user:
        return '—'
    full = user.get_full_name().strip()
    return f"{full}" if full else user.username


def _split_principal_interest(repayment_amount, interest_rate_annual, loan_amount, total_repayment):
    """
    Given a repayment amount, split it into principal and interest
    proportional to the loan's interest ratio.

    interest_ratio = total_interest / total_repayment
    interest_part  = repayment_amount * interest_ratio
    principal_part = repayment_amount - interest_part
    """
    if not total_repayment or total_repayment == 0:
        return repayment_amount, Decimal('0.00')

    interest_ratio = (total_repayment - loan_amount) / total_repayment
    interest_part  = (repayment_amount * interest_ratio).quantize(Decimal('1'))
    principal_part = repayment_amount - interest_part
    return principal_part, interest_part


# ══════════════════════════════════════════════════════════════════════
#  PAGE 1 — Date range filter
# ══════════════════════════════════════════════════════════════════════

def loan_collection_statement(request):
    """Renders the date-range filter form only."""
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/loan_collection_filter.html', context)


# ══════════════════════════════════════════════════════════════════════
#  PAGE 2 — Report
# ══════════════════════════════════════════════════════════════════════

# def loan_collection_statement_report(request):
#     if request.method != 'POST':
#         return redirect('loan_collection_statement')

#     base_ctx        = get_base_context(request)
#     filter_office   = base_ctx['filter_office']
#     selected_office = base_ctx['selected_office']
#     branch_name     = selected_office.name.upper() if selected_office else 'All Branches'
#     office_name     = filter_office.name if filter_office else None

#     date_from_str = request.POST.get('date_from', '')
#     date_to_str   = request.POST.get('date_to',   '')

#     try:
#         date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
#         date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
#     except ValueError:
#         return redirect('loan_collection_statement')

#     date_from_display = date_from.strftime('%d/%m/%Y')
#     date_to_display   = date_to.strftime('%d/%m/%Y')

#     # ── DateTimeField range: start of date_from → end of date_to ─────
#     dt_from = datetime.datetime.combine(date_from, datetime.time.min)  # 00:00:00
#     dt_to   = datetime.datetime.combine(date_to,   datetime.time.max)  # 23:59:59.999999

#     raw = []

#     # ── 1. Loan Repayments ────────────────────────────────────────────
#     repayments = LoanRepayment.objects.filter(
#         created_at__range=(dt_from, dt_to),        # ← changed
#     ).select_related(
#         'loan_application__client',
#         'processed_by',
#     ).order_by('created_at', 'id')                 # ← changed

#     if office_name:
#         repayments = repayments.filter(loan_application__office=office_name)

#     for r in repayments:
#         loan      = r.loan_application
#         client    = loan.client
#         rate      = loan.interest_rate or Decimal('0')
#         total_rep = loan.total_repayment_amount or Decimal('0')
#         loan_amt  = loan.loan_amount or Decimal('0')

#         principal, interest = _split_principal_interest(
#             r.repayment_amount, rate, loan_amt, total_rep
#         )

#         raw.append({
#             'sort_key':     (r.created_at, 0, r.id),   # ← changed
#             'date':         r.created_at.date(),        # ← changed
#             'receipt_no':   str(r.id).zfill(6),
#             'name':         f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
#             'description':  'Loan payment',
#             'rate':         rate,
#             'principal':    principal,
#             'interest':     interest,
#             'total':        r.repayment_amount,
#             'collected_by': _collector_name(r.processed_by),
#         })

#     # ── 2. Loan Top-up clearances ─────────────────────────────────────
#     topups = LoanTopup.objects.filter(
#         created_at__range=(dt_from, dt_to),        # ← changed
#     ).select_related(
#         'loan_application__client',
#         'processed_by',
#     ).order_by('created_at', 'id')                 # ← changed

#     if office_name:
#         topups = topups.filter(loan_application__office=office_name)

#     for t in topups:
#         loan        = t.loan_application
#         client      = loan.client
#         rate        = loan.interest_rate or Decimal('0')
#         total_rep   = loan.total_repayment_amount or Decimal('0')
#         loan_amt    = loan.loan_amount or Decimal('0')
#         old_balance = t.old_balance_cleared or Decimal('0')   # ← use topup's own field, not loan's remaining

#         if old_balance <= 0:
#             continue

#         principal, interest = _split_principal_interest(
#             old_balance, rate, loan_amt, total_rep
#         )

#         raw.append({
#             'sort_key':     (t.created_at, 1, t.id),   # ← changed
#             'date':         t.created_at.date(),        # ← changed
#             'receipt_no':   str(t.id).zfill(6),
#             'name':         f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
#             'description':  'Clearance loan balance for top-up',
#             'rate':         rate,
#             'principal':    principal,
#             'interest':     interest,
#             'total':        old_balance,
#             'collected_by': _collector_name(t.processed_by),
#         })

#     raw.sort(key=lambda e: e['sort_key'])

#     prev_date = None
#     for idx, entry in enumerate(raw, start=1):
#         entry['sn']        = idx
#         entry['hide_date'] = (entry['date'] == prev_date)
#         prev_date          = entry['date']

#     grand_principal = sum(e['principal'] for e in raw)
#     grand_interest  = sum(e['interest']  for e in raw)
#     grand_total     = sum(e['total']     for e in raw)

#     return render(request, 'app/loan_collection_report.html', {
#         **base_ctx,
#         'rows':              raw,
#         'grand_principal':   grand_principal,
#         'grand_interest':    grand_interest,
#         'grand_total':       grand_total,
#         'date_from_display': date_from_display,
#         'date_to_display':   date_to_display,
#         'branch_name':       branch_name,
#     })
    

def loan_collection_statement_report(request):
    if request.method != 'POST':
        return redirect('loan_collection_statement')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'
    office_name     = filter_office.name if filter_office else None

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('loan_collection_statement')

    date_from_display = date_from.strftime('%d/%m/%Y')
    date_to_display   = date_to.strftime('%d/%m/%Y')

    dt_from = datetime.datetime.combine(date_from, datetime.time.min)
    dt_to   = datetime.datetime.combine(date_to,   datetime.time.max)

    # ── Pre-compute topup clearance repayment IDs to exclude ─────────
    # Fetch all topups in range (no office filter needed here, we filter repayments after)
    topup_repayment_ids_to_exclude = set()

    candidate_topups = LoanTopup.objects.filter(
        created_at__range=(dt_from, dt_to),
        old_balance_cleared__gt=0,
    ).select_related('loan_application')

    if office_name:
        candidate_topups = candidate_topups.filter(loan_application__office=office_name)

    for t in candidate_topups:
        match = LoanRepayment.objects.filter(
            loan_application=t.loan_application,
            repayment_amount=t.old_balance_cleared,
            created_at__date=t.created_at.date(),
        ).values_list('id', flat=True).first()

        if match:
            topup_repayment_ids_to_exclude.add(match)

    raw = []

    # ── 1. Loan Repayments ────────────────────────────────────────────
    repayments = LoanRepayment.objects.filter(
        created_at__range=(dt_from, dt_to),
    ).exclude(
        id__in=topup_repayment_ids_to_exclude,  # ✅ key exclusion
    ).select_related(
        'loan_application__client',
        'processed_by',
    ).order_by('created_at', 'id')

    if office_name:
        repayments = repayments.filter(loan_application__office=office_name)

    for r in repayments:
        loan      = r.loan_application
        client    = loan.client
        rate      = loan.interest_rate or Decimal('0')
        total_rep = loan.total_repayment_amount or Decimal('0')
        loan_amt  = loan.loan_amount or Decimal('0')

        principal, interest = _split_principal_interest(
            r.repayment_amount, rate, loan_amt, total_rep
        )

        raw.append({
            'sort_key':     (r.created_at, 0, r.id),
            'date':         r.created_at.date(),
            'receipt_no':   str(r.id).zfill(6),
            'name':         f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'description':  'Loan payment',
            'rate':         rate,
            'principal':    principal,
            'interest':     interest,
            'total':        r.repayment_amount,
            'collected_by': _collector_name(r.processed_by),
        })

    # ── 2. Loan Top-up clearances ─────────────────────────────────────
    topups = LoanTopup.objects.filter(
        created_at__range=(dt_from, dt_to),
    ).select_related(
        'loan_application__client',
        'processed_by',
    ).order_by('created_at', 'id')

    if office_name:
        topups = topups.filter(loan_application__office=office_name)

    for t in topups:
        loan        = t.loan_application
        client      = loan.client
        rate        = loan.interest_rate or Decimal('0')
        total_rep   = loan.total_repayment_amount or Decimal('0')
        loan_amt    = loan.loan_amount or Decimal('0')
        old_balance = t.old_balance_cleared or Decimal('0')

        if old_balance <= 0:
            continue

        principal, interest = _split_principal_interest(
            old_balance, rate, loan_amt, total_rep
        )

        raw.append({
            'sort_key':     (t.created_at, 1, t.id),
            'date':         t.created_at.date(),
            'receipt_no':   str(t.id).zfill(6),
            'name':         f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'description':  'Clearance loan balance for top-up',
            'rate':         rate,
            'principal':    principal,
            'interest':     interest,
            'total':        old_balance,
            'collected_by': _collector_name(t.processed_by),
        })

    raw.sort(key=lambda e: e['sort_key'])

    prev_date = None
    for idx, entry in enumerate(raw, start=1):
        entry['sn']        = idx
        entry['hide_date'] = (entry['date'] == prev_date)
        prev_date          = entry['date']

    grand_principal = sum(e['principal'] for e in raw)
    grand_interest  = sum(e['interest']  for e in raw)
    grand_total     = sum(e['total']     for e in raw)

    return render(request, 'app/loan_collection_report.html', {
        **base_ctx,
        'rows':              raw,
        'grand_principal':   grand_principal,
        'grand_interest':    grand_interest,
        'grand_total':       grand_total,
        'date_from_display': date_from_display,
        'date_to_display':   date_to_display,
        'branch_name':       branch_name,
    })
    
# ===================================================================================================  
    
    

def _full_name(user):
    if not user:
        return '—'
    full = user.get_full_name().strip()
    return f"{full}" if full else user.username


def bank_transfer_expenses(request):
    return render(request, 'app/bank_transfer_filter.html', {
        **get_base_context(request),
    })


def bank_transfer_expenses_report(request):
    if request.method != 'POST':
        return redirect('bank_transfer_expenses')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('bank_transfer_expenses')

    date_from_display = date_from.strftime('%d/%m/%Y')
    date_to_display   = date_to.strftime('%d/%m/%Y')

    # ── DateTimeField range: start of date_from → end of date_to ─────
    dt_from = datetime.datetime.combine(date_from, datetime.time.min)  # 00:00:00
    dt_to   = datetime.datetime.combine(date_to,   datetime.time.max)  # 23:59:59.999999

    office_txns = OfficeTransaction.objects.filter(
        created_at__range=(dt_from, dt_to),         # ← changed
        transaction_method='bank',
    ).select_related('office_to', 'processed_by').order_by('created_at', 'id')  # ← changed

    if filter_office:
        office_txns = office_txns.filter(office_from=filter_office)

    rows = []
    for txn in office_txns:
        dest_name = txn.office_to.name if txn.office_to else 'headquarter'
        rows.append({
            'txn_id':       txn.id,
            'date':         txn.created_at.date(),          # ← changed
            'receipt_no':   str(txn.id).zfill(6),
            'name':         _full_name(txn.processed_by),
            'description':  f"Amount to Branch [{dest_name}]",
            'amount':       txn.amount,
            'processed_by': _full_name(txn.processed_by),
            'attachment':   txn.attachment,
        })

    return render(request, 'app/bank_transfer_report.html', {
        **base_ctx,
        'rows':              rows,
        'grand_total':       sum(r['amount'] for r in rows),
        'date_from_display': date_from_display,
        'date_to_display':   date_to_display,
        'branch_name':       branch_name,
    })



def delete_office_transaction(request, txn_id):
    if request.method != 'POST':
        return redirect('bank_transfer_expenses')

    txn = get_object_or_404(OfficeTransaction, pk=txn_id)

    with transaction.atomic():
        method = (txn.transaction_method or '').lower()

        # ── Get latest balances for both branches ─────────────────────────
        balance_from = BranchBalance.objects.select_for_update().filter(branch=txn.office_from).last()
        balance_to   = BranchBalance.objects.select_for_update().filter(branch=txn.office_to).last()

        # ── Restore sender's balance ──────────────────────────────────────
        if balance_from:
            BranchBalance.objects.create(
                branch=txn.office_from,
                office_balance=(
                    balance_from.office_balance + txn.amount
                    if method == 'cash'
                    else balance_from.office_balance
                ),
                bank_balance=(
                    balance_from.bank_balance + txn.amount
                    if method == 'bank'
                    else balance_from.bank_balance
                ),
                updated_by=request.user,
            )

        # ── Deduct from receiver's balance ────────────────────────────────
        if balance_to:
            BranchBalance.objects.create(
                branch=txn.office_to,
                office_balance=(
                    balance_to.office_balance - txn.amount
                    if method == 'cash'
                    else balance_to.office_balance
                ),
                bank_balance=(
                    balance_to.bank_balance - txn.amount
                    if method == 'bank'
                    else balance_to.bank_balance
                ),
                updated_by=request.user,
            )

        # ── Delete linked HQTransaction ───────────────────────────────────
        HQTransaction.objects.filter(
            from_branch=txn.office_from,
            to_branch=txn.office_to,
            amount=txn.amount,
        ).order_by('-processed_at').first().delete() if HQTransaction.objects.filter(
            from_branch=txn.office_from,
            to_branch=txn.office_to,
            amount=txn.amount,
        ).exists() else None

        txn.delete()

    messages.success(request, f"Transaction #{str(txn_id).zfill(6)} deleted and balances reversed.")
    return redirect('bank_transfer_expenses')


def bank_cash_transfer(request):
    return render(request, 'app/bank_cash_transfer_filter.html', {
        **get_base_context(request),
    })


def bank_cash_transfer_report(request):
    if request.method != 'POST':
        return redirect('bank_cash_transfer')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('bank_cash_transfer')

    date_from_display = date_from.strftime('%d/%m/%Y')
    date_to_display   = date_to.strftime('%d/%m/%Y')

    # ── DateTimeField range: start of date_from → end of date_to ─────
    dt_from = datetime.datetime.combine(date_from, datetime.time.min)  # 00:00:00
    dt_to   = datetime.datetime.combine(date_to,   datetime.time.max)  # 23:59:59.999999

    txns = BankCashTransaction.objects.filter(
        created_at__range=(dt_from, dt_to),             # ← changed
    ).select_related('office_from').order_by('created_at', 'id')  # ← changed

    if filter_office:
        txns = txns.filter(office_from=filter_office)

    rows = []
    for txn in txns:
        src  = txn.source.lower()
        dest = txn.destination.lower()
        if src == 'bank' and dest == 'cash':
            desc = 'Bank to Cash transfer'
        elif src == 'cash' and dest == 'bank':
            desc = 'Cash to Bank transfer'
        else:
            desc = f"{txn.source.title()} to {txn.destination.title()} transfer"

        rows.append({
            'txn_id':       txn.id,
            'date':         txn.created_at.date(),          # ← changed
            'receipt_no':   str(txn.id).zfill(6),
            'name':         txn.office_from.name if txn.office_from else '—',
            'description':  desc,
            'amount':       txn.amount,
            'processed_by': _full_name(request.user),
            'attachment':   txn.attachment,
        })

    return render(request, 'app/bank_cash_transfer_report.html', {
        **base_ctx,
        'rows':              rows,
        'grand_total':       sum(r['amount'] for r in rows),
        'date_from_display': date_from_display,
        'date_to_display':   date_to_display,
        'branch_name':       branch_name,
    })


def delete_bank_cash_transaction(request, txn_id):
    if request.method != 'POST':
        return redirect('bank_cash_transfer')

    txn = get_object_or_404(BankCashTransaction, pk=txn_id)

    with transaction.atomic():
        balance = BranchBalance.objects.select_for_update().filter(branch=txn.office_from).last()

        if not balance:
            messages.error(request, f'No balance record found for {txn.office_from.name}.')
            return redirect('bank_cash_transfer')

        src  = txn.source.lower()
        dest = txn.destination.lower()

        new_office_balance = balance.office_balance
        new_bank_balance   = balance.bank_balance

        if src == 'cash' and dest == 'bank':
            # Original: cash ↓, bank ↑  →  Reverse: cash ↑, bank ↓
            new_office_balance += txn.amount
            new_bank_balance   -= txn.amount

        elif src == 'bank' and dest == 'cash':
            # Original: bank ↓, cash ↑  →  Reverse: bank ↑, cash ↓
            new_bank_balance   += txn.amount
            new_office_balance -= txn.amount

        BranchBalance.objects.create(
            branch=txn.office_from,
            office_balance=new_office_balance,
            bank_balance=new_bank_balance,
            updated_by=request.user,
        )

        # Delete linked HQTransaction (same office_from == office_to pattern)
        HQTransaction.objects.filter(
            from_branch=txn.office_from,
            to_branch=txn.office_from,
            amount=txn.amount,
        ).order_by('-processed_at').first().delete() if HQTransaction.objects.filter(
            from_branch=txn.office_from,
            to_branch=txn.office_from,
            amount=txn.amount,
        ).exists() else None

        txn.delete()

    messages.success(request, f"Transaction #{str(txn_id).zfill(6)} deleted and balances reversed.")
    return redirect('bank_cash_transfer')



MONTH_NAMES = {
    1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
    7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
}


def loan_issued_filter(request):
    return render(request, 'app/loan_issued_filter.html', {
        **get_base_context(request),
    })


def loan_issued_report(request):
    if request.method != 'POST':
        return redirect('loan_issued_filter')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('loan_issued_filter')

    date_from_display = date_from.strftime('%d/%m/%Y')
    date_to_display   = date_to.strftime('%d/%m/%Y')

    loans = LoanApplication.objects.filter(
        created_at__date__range=(date_from, date_to),
    ).select_related('client', 'processed_by').order_by('created_at', 'id')

    if filter_office:
        loans = loans.filter(office=filter_office.name)

    rows = []
    grand_loan_amount         = Decimal('0.00')
    grand_interest_amount     = Decimal('0.00')
    grand_total_amount        = Decimal('0.00')
    grand_monthly_installment = Decimal('0.00')

    for loan in loans:
        client = loan.client

        if loan.first_repayment_date:
            m = loan.first_repayment_date
            starting_month = f"{MONTH_NAMES[m.month]}/{m.year}"
        else:
            starting_month = '—'

        loan_amount         = loan.loan_amount            or Decimal('0.00')
        interest_amount     = loan.total_interest_amount  or Decimal('0.00')
        total_amount        = loan.total_repayment_amount or Decimal('0.00')
        monthly_installment = loan.monthly_installment    or Decimal('0.00')

        rows.append({
            'loan':                loan,
            'date':                loan.application_date,
            'name':                f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'check_no':            client.checkno or client.employmentcardno or '—',
            'mobile':              client.phonenumber or '—',
            'work_station':        client.kaziyako or client.idara or client.employername or '—',
            'loan_id_label':       f"{(loan.office or 'loan').lower()}-{loan.id}",
            'rate_type':           'Flat',
            'starting_month':      starting_month,
            'loan_amount':         loan_amount,
            'period':              loan.payment_period_months or 0,
            'interest_rate':       loan.interest_rate or Decimal('0.00'),
            'interest_amount':     interest_amount,
            'total_amount':        total_amount,
            'monthly_installment': monthly_installment,
        })

        grand_loan_amount         += loan_amount
        grand_interest_amount     += interest_amount
        grand_total_amount        += total_amount
        grand_monthly_installment += monthly_installment

    return render(request, 'app/loan_issued_report.html', {
        **base_ctx,
        'rows':                      rows,
        'grand_loan_amount':         grand_loan_amount,
        'grand_interest_amount':     grand_interest_amount,
        'grand_total_amount':        grand_total_amount,
        'grand_monthly_installment': grand_monthly_installment,
        'date_from_display':         date_from_display,
        'date_to_display':           date_to_display,
        'branch_name':               branch_name,
    })
    

def loan_issued_report_edit(request, loan_id):
    from decimal import Decimal

    loan   = get_object_or_404(LoanApplication, id=loan_id)
    client = loan.client
    name   = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
    loan_id_label = f"{(loan.office or '').lower()}-{loan.id}"

    if request.method == 'POST':
        try:
            new_amount_str       = request.POST.get('loan_amount', '').replace(',', '')
            new_application_date = request.POST.get('application_date')

            new_amount = Decimal(new_amount_str)

            # Recalculate derived fields
            I = loan.interest_rate
            N = loan.payment_period_months

            from decimal import ROUND_HALF_UP
            from dateutil.relativedelta import relativedelta
            import datetime

            total_interest  = (I / Decimal('100')) * new_amount
            total_repayment = new_amount + total_interest
            monthly         = total_repayment / Decimal(str(N))

            loan.loan_amount            = new_amount
            loan.interest_amount        = total_interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            loan.total_interest_amount  = loan.interest_amount
            loan.total_repayment_amount = total_repayment.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            loan.monthly_installment    = monthly.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

            # Update repayment_amount_remaining by the difference in total_repayment
            old_total = loan.total_repayment_amount  # already updated above
            loan.repayment_amount_remaining = total_repayment.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

            if new_application_date:
                parsed_date = datetime.datetime.strptime(new_application_date, '%Y-%m-%d').date()
                loan.application_date = parsed_date

                # Recalculate first_repayment_date based on new application date
                if parsed_date.day <= 18:
                    loan.first_repayment_date = parsed_date.replace(day=28)
                else:
                    loan.first_repayment_date = (
                        parsed_date + relativedelta(months=1)
                    ).replace(day=28)

            loan.save()
            messages.success(request, f'Loan {loan_id_label} updated successfully.')
            return redirect('loan_issued_report',)

        except Exception as e:
            messages.error(request, f'Error updating loan: {str(e)}')

    return render(request, 'app/loan_issued_report_edit.html', {
        'loan':          loan,
        'name':          name,
        'loan_id_label': loan_id_label,
    })


def loan_report(request):
    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    # loans = LoanApplication.objects.all().select_related(
    #     'client'
    # ).prefetch_related('repayments').order_by(
    #     'client__lastname', 'client__firstname', 'id'
    # )
    loans = LoanApplication.objects.filter(
        is_approved=False,
        repayment_amount_remaining__gt=0,
    ).select_related(
        'client'
    ).prefetch_related('repayments').order_by(
        'client__lastname', 'client__firstname', 'id'
    )

    if filter_office:
        loans = loans.filter(office=filter_office.name)

    rows = []
    grand_loan_amount     = Decimal('0.00')
    grand_interest_amount = Decimal('0.00')
    grand_total_amount    = Decimal('0.00')
    grand_paid_amount     = Decimal('0.00')
    grand_balance         = Decimal('0.00')

    for loan in loans:
        client = loan.client

        paid_amount     = sum(r.repayment_amount for r in loan.repayments.all())
        loan_amount     = loan.loan_amount            or Decimal('0.00')
        interest_amount = loan.total_interest_amount  or Decimal('0.00')
        total_amount    = loan.total_repayment_amount or Decimal('0.00')
        balance         = max(loan.repayment_amount_remaining or Decimal('0.00'), Decimal('0.00'))

        rows.append({
            'name':            f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'check_no':        client.checkno or client.employmentcardno or '—',
            'mobile':          client.phonenumber or '—',
            'loan_type':       loan.loan_type or '—',
            'loan_amount':     loan_amount,
            'interest_amount': interest_amount,
            'total_amount':    total_amount,
            'paid_amount':     paid_amount if paid_amount > 0 else None,
            'balance':         balance,
        })

        grand_loan_amount     += loan_amount
        grand_interest_amount += interest_amount
        grand_total_amount    += total_amount
        grand_paid_amount     += paid_amount
        grand_balance         += balance

    return render(request, 'app/loan_report.html', {
        **base_ctx,
        'rows':                  rows,
        'grand_loan_amount':     grand_loan_amount,
        'grand_interest_amount': grand_interest_amount,
        'grand_total_amount':    grand_total_amount,
        'grand_paid_amount':     grand_paid_amount,
        'grand_balance':         grand_balance,
        'branch_name':           branch_name,
    })
    
    

def no_loan_customers(request):
    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    # Base queryset — scope to selected office via registered_office
    base_qs = Client.objects.all()
    if filter_office:
        base_qs = base_qs.filter(registered_office=filter_office)

    # Clients with NO loans at all
    clients_no_loans = base_qs.filter(
        loan_applications__isnull=True,
    ).distinct()

    # Clients whose ALL loans are fully paid
    clients_all_paid = base_qs.exclude(
        loan_applications__repayment_amount_remaining__gt=0
    ).exclude(
        loan_applications__isnull=True
    ).distinct()

    # Merge and de-duplicate
    all_ids = set(
        list(clients_no_loans.values_list('id', flat=True)) +
        list(clients_all_paid.values_list('id', flat=True))
    )

    clients = Client.objects.filter(
        id__in=all_ids
    ).order_by('lastname', 'firstname')

    return render(request, 'app/no_loan_customers.html', {
        **base_ctx,
        'clients':     clients,
        'branch_name': branch_name,
    })
    
    
    
    
    
MONTH_NAMES = {
    1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
    7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
}


def _month_label(d):
    """Return 'Mon/YYYY' label from a date, e.g. 'Feb/2026'."""
    return f"{MONTH_NAMES[d.month]}/{d.year}"


def _classify_status(expired_days):
    """
    Classify loan status by number of days past end month:
      1 – 30   → Substandard
      31 – 90  → Doubtful
      91+      → Loss
    """
    if expired_days <= 30:
        return 'Substandard'
    elif expired_days <= 90:
        return 'Doubtful'
    else:
        return 'Loss'



def expired_loans(request):
    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    today       = datetime.date.today()
    up_to_label = _month_label(today)

    loans = LoanApplication.objects.filter(
        repayment_amount_remaining__gt=0,
    ).select_related('client').prefetch_related('repayments').order_by(
        'application_date', 'id'
    )

    if filter_office:
        loans = loans.filter(office=filter_office.name)

    raw = []

    for loan in loans:
        if not loan.first_repayment_date or not loan.payment_period_months:
            continue

        periods  = loan.payment_period_months
        end_date = loan.first_repayment_date + relativedelta(months=periods - 1)

        if end_date >= today:
            continue

        expired_days = (today - end_date).days
        if expired_days <= 0:
            continue

        client      = loan.client
        paid_amount = sum(r.repayment_amount for r in loan.repayments.all())
        outstanding = loan.repayment_amount_remaining or Decimal('0.00')

        raw.append({
            'loan':          loan,
            'name':          f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'check_no':      client.checkno or client.employmentcardno or '—',
            'contact':       client.phonenumber or '—',
            'loan_type':     loan.loan_type or '—',
            'decision_date': loan.application_date,
            'start_month':   _month_label(loan.first_repayment_date),
            'end_month':     _month_label(end_date),
            'loaned_amount': loan.loan_amount or Decimal('0.00'),
            'paid_amount':   paid_amount,
            'outstanding':   outstanding,
            'expired_days':  expired_days,
            'status':        _classify_status(expired_days),
        })

    raw.sort(key=lambda r: r['expired_days'])

    return render(request, 'app/expired_loans.html', {
        **base_ctx,
        'rows':              raw,
        'total_loaned':      sum(r['loaned_amount'] for r in raw),
        'total_paid':        sum(r['paid_amount']   for r in raw),
        'total_outstanding': sum(r['outstanding']   for r in raw),
        'branch_name':       branch_name,
        'up_to_label':       up_to_label.upper(),
    })
    
    
    
    
    

MONTH_NAMES = {
    1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
    7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
}


def _month_label(d):
    return f"{MONTH_NAMES[d.month]}/{d.year}"


def monthly_outstanding_filter(request):
    return render(request, 'app/monthly_outstanding_filter.html', {
        **get_base_context(request),
    })


def monthly_outstanding_report(request):
    if request.method != 'POST':
        return redirect('monthly_outstanding_filter')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    selected_str = request.POST.get('selected_month', '')
    try:
        selected_date = datetime.datetime.strptime(selected_str, '%Y-%m-%d').date()
    except ValueError:
        return redirect('monthly_outstanding_filter')

    sel_year  = selected_date.year
    sel_month = selected_date.month

    last_day    = calendar.monthrange(sel_year, sel_month)[1]
    month_start = datetime.date(sel_year, sel_month, 1)
    month_end   = datetime.date(sel_year, sel_month, last_day)
    month_label = _month_label(selected_date).upper()

    loans = LoanApplication.objects.filter(
        repayment_amount_remaining__gt=0,
    ).select_related('client').prefetch_related('repayments').order_by(
        'client__lastname', 'client__firstname', 'id'
    )

    if filter_office:
        loans = loans.filter(office=filter_office.name)

    rows = []

    for loan in loans:
        if not loan.first_repayment_date or not loan.payment_period_months:
            continue
        if not loan.monthly_installment:
            continue

        periods          = loan.payment_period_months
        first_repay_date = loan.first_repayment_date
        monthly_inst     = loan.monthly_installment

        schedule_months = [
            ((first_repay_date + relativedelta(months=i)).year,
             (first_repay_date + relativedelta(months=i)).month)
            for i in range(periods)
        ]

        slots_in_month = sum(
            1 for (y, m) in schedule_months if y == sel_year and m == sel_month
        )

        if slots_in_month == 0:
            continue

        amount_to_be_paid = monthly_inst * slots_in_month

        paid_this_month = sum(
            r.repayment_amount or Decimal('0.00')
            for r in loan.repayments.all()
            if r.repayment_date and month_start <= r.repayment_date <= month_end
        )

        not_paid    = max(amount_to_be_paid - paid_this_month, Decimal('0.00'))
        outstanding = loan.repayment_amount_remaining or Decimal('0.00')
        client      = loan.client

        rows.append({
            'name':              f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'check_no':          client.checkno or client.employmentcardno or '—',
            'contact':           client.phonenumber or '—',
            'amount_to_be_paid': amount_to_be_paid,
            'paid_this_month':   paid_this_month,
            'not_paid':          not_paid,
            'outstanding_total': outstanding,
        })

    return render(request, 'app/monthly_outstanding_report.html', {
        **base_ctx,
        'rows':                  rows,
        'month_label':           month_label,
        'branch_name':           branch_name,
        'total_amount_to_pay':   sum(r['amount_to_be_paid'] for r in rows),
        'total_paid_this_month': sum(r['paid_this_month']   for r in rows),
        'total_not_paid':        sum(r['not_paid']          for r in rows),
        'total_outstanding':     sum(r['outstanding_total'] for r in rows),
    })
    
    
    
    
    

def expenses_filter(request):
    return render(request, 'app/expenses_filter.html', {
        **get_base_context(request),
    })


def expenses_report(request):
    if request.method != 'POST':
        return redirect('expenses_filter')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('expenses_filter')

    date_from_display = date_from.strftime('%d/%m/%Y')
    date_to_display   = date_to.strftime('%d/%m/%Y')

    expenses = Expense.objects.filter(
        transaction_date__range=(date_from, date_to),
    ).select_related(
        'transaction_type',
        'recorded_by',
    ).order_by('transaction_date', 'id')

    if filter_office:
        expenses = expenses.filter(office=filter_office.name)

    rows      = []
    prev_date = None

    for exp in expenses:
        category    = exp.transaction_type.name if exp.transaction_type else 'Expense'
        description = (exp.description or '').strip()
        is_bank     = (getattr(exp, 'payment_method', 'cash') or 'cash').lower() == 'bank'

        rows.append({
            'date':        exp.transaction_date,
            'receipt_no':  str(exp.id).zfill(6),
            'category':    category,
            'description': description,
            'amount':      exp.amount or Decimal('0.00'),
            'is_bank':     is_bank,
            'attachment':  exp.attachment.url if exp.attachment else None,
            'hide_date':   (exp.transaction_date == prev_date),
        })
        prev_date = exp.transaction_date

    return render(request, 'app/expenses_report.html', {
        **base_ctx,
        'rows':              rows,
        'grand_total':       sum(r['amount'] for r in rows),
        'date_from_display': date_from_display,
        'date_to_display':   date_to_display,
        'branch_name':       branch_name,
    })
    
    
    
    
    
    
def financial_statement_filter(request):
    return render(request, 'app/financial_statement_filter.html', {
        **get_base_context(request),
    })


def financial_statement_report(request):
    if request.method != 'POST':
        return redirect('financial_statement_filter')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    try:
        start_date = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        return redirect('financial_statement_filter')

    date_from_display = start_date.strftime('%d/%m/%Y')
    date_to_display   = end_date.strftime('%d/%m/%Y')

    # Convert dates to datetime range for DateTimeField compatibility
    start_dt = datetime.datetime.combine(start_date, datetime.time.min)  # 00:00:00
    end_dt   = datetime.datetime.combine(end_date,   datetime.time.max)  # 23:59:59.999999

    # ── Scope helpers ─────────────────────────────────────────────────────────────
    def rep_qs(q):
        qs = LoanRepayment.objects.filter(q)
        if filter_office:
            qs = qs.filter(loan_application__office__iexact=filter_office.name)
        return qs

    def loan_qs(q):
        qs = LoanApplication.objects.filter(q)
        if filter_office:
            qs = qs.filter(office__iexact=filter_office.name)
        return qs

    def nyo_qs(q):
        qs = Nyongeza.objects.filter(q)
        if filter_office:
            qs = qs.filter(Office=filter_office)
        return qs

    def exp_qs(q):
        qs = Expense.objects.filter(q)
        if filter_office:
            qs = qs.filter(office__iexact=filter_office.name)
        return qs

    def transfer_in_qs(q):
        qs = OfficeTransaction.objects.filter(q)
        if filter_office:
            qs = qs.filter(office_to=filter_office)
        return qs

    def transfer_out_qs(q):
        qs = OfficeTransaction.objects.filter(q)
        if filter_office:
            qs = qs.filter(office_from=filter_office)
        return qs

    def transfer_out_non_hq_qs(q):
        qs = OfficeTransaction.objects.filter(q)
        if filter_office:
            qs = qs.filter(office_from=filter_office)
        return qs.exclude(office_to__name__iexact='HQ')

    def transfer_out_hq_qs(q):
        qs = OfficeTransaction.objects.filter(q)
        if filter_office:
            qs = qs.filter(office_from=filter_office)
        return qs.filter(office_to__name__iexact='HQ')

    def bank_charge_qs(q):
        qs = BankCharge.objects.filter(q)
        if filter_office:
            qs = qs.filter(office__iexact=filter_office.name)
        return qs

    # ── Opening Balance (everything BEFORE start_date) ────────────────────────────
    # rep_b          = rep_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')
    # nyo_b          = nyo_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']           or Decimal('0')
    # transfer_in_b  = transfer_in_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']   or Decimal('0')
    rep_b         = rep_qs(Q(created_at__lt=start_dt)).exclude(loan_application__loan_type__iexact='Hazina').aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')
    hazina_rep_b  = rep_qs(Q(created_at__lt=start_dt)).filter(loan_application__loan_type__iexact='Hazina').aggregate(t=Sum('repayment_amount'))['t']  or Decimal('0')
    nyo_b         = nyo_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']           or Decimal('0')
    transfer_in_b = transfer_in_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']   or Decimal('0')

    exp_b          = exp_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']           or Decimal('0')
    loan_b         = loan_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('loan_amount'))['t']     or Decimal('0')
    transfer_out_b = transfer_out_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']  or Decimal('0')
    bank_charge_b  = bank_charge_qs(Q(created_at__lt=start_dt)).aggregate(t=Sum('amount'))['t']  or Decimal('0')

    # opening_stock = (rep_b + nyo_b + transfer_in_b) - (exp_b + loan_b + transfer_out_b + bank_charge_b)
    opening_stock = (rep_b + hazina_rep_b + nyo_b + transfer_in_b) - (exp_b + loan_b + transfer_out_b + bank_charge_b)

    # ── Period date filters (all using created_at) ────────────────────────────────
    pq = Q(created_at__gte=start_dt, created_at__lte=end_dt)  # one Q reused for all models

    # ── INCOME (Inflows) ────────────────────────────────────────────────────────── BEFORE
    # total_mapato       = rep_qs(pq).aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')
    # total_nyongeza     = nyo_qs(pq).aggregate(t=Sum('amount'))['t']           or Decimal('0')
    # total_hazina       = Decimal('0')
    # total_transfers_in = transfer_in_qs(pq).aggregate(t=Sum('amount'))['t']  or Decimal('0')
    
    # ── INCOME (Inflows) ────────────────────────────────────────────────────────── AFTER
    total_mapato       = rep_qs(pq).exclude(loan_application__loan_type__iexact='Hazina').aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')
    total_nyongeza     = nyo_qs(pq).aggregate(t=Sum('amount'))['t']           or Decimal('0')
    total_hazina       = rep_qs(pq).filter(loan_application__loan_type__iexact='Hazina').aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')
    total_transfers_in = transfer_in_qs(pq).aggregate(t=Sum('amount'))['t']  or Decimal('0')

    total_income_with_opening = (
        opening_stock
        + total_mapato
        + total_nyongeza
        + total_hazina
        + total_transfers_in
    )

    # ── EXPENDITURE ───────────────────────────────────────────────────────────────
    period_expenses = exp_qs(pq).select_related('transaction_type')
    period_loans    = loan_qs(pq)

    # ── 0. FOMU (Loan Disbursements) ──────────────────────────────────────────────
    total_loans_disbursed = period_loans.aggregate(t=Sum('loan_amount'))['t']                                   or Decimal('0')
    loan_cash_amount      = period_loans.filter(transaction_method='cash').aggregate(t=Sum('loan_amount'))['t'] or Decimal('0')
    loan_bank_amount      = period_loans.filter(transaction_method='bank').aggregate(t=Sum('loan_amount'))['t'] or Decimal('0')

    # ── 1. MATUMIZI OFISINI ───────────────────────────────────────────────────────
    office_expense_categories = ExpenseCategory.objects.all().order_by('name')

    matumizi_ofisini_rows = []

    for cat in office_expense_categories:
        cat_exps    = period_expenses.filter(transaction_type=cat)
        cash_amount = cat_exps.filter(payment_method='cash').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        bank_amount = cat_exps.filter(payment_method='bank').aggregate(t=Sum('amount'))['t'] or Decimal('0')
        total       = cash_amount + bank_amount
        if total > 0:
            matumizi_ofisini_rows.append({
                'name':        cat.name,
                'cash_amount': cash_amount if cash_amount > 0 else None,
                'bank_amount': bank_amount if bank_amount > 0 else None,
                'total':       total,
            })

    uncategorised = period_expenses.filter(transaction_type__isnull=True)
    unc_cash  = uncategorised.filter(payment_method='cash').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    unc_bank  = uncategorised.filter(payment_method='bank').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    unc_total = unc_cash + unc_bank
    if unc_total > 0:
        matumizi_ofisini_rows.append({
            'name':        'Matumizi Mengineyo',
            'cash_amount': unc_cash if unc_cash > 0 else None,
            'bank_amount': unc_bank if unc_bank > 0 else None,
            'total':       unc_total,
        })

    total_matumizi_ofisini = sum(r['total'] for r in matumizi_ofisini_rows)

    # ── 2. MATUMIZI BENKI-[KITUO] ─────────────────────────────────────────────────
    transfers_kituo      = transfer_out_non_hq_qs(pq).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_matumizi_kituo = transfers_kituo

    # ── 3. MATUMIZI BENKI-[MKURUGENZI] ───────────────────────────────────────────
    total_matumizi_mkurugenzi = transfer_out_hq_qs(pq).aggregate(t=Sum('amount'))['t'] or Decimal('0')

    # ── 4. MAKATO BANK ────────────────────────────────────────────────────────────
    period_bank_charges = bank_charge_qs(pq)
    makato_benki_cash   = period_bank_charges.filter(payment_method='cash').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    makato_benki_bank   = period_bank_charges.filter(payment_method='bank').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_makato_benki  = makato_benki_cash + makato_benki_bank

    # ── Grand total outflow ───────────────────────────────────────────────────────
    total_outflow_current = (
        total_loans_disbursed
        + total_matumizi_ofisini
        + total_matumizi_kituo
        + total_matumizi_mkurugenzi
        + total_makato_benki
    )

    # ── Live closing balances (unchanged) ─────────────────────────────────────────
    if filter_office:
        latest = BranchBalance.objects.filter(
            branch=filter_office
        ).order_by('-last_updated').first()
        cash_in_office = latest.office_balance if latest else Decimal('0')
        cash_in_bank   = latest.bank_balance   if latest else Decimal('0')
    else:
        cash_in_office = Decimal('0')
        cash_in_bank   = Decimal('0')
        for office in Office.objects.all():
            latest = BranchBalance.objects.filter(branch=office).order_by('-last_updated').first()
            if latest:
                cash_in_office += latest.office_balance
                cash_in_bank   += latest.bank_balance

    return render(request, 'app/financial_statement_report.html', {
        **base_ctx,
        'branch_name':       branch_name,
        'date_from_display': date_from_display,
        'date_to_display':   date_to_display,

        # ── Opening balance
        'opening_stock': opening_stock,

        # ── Income
        'total_mapato':              total_mapato       if total_mapato       else None,
        'total_nyongeza':            total_nyongeza     if total_nyongeza     else None,
        'total_hazina':              total_hazina       if total_hazina       else None,
        'total_transfers_in':        total_transfers_in if total_transfers_in else None,
        'total_income_with_opening': total_income_with_opening,

        # ── Expenditure component 0: FOMU (Loan Disbursements)
        'total_loans_disbursed': total_loans_disbursed if total_loans_disbursed else None,
        'loan_cash_amount':      loan_cash_amount      if loan_cash_amount      else None,
        'loan_bank_amount':      loan_bank_amount      if loan_bank_amount      else None,

        # ── Expenditure component 1: Matumizi Ofisini (expenses only, no loans)
        'matumizi_ofisini_rows':  matumizi_ofisini_rows,
        'total_matumizi_ofisini': total_matumizi_ofisini if total_matumizi_ofisini else None,

        # ── Expenditure component 2: Matumizi Benki-[Kituo]
        'transfers_kituo':      transfers_kituo      if transfers_kituo      else None,
        'total_matumizi_kituo': total_matumizi_kituo if total_matumizi_kituo else None,

        # ── Expenditure component 3: Matumizi Benki-[Mkurugenzi]
        'total_matumizi_mkurugenzi': total_matumizi_mkurugenzi if total_matumizi_mkurugenzi else None,

        # ── Expenditure component 4: Makato Bank
        'makato_benki_cash':  makato_benki_cash  if makato_benki_cash  else None,
        'makato_benki_bank':  makato_benki_bank  if makato_benki_bank  else None,
        'total_makato_benki': total_makato_benki if total_makato_benki else None,

        # ── Grand total outflow
        'total_outflow_current': total_outflow_current,

        # ── Closing balances
        'cash_in_office': cash_in_office,
        'cash_in_bank':   cash_in_bank,
        'total_cash':     cash_in_office + cash_in_bank,
    })
    
 
   
    
def monthly_repayment_filter(request):
    return render(request, 'app/monthly_repayment_filter.html', {
        **get_base_context(request),
    })
    
MONTH_NAMES = {
    1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
    7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
}


def _parse_month(mm_yyyy):
    """Parse 'MM-YYYY' → (month:int, year:int)."""
    parts = mm_yyyy.strip().split('-')
    if len(parts) != 2:
        return None, None
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None, None






# def monthly_repayment_report(request):
#     if request.method != 'POST':
#         return redirect('monthly_repayment_filter')

#     base_ctx        = get_base_context(request)
#     filter_office   = base_ctx['filter_office']
#     selected_office = base_ctx['selected_office']
#     branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

#     month_from_str = request.POST.get('month_from', '')
#     month_to_str   = request.POST.get('month_to',   '')

#     from_month, from_year = _parse_month(month_from_str)
#     to_month,   to_year   = _parse_month(month_to_str)

#     if not from_month or not to_month:
#         return redirect('monthly_repayment_filter')

#     period_months = []
#     y, m = from_year, from_month
#     while (y < to_year) or (y == to_year and m <= to_month):
#         period_months.append((y, m))
#         m += 1
#         if m > 12:
#             m = 1
#             y += 1

#     months_data = []

#     for (year, month) in period_months:
#         last_day    = calendar.monthrange(year, month)[1]
#         month_start = datetime.date(year, month, 1)
#         month_end   = datetime.date(year, month, last_day)
#         label       = f"{MONTH_NAMES[month]}-{year}"

#         raw = []

#         # ── LoanRepayment records ─────────────────────────────────
#         repayments = LoanRepayment.objects.filter(
#             payment_month__range=(month_start, month_end),
#         ).select_related(
#             'loan_application', 'loan_application__client', 'processed_by'
#         ).order_by('payment_month', 'id')

#         if filter_office:
#             repayments = repayments.filter(loan_application__office=filter_office.name)

#         for rep in repayments:
#             loan   = rep.loan_application
#             client = loan.client
#             raw.append({
#                 'sort_key':      (rep.payment_month, 0, rep.id),
#                 'date':          rep.repayment_date,
#                 'receipt_no':    str(rep.id).zfill(6),
#                 'name':          f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
#                 'check_no':      client.checkno or client.employmentcardno or '—',
#                 'loan_id_label': f"{(loan.office or 'loan').lower()}-{loan.id}",
#                 'description':   'Loan payment',
#                 'amount':        rep.repayment_amount or Decimal('0'),
#                 'row_key':       f"repayment-{rep.id}",
#                 'record_type':   'repayment',
#                 'record_id':     rep.id,
#             })

#         # ── LoanTopup records ─────────────────────────────────────
#         topups = LoanTopup.objects.filter(
#             payment_month__range=(month_start, month_end),
#         ).select_related(
#             'loan_application', 'loan_application__client', 'processed_by'
#         ).order_by('payment_month', 'id')

#         if filter_office:
#             topups = topups.filter(loan_application__office=filter_office.name)

#         for topup in topups:
#             loan    = topup.loan_application
#             client  = loan.client
#             cleared = topup.old_balance_cleared or Decimal('0')
#             raw.append({
#                 'sort_key':      (topup.payment_month, 1, topup.id),
#                 'date':          topup.topup_date,
#                 'receipt_no':    str(topup.id).zfill(6),
#                 'name':          f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
#                 'check_no':      client.checkno or client.employmentcardno or '—',
#                 'loan_id_label': f"{(loan.office or 'loan').lower()}-{loan.id}",
#                 'description':   'Clearance loan balance for top-up',
#                 'amount':        cleared,
#                 'row_key':       f"topup-{topup.id}",
#                 'record_type':   'topup',
#                 'record_id':     topup.id,
#             })

#         raw.sort(key=lambda r: r['sort_key'])

#         months_data.append({
#             'label':       label,
#             'rows':        raw,
#             'grand_total': sum(r['amount'] for r in raw),
#         })

#     return render(request, 'app/monthly_repayment_report.html', {
#         **base_ctx,
#         'branch_name': branch_name,
#         'months':      months_data,
#     })


def monthly_repayment_report(request):
    if request.method != 'POST':
        return redirect('monthly_repayment_filter')

    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    month_from_str = request.POST.get('month_from', '')
    month_to_str   = request.POST.get('month_to',   '')

    from_month, from_year = _parse_month(month_from_str)
    to_month,   to_year   = _parse_month(month_to_str)

    if not from_month or not to_month:
        return redirect('monthly_repayment_filter')

    # Full date range based on transaction date (repayment_date / topup_date)
    range_start = datetime.date(from_year, from_month, 1)
    range_end   = datetime.date(to_year, to_month, calendar.monthrange(to_year, to_month)[1])

    # ── LoanRepayment: filter by transaction date, group by payment_month ──
    repayments = LoanRepayment.objects.filter(
        repayment_date__range=(range_start, range_end),
    ).select_related(
        'loan_application', 'loan_application__client', 'processed_by'
    ).order_by('payment_month', 'id')

    if filter_office:
        repayments = repayments.filter(loan_application__office=filter_office.name)

    # ── LoanTopup: filter by transaction date, group by payment_month ──
    topups = LoanTopup.objects.filter(
        topup_date__range=(range_start, range_end),
    ).select_related(
        'loan_application', 'loan_application__client', 'processed_by'
    ).order_by('payment_month', 'id')

    if filter_office:
        topups = topups.filter(loan_application__office=filter_office.name)

    # ── Group into buckets by payment_month ───────────────────────────────
    buckets = {}  # key: (year, month) → list of row dicts

    for rep in repayments:
        loan   = rep.loan_application
        client = loan.client
        key    = (rep.payment_month.year, rep.payment_month.month)
        buckets.setdefault(key, []).append({
            'sort_key':      (rep.payment_month, 0, rep.id),
            'date':          rep.repayment_date,
            'receipt_no':    str(rep.id).zfill(6),
            'name':          f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'check_no':      client.checkno or client.employmentcardno or '—',
            'loan_id_label': f"{(loan.office or 'loan').lower()}-{loan.id}",
            'description':   'Loan payment',
            'amount':        rep.repayment_amount or Decimal('0'),
            'row_key':       f"repayment-{rep.id}",
            'record_type':   'repayment',
            'record_id':     rep.id,
        })

    for topup in topups:
        loan    = topup.loan_application
        client  = loan.client
        cleared = topup.old_balance_cleared or Decimal('0')
        key     = (topup.payment_month.year, topup.payment_month.month)
        buckets.setdefault(key, []).append({
            'sort_key':      (topup.payment_month, 1, topup.id),
            'date':          topup.topup_date,
            'receipt_no':    str(topup.id).zfill(6),
            'name':          f"{client.firstname} {client.middlename or ''} {client.lastname}".strip(),
            'check_no':      client.checkno or client.employmentcardno or '—',
            'loan_id_label': f"{(loan.office or 'loan').lower()}-{loan.id}",
            'description':   'Clearance loan balance for top-up',
            'amount':        cleared,
            'row_key':       f"topup-{topup.id}",
            'record_type':   'topup',
            'record_id':     topup.id,
        })

    # ── Build months_data sorted by payment_month ─────────────────────────
    months_data = []
    for key in sorted(buckets.keys()):
        year, month = key
        rows = sorted(buckets[key], key=lambda r: r['sort_key'])
        months_data.append({
            'label':       f"{MONTH_NAMES[month]}-{year}",
            'rows':        rows,
            'grand_total': sum(r['amount'] for r in rows),
        })

    return render(request, 'app/monthly_repayment_report.html', {
        **base_ctx,
        'branch_name': branch_name,
        'months':      months_data,
    })

# def edit_repayment(request, repayment_type, repayment_id):
#     """Single-record edit page (image 3)."""
#     from django.shortcuts import get_object_or_404

#     if repayment_type == 'repayment':
#         obj = get_object_or_404(LoanRepayment, id=repayment_id)
#         client = obj.loan_application.client
#         name   = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
#         receipt_no       = str(obj.id).zfill(6)
#         amount           = obj.repayment_amount
#         transaction_method = obj.transaction_method or 'cash'
#         payment_month    = obj.payment_month
#         repayment_date   = obj.repayment_date
#     else:  # topup
#         obj = get_object_or_404(LoanTopup, id=repayment_id)
#         client = obj.loan_application.client
#         name   = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
#         receipt_no       = str(obj.id).zfill(6)
#         amount           = obj.old_balance_cleared
#         transaction_method = obj.transaction_method or 'cash'
#         payment_month    = obj.payment_month
#         repayment_date   = obj.topup_date

#     if request.method == 'POST':
#         new_payment_month    = request.POST.get('payment_month')
#         new_transaction_method = request.POST.get('transaction_method', transaction_method)
#         new_repayment_date   = request.POST.get('transaction_date')

#         try:
#             if new_payment_month:
#                 obj.payment_month = new_payment_month
#             if new_transaction_method:
#                 obj.transaction_method = new_transaction_method
#             if new_repayment_date:
#                 if repayment_type == 'repayment':
#                     obj.repayment_date = new_repayment_date
#                 else:
#                     obj.topup_date = new_repayment_date
#             obj.save()
#             messages.success(request, 'Record updated successfully.')
#         except Exception as e:
#             messages.error(request, f'Error: {str(e)}')

#         return redirect('monthly_repayment_filter')

#     return render(request, 'app/edit_repayment.html', {
#         **get_base_context(request),
#         'repayment_type':    repayment_type,
#         'repayment_id':      repayment_id,
#         'name':              name,
#         'receipt_no':        receipt_no,
#         'amount':            amount,
#         'transaction_method': transaction_method,
#         'payment_month':     payment_month,
#         'repayment_date':    repayment_date,
#     })


def edit_repayment(request, repayment_type, repayment_id):
    """Single-record edit page."""
    from django.shortcuts import get_object_or_404

    if repayment_type == 'repayment':
        obj = get_object_or_404(LoanRepayment, id=repayment_id)
        client = obj.loan_application.client
        name               = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
        receipt_no         = str(obj.id).zfill(6)
        amount             = obj.repayment_amount
        transaction_method = obj.transaction_method or 'cash'
        payment_month      = obj.payment_month
        repayment_date     = obj.repayment_date

    else:  # topup
        obj = get_object_or_404(LoanTopup, id=repayment_id)
        client = obj.loan_application.client
        name               = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
        receipt_no         = str(obj.id).zfill(6)
        amount             = obj.old_balance_cleared
        transaction_method = obj.transaction_method or 'cash'
        payment_month      = obj.payment_month
        repayment_date     = obj.topup_date

    if request.method == 'POST':
        new_payment_month      = request.POST.get('payment_month')
        new_transaction_method = request.POST.get('transaction_method', transaction_method)
        new_repayment_date     = request.POST.get('transaction_date')

        try:
            if repayment_type == 'topup':
                # Find linked LoanRepayment BEFORE we change payment_month on obj
                # Try matching by payment_month first (for records that have it)
                linked_repayment = LoanRepayment.objects.filter(
                    loan_application=obj.loan_application,
                    payment_month=obj.payment_month,
                ).first()

                # Fall back to latest repayment if payment_month was NULL
                if not linked_repayment:
                    linked_repayment = LoanRepayment.objects.filter(
                        loan_application=obj.loan_application,
                    ).order_by('-created_at').first()

            # ── Update the main object (LoanRepayment or LoanTopup) ──
            if new_payment_month:
                obj.payment_month = new_payment_month
            if new_transaction_method:
                obj.transaction_method = new_transaction_method
            if new_repayment_date:
                if repayment_type == 'repayment':
                    obj.repayment_date = new_repayment_date
                else:
                    obj.topup_date = new_repayment_date
            obj.save()

            # ── Sync payment_month AND repayment_date to linked LoanRepayment ──
            if repayment_type == 'topup' and new_payment_month:
                if linked_repayment:
                    # Update existing linked repayment
                    linked_repayment.payment_month  = new_payment_month
                    linked_repayment.repayment_date = new_payment_month
                    linked_repayment.save()
                else:
                    # No repayment found at all — create one linked to this topup's loan
                    LoanRepayment.objects.create(
                        loan_application   = obj.loan_application,
                        repayment_amount   = obj.old_balance_cleared,
                        repayment_date     = new_payment_month,
                        payment_month      = new_payment_month,
                        transaction_method = new_transaction_method or obj.transaction_method,
                        processed_by       = obj.processed_by,
                    )

            messages.success(request, 'Record updated successfully.')

        except Exception as e:
            messages.error(request, f'Error: {str(e)}')

        return redirect('monthly_repayment_filter')

    return render(request, 'app/edit_repayment.html', {
        **get_base_context(request),
        'repayment_type':     repayment_type,
        'repayment_id':       repayment_id,
        'name':               name,
        'receipt_no':         receipt_no,
        'amount':             amount,
        'transaction_method': transaction_method,
        'payment_month':      payment_month,
        'repayment_date':     repayment_date,
    })


# def bulk_update_payment_month(request):
#     if request.method != 'POST':
#         return redirect('monthly_repayment_filter')

#     selected_items   = request.POST.getlist('selected_items')
#     transaction_date = request.POST.get('transaction_date') or None

#     if not selected_items or not transaction_date:
#         messages.error(request, 'Please select records and fill the transaction date.')
#         return redirect(request.META.get('HTTP_REFERER', 'monthly_repayment_filter'))

#     updated = 0
#     for item in selected_items:
#         try:
#             rtype, rid = item.split('-', 1)
#             if rtype == 'repayment':
#                 LoanRepayment.objects.filter(id=int(rid)).update(
#                     payment_month=transaction_date 
#                 )
#             elif rtype == 'topup':
#                 LoanTopup.objects.filter(id=int(rid)).update(
#                     payment_month=transaction_date
#                 )
#             updated += 1
#         except Exception:
#             continue

#     messages.success(request, f'{updated} record(s) updated successfully.')
#     return redirect(request.META.get('HTTP_REFERER', 'monthly_repayment_filter'))

def bulk_update_payment_month(request):
    if request.method != 'POST':
        return redirect('monthly_repayment_filter')

    selected_items   = request.POST.getlist('selected_items')
    transaction_date = request.POST.get('transaction_date') or None

    if not selected_items or not transaction_date:
        messages.error(request, 'Please select records and fill the transaction date.')
        return redirect(request.META.get('HTTP_REFERER', 'monthly_repayment_filter'))

    updated = 0
    for item in selected_items:
        try:
            rtype, rid = item.split('-', 1)

            if rtype == 'repayment':
                LoanRepayment.objects.filter(id=int(rid)).update(
                    payment_month=transaction_date
                )
                updated += 1

            elif rtype == 'topup':
                topup = LoanTopup.objects.select_related(
                    'loan_application', 'processed_by'
                ).filter(id=int(rid)).first()

                if not topup:
                    continue

                # Find linked LoanRepayment BEFORE updating payment_month
                linked_repayment = LoanRepayment.objects.filter(
                    loan_application=topup.loan_application,
                    payment_month=topup.payment_month,
                ).first()

                # Fall back to latest repayment if payment_month was NULL
                if not linked_repayment:
                    linked_repayment = LoanRepayment.objects.filter(
                        loan_application=topup.loan_application,
                    ).order_by('-created_at').first()

                # Update the topup
                topup.payment_month = transaction_date
                topup.save()

                # Sync to linked LoanRepayment
                if linked_repayment:
                    linked_repayment.payment_month  = transaction_date
                    linked_repayment.repayment_date = transaction_date
                    linked_repayment.save()
                else:
                    # No repayment found — create one
                    LoanRepayment.objects.create(
                        loan_application   = topup.loan_application,
                        repayment_amount   = topup.old_balance_cleared,
                        repayment_date     = transaction_date,
                        payment_month      = transaction_date,
                        transaction_method = topup.transaction_method,
                        processed_by       = topup.processed_by,
                    )

                updated += 1

        except Exception:
            continue

    messages.success(request, f'{updated} record(s) updated successfully.')
    return redirect(request.META.get('HTTP_REFERER', 'monthly_repayment_filter'))
# ============================================================================================  
    
    

def expense_category_list(request):
    """List all expense categories."""
    categories = ExpenseCategory.objects.all().order_by('name')
    return render(request, 'app/expense_category.html', {
        **get_base_context(request),
        'categories': categories,
    })


def expense_category_add(request):
    """Add a new expense category."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Category name is required.')
            return redirect('expense_category_list')

        if ExpenseCategory.objects.filter(name__iexact=name).exists():
            messages.error(request, f'Category "{name}" already exists.')
            return redirect('expense_category_list')

        ExpenseCategory.objects.create(name=name)
        messages.success(request, f'Expense category "{name}" added successfully.')
        return redirect('expense_category_list')

    return redirect('expense_category_list')


def expense_category_update(request, pk):
    """Update an existing expense category."""
    category = get_object_or_404(ExpenseCategory, pk=pk)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'Category name is required.')
            return redirect('expense_category_list')

        # Check for duplicate (excluding current record)
        if ExpenseCategory.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f'Category "{name}" already exists.')
            return redirect('expense_category_list')

        category.name = name
        category.save()
        messages.success(request, f'Expense category updated to "{name}" successfully.')
        return redirect('expense_category_list')

    return redirect('expense_category_list')


def expense_category_delete(request, pk):
    """Delete an expense category."""
    category = get_object_or_404(ExpenseCategory, pk=pk)

    if request.method == 'POST':
        name = category.name
        category.delete()
        messages.success(request, f'Expense category "{name}" deleted successfully.')
        return redirect('expense_category_list')

    return redirect('expense_category_list')


def expense_category_detail(request, pk):
    """Return category details as JSON (for AJAX view)."""
    category = get_object_or_404(ExpenseCategory, pk=pk)
    data = {
        'id': category.id,
        'name': category.name,
        'created_at': category.created_at.strftime('%Y-%m-%d'),
    }
    return JsonResponse(data)




from django.views.decorators.http import require_POST
def staff_salary_list(request):
    """
    Display all staff with their current salary and deduction amounts.
    """
    staff_list = CustomUser.objects.filter(
        is_active=True
    ).exclude(
        is_superuser=True
    ).order_by('first_name', 'last_name')

    context = {
        **get_base_context(request),
        'staff_list': staff_list,
        'total_staff': staff_list.count(),
    }
    return render(request, 'app/staff_salary_list.html', context)


def staff_salary_setting(request):
    """
    Display editable table for setting staff basic salary and deduction amounts.
    """
    staff_list = CustomUser.objects.filter(
        is_active=True
    ).exclude(
        is_superuser=True
    ).order_by('first_name', 'last_name')

    context = {
        **get_base_context(request), 
        'staff_list': staff_list,
        'total_staff': staff_list.count(),
    }
    return render(request, 'app/staff_salary_setting.html', context)


@require_POST
def staff_salary_update(request):
    """
    Handle bulk salary and deduction update for all staff.
    Supports both AJAX (returns JSON) and standard form submit.
    """
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    updated_count = 0
    errors = []

    staff_list = CustomUser.objects.filter(
        is_active=True
    ).exclude(is_superuser=True)

    for staff in staff_list:
        salary_key = f'salary_{staff.id}'
        deduction_key = f'deduction_{staff.id}'

        salary_val = request.POST.get(salary_key)
        deduction_val = request.POST.get(deduction_key)

        changed = False

        if salary_val is not None:
            try:
                new_salary = float(salary_val)
                if staff.salary != new_salary:
                    staff.salary = new_salary
                    changed = True
            except (ValueError, TypeError):
                errors.append(f'Invalid salary value for {staff.get_full_name()}')

        if deduction_val is not None:
            try:
                new_deduction = float(deduction_val)
                if staff.deduction_amount != new_deduction:
                    staff.deduction_amount = new_deduction
                    changed = True
            except (ValueError, TypeError):
                errors.append(f'Invalid deduction value for {staff.get_full_name()}')

        if changed:
            staff.save(update_fields=['salary', 'deduction_amount'])
            updated_count += 1

    if is_ajax:
        if errors:
            return JsonResponse({'success': False, 'errors': errors})
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} staff record(s) updated successfully.'
        })

    # Standard form submit fallback
    if errors:
        messages.error(request, f'Some errors occurred: {"; ".join(errors)}')
    else:
        messages.success(
            request,
            f'Salaries updated successfully. {updated_count} record(s) changed.'
        )
    return redirect('staff_salary_list')


def staff_salary_detail_json(request, pk):
    """
    Return a single staff member's salary info as JSON (for AJAX use).
    """
    try:
        staff = CustomUser.objects.get(pk=pk, is_active=True)
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Staff not found'}, status=404)

    return JsonResponse({
        'id': staff.id,
        'name': staff.get_full_name() or staff.username,
        'role': staff.role or '',
        'salary': float(staff.salary or 0),
        'deduction_amount': float(staff.deduction_amount or 0),
        'employee_id': staff.employee_id or '',
    })
    
    
    
    
def salary_advance_list(request):
    """List all salary advances."""
    advances = SalaryAdvance.objects.select_related('employee').all()
    context = {
        **get_base_context(request),
        'advances': advances,
        }
    return render(request, 'app/salary_advance_list.html', context)

def salary_advance_create(request):
    """Render the salary advance creation form."""
    staff_list = CustomUser.objects.filter(
        is_active=True
    ).exclude(is_superuser=True).order_by('first_name', 'last_name')

    # Build account options (can be extended from ExpenseCategory model)
    account_options = [
        'Salary Advance',
        'Loan',
        'Emergency Advance',
        'School Fees Advance',
        'Medical Advance',
    ]

    context = {
        **get_base_context(request),
        'staff_list': staff_list,
        'account_options': account_options,
        'period_options': range(1, 25),  # 1–24 months
        'today': date.today().strftime('%m-%Y'),
    }
    return render(request, 'app/salary_advance_form.html', context)


@require_POST
def salary_advance_store(request):
    """Handle form submission to create a salary advance."""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    employee_id = request.POST.get('employee_id')
    account = request.POST.get('account', 'Salary Advance').strip()
    amount_str = request.POST.get('amount', '0').replace(',', '')
    period_str = request.POST.get('period', '1')
    starting_month_str = request.POST.get('starting_payment_month', '')  # MM-YYYY

    # Validate
    errors = {}
    if not employee_id:
        errors['employee'] = 'Please select a staff member.'
    if not amount_str or float(amount_str) <= 0:
        errors['amount'] = 'Please enter a valid amount.'
    if not period_str or int(period_str) < 1:
        errors['period'] = 'Please select a valid period.'
    if not starting_month_str:
        errors['starting_month'] = 'Please enter the starting payment month.'

    if errors:
        if is_ajax:
            return JsonResponse({'success': False, 'errors': errors}, status=400)
        messages.error(request, 'Please fix the errors below.')
        return redirect('salary_advance_create')

    try:
        employee = CustomUser.objects.get(pk=employee_id)
        amount = float(amount_str)
        period = int(period_str)

        # Parse MM-YYYY → date (1st of that month)
        month, year = starting_month_str.split('-')
        start_date = date(int(year), int(month), 1)

        advance = SalaryAdvance(
            employee=employee,
            account=account,
            amount=amount,
            payment_period=period,
            starting_payment_month=start_date,
        )
        advance.save()  # triggers auto-calculation in model.save()

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'Salary advance for {employee.get_full_name()} saved successfully.',
                'id': advance.id,
            })

        messages.success(
            request,
            f'Salary advance for {employee.get_full_name()} created successfully.'
        )
        return redirect('salary_advance_list')

    except CustomUser.DoesNotExist:
        if is_ajax:
            return JsonResponse({'success': False, 'errors': {'employee': 'Staff not found.'}}, status=404)
        messages.error(request, 'Selected staff member not found.')
        return redirect('salary_advance_create')

    except Exception as e:
        if is_ajax:
            return JsonResponse({'success': False, 'errors': {'general': str(e)}}, status=500)
        messages.error(request, f'An error occurred: {e}')
        return redirect('salary_advance_create')


def salary_advance_preview(request):
    """
    AJAX endpoint: given amount, period, starting_month
    returns schedule preview (month + installment rows).
    """
    amount_str = request.GET.get('amount', '0').replace(',', '')
    period_str = request.GET.get('period', '1')
    starting_str = request.GET.get('starting_month', '')

    try:
        amount = float(amount_str) if amount_str else 0
        period = int(period_str) if period_str else 1
        installment = round(amount / period, 2) if period > 0 else 0

        schedule = []
        if starting_str:
            month, year = starting_str.split('-')
            start_date = date(int(year), int(month), 1)
            for i in range(period):
                d = start_date + relativedelta(months=i)
                schedule.append({
                    'month': d.strftime('%m-%Y'),
                    'amount': installment,
                })
            ending = start_date + relativedelta(months=period - 1)
            ending_str = ending.strftime('%m-%Y')
        else:
            ending_str = ''

        return JsonResponse({
            'success': True,
            'installment': installment,
            'ending_month': ending_str,
            'schedule': schedule,
            'total': amount,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


def get_employee_info(request):
    """AJAX: return employee role/designation for a given employee id."""
    employee_id = request.GET.get('employee_id')
    try:
        emp = CustomUser.objects.get(pk=employee_id)
        return JsonResponse({
            'success': True,
            'name': emp.get_full_name() or emp.username,
            'role': emp.role or '',
            'salary': float(emp.salary or 0),
        })
    except CustomUser.DoesNotExist:
        return JsonResponse({'success': False}, status=404)
    
    
    
    
from django.db.models import Count, Sum, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce  
def loans_issued_report_filter(request):
    """
    Page 1: Date filter form for loans issued report.
    """
    today = date.today()
    default_from = today.replace(day=1)   # first of current month

    context = {
        'date_from': request.GET.get('date_from', default_from.strftime('%Y-%m-%d')),
        'date_to':   request.GET.get('date_to',   today.strftime('%Y-%m-%d')),
        **get_base_context(request),
    }
    return render(request, 'app/loans_issued_report_filter.html', context)



def loans_issued_report_result(request):
    """
    Page 2: Branch-by-branch summary of loans issued in the selected date range.

    LoanApplication fields used:
        application_date          – DateField
        office                    – CharField  (branch name stored as text)
        loan_amount               – principal issued
        total_interest_amount     – total interest  (auto-calculated on save)
        total_repayment_amount    – loan + interest (auto-calculated on save)
        status                    – 'Approved' by default
    """
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to',   '')

    # ── Validate dates ────────────────────────────────────────────────────────
    if not date_from_str or not date_to_str:
        messages.error(request, 'Please select both start and end dates.')
        return redirect('loans_issued_report_filter')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('loans_issued_report_filter')

    if date_from > date_to:
        messages.error(request, '"Date from" cannot be after "Date to".')
        return redirect('loans_issued_report_filter')

    # ── Base queryset ─────────────────────────────────────────────────────────
    # LoanApplication.application_date is set via auto_now_add in the model save()
    # office is a CharField containing the branch name
    qs = LoanApplication.objects.filter(
        application_date__gte=date_from,
        application_date__lte=date_to,
    )
    # Uncomment to restrict to approved loans only:
    # qs = qs.filter(status='Approved')

    # ── Group by office (CharField) and aggregate ─────────────────────────────
    branch_data = (
        qs
        .values('office')
        .annotate(
            no_of_loans    = Count('id'),
            loaned_amount  = Coalesce(Sum('loan_amount'),            Decimal('0')),
            interest_amount= Coalesce(Sum('total_interest_amount'), Decimal('0')),
            total_return   = Coalesce(Sum('total_repayment_amount'), Decimal('0')),
        )
        .order_by('office')
    )

    # Build a clean list with friendly keys for the template
    branch_summary = [
        {
            'branch_name':     row['office'] or 'N/A',
            'no_of_loans':     row['no_of_loans'],
            'loaned_amount':   row['loaned_amount'],
            'interest_amount': row['interest_amount'],
            'total_return':    row['total_return'],
        }
        for row in branch_data
    ]

    # ── Grand totals ──────────────────────────────────────────────────────────
    totals = {
        'no_of_loans':     sum(r['no_of_loans']     for r in branch_summary),
        'loaned_amount':   sum(r['loaned_amount']   for r in branch_summary),
        'interest_amount': sum(r['interest_amount'] for r in branch_summary),
        'total_return':    sum(r['total_return']    for r in branch_summary),
    }

    context = {
        'date_from':      date_from,
        'date_to':        date_to,
        'branch_summary': branch_summary,
        'totals':         totals,
        **get_base_context(request),
    }
    return render(request, 'app/loans_issued_report_result.html', context)








def monthly_outstanding_filter_summary(request):
    """
    Page 1: Pick a month/date — show all loans still outstanding up to
    the end of that month.
    """
    today = date.today()
    context = {
        'selected_month': today.strftime('%Y-%m-%d'),
        **get_base_context(request),
    }
    return render(request, 'app/monthly_outstanding_filter_summary.html', context)


def monthly_outstanding_result_summary(request):
    """
    Page 2: For the selected month/date, show loans that:
      - Were issued ON OR BEFORE the last day of that month
        (application_date <= end_of_month)
      - Still have a remaining balance > 0
        (repayment_amount_remaining > 0)

    Groups by office (CharField) and sums repayment_amount_remaining.

    LoanApplication fields used:
        application_date            – DateField
        office                      – CharField  (branch name as text)
        repayment_amount_remaining  – DecimalField  (outstanding balance)
    """
    month_str = request.GET.get('month', '')

    # ── Validate ──────────────────────────────────────────────────────────────
    if not month_str:
        messages.error(request, 'Please select a month.')
        return redirect('monthly_outstanding_filter')

    try:
        selected_date = datetime.datetime.strptime(month_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('monthly_outstanding_filter')

    # Get the last day of the selected month
    import calendar
    last_day = calendar.monthrange(selected_date.year, selected_date.month)[1]
    end_of_month = selected_date.replace(day=last_day)

    # ── Query ─────────────────────────────────────────────────────────────────
    # Outstanding = loans issued up to end of that month that still have balance
    qs = LoanApplication.objects.filter(
        application_date__lte=end_of_month,          # issued on or before month end
        repayment_amount_remaining__gt=Decimal('0'), # still has outstanding balance
    )

    branch_data = (
        qs
        .values('office')
        .annotate(
            no_of_loans       = Count('id'),
            outstanding_amount= Coalesce(
                Sum('repayment_amount_remaining'), Decimal('0')
            ),
        )
        .order_by('office')
    )

    branch_summary = [
        {
            'branch_name':       row['office'] or 'N/A',
            'no_of_loans':       row['no_of_loans'],
            'outstanding_amount': row['outstanding_amount'],
        }
        for row in branch_data
    ]

    # ── Grand totals ──────────────────────────────────────────────────────────
    totals = {
        'no_of_loans':       sum(r['no_of_loans']        for r in branch_summary),
        'outstanding_amount': sum(r['outstanding_amount'] for r in branch_summary),
    }

    context = {
        'selected_month': selected_date,
        'end_of_month':   end_of_month,
        'branch_summary': branch_summary,
        'totals':         totals,
        **get_base_context(request),
    }
    return render(request, 'app/monthly_outstanding_result_summary.html', context)







def classify_loan(loan, today):
    """
    Classify a loan by how many months overdue it is.

    Standard microfinance classification (days past due):
        Current      : 0   – 30  days overdue  (on track or just missed)
        ESM          : 31  – 60  days overdue
        Substandard  : 61  – 90  days overdue
        Doubtful     : 91  – 180 days overdue
        Loss         : >180 days overdue

    We use first_repayment_date as the reference start date.
    If repayment_amount_remaining == 0 the loan is fully paid — excluded.
    """
    if not loan.first_repayment_date:
        return 'loss'   # no repayment date set → treat as loss

    days_overdue = (today - loan.first_repayment_date).days

    if days_overdue <= 30:
        return 'current'
    elif days_overdue <= 60:
        return 'esm'
    elif days_overdue <= 90:
        return 'substandard'
    elif days_overdue <= 180:
        return 'doubtful'
    else:
        return 'loss'



def expired_loans_report_summary(request):
    """
    One-page report: expired / overdue loans up to today, summarised by branch.

    "Expired" means the loan's first_repayment_date has passed and there is
    still a remaining balance (repayment_amount_remaining > 0).

    Columns per branch:
        Loan Issued   – original loan_amount
        Outstanding   – repayment_amount_remaining
        Current       – count of loans 0-30 days overdue
        ESM           – count 31-60 days overdue
        Substandard   – count 61-90 days overdue
        Doubtful      – count 91-180 days overdue
        Loss          – count >180 days overdue
        Total         – sum of all classified counts
    """
    today = date.today()

    # All loans that are past their first repayment date and still outstanding
    expired_qs = LoanApplication.objects.filter(
        first_repayment_date__lt=today,
        repayment_amount_remaining__gt=Decimal('0'),
    ).values(
        'office',
        'loan_amount',
        'repayment_amount_remaining',
        'first_repayment_date',
    ).order_by('office')

    # ── Build per-branch summary in Python ───────────────────────────────────
    branch_map = {}   # key = office name

    for loan in expired_qs:
        branch = loan['office'] or 'N/A'
        cls    = classify_loan_dict(loan, today)

        if branch not in branch_map:
            branch_map[branch] = {
                'branch_name':   branch,
                'loan_issued':   Decimal('0'),
                'outstanding':   Decimal('0'),
                'current':       0,
                'esm':           0,
                'substandard':   0,
                'doubtful':      0,
                'loss':          0,
            }

        branch_map[branch]['loan_issued']  += Decimal(str(loan['loan_amount']))
        branch_map[branch]['outstanding']  += Decimal(str(loan['repayment_amount_remaining']))
        branch_map[branch][cls]            += 1

    # Sort branches alphabetically
    branch_summary = sorted(branch_map.values(), key=lambda x: x['branch_name'])

    # Add total column per branch
    for row in branch_summary:
        row['total'] = row['current'] + row['esm'] + row['substandard'] + row['doubtful'] + row['loss']

    # ── Grand totals ──────────────────────────────────────────────────────────
    totals = {
        'loan_issued':  sum(r['loan_issued']  for r in branch_summary),
        'outstanding':  sum(r['outstanding']  for r in branch_summary),
        'current':      sum(r['current']      for r in branch_summary),
        'esm':          sum(r['esm']          for r in branch_summary),
        'substandard':  sum(r['substandard']  for r in branch_summary),
        'doubtful':     sum(r['doubtful']     for r in branch_summary),
        'loss':         sum(r['loss']         for r in branch_summary),
        'total':        sum(r['total']        for r in branch_summary),
    }

    context = {
        'today':          today,
        'branch_summary': branch_summary,
        'totals':         totals,
        **get_base_context(request),
    }
    return render(request, 'app/expired_loans_report_summary.html', context)


# ── Helper: classify from dict (used in loop above) ──────────────────────────
def classify_loan_dict(loan_dict, today):
    first_date = loan_dict.get('first_repayment_date')
    if not first_date:
        return 'loss'
    days_overdue = (today - first_date).days
    if days_overdue <= 30:
        return 'current'
    elif days_overdue <= 60:
        return 'esm'
    elif days_overdue <= 90:
        return 'substandard'
    elif days_overdue <= 180:
        return 'doubtful'
    else:
        return 'loss'
    
    
    
def bank_transfer_expenses_filter(request):
    """
    Page 1: Date range filter form for bank transfer expenses report.
    """
    today = date.today()
    default_from = today.replace(day=1)  # first of current month
    context = {
        'date_from': request.GET.get('date_from', default_from.strftime('%Y-%m-%d')),
        'date_to':   request.GET.get('date_to',   today.strftime('%Y-%m-%d')),
        **get_base_context(request),
    }
    return render(request, 'app/bank_transfer_expenses_filter.html', context)



def bank_transfer_expenses_result(request):
    """
    Page 2: Summary of HQ transfers RECEIVED by each branch within date range.

    Uses HQTransaction model:
        transaction_date  – DateField
        to_branch         – FK to Office  (the receiving branch)
        amount            – DecimalField
    """
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to',   '')

    # ── Validate ──────────────────────────────────────────────────────────────
    if not date_from_str or not date_to_str:
        messages.error(request, 'Please select both start and end dates.')
        return redirect('bank_transfer_expenses_filter')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('bank_transfer_expenses_filter')

    if date_from > date_to:
        messages.error(request, '"Date from" cannot be after "Date to".')
        return redirect('bank_transfer_expenses_filter')

    # ── All offices sorted alphabetically ─────────────────────────────────────
    all_offices = Office.objects.exclude(name__iexact='HQ').order_by('name')

    # ── HQ transfers received by each branch in date range ────────────────────
    hq_received = (
        HQTransaction.objects.filter(
            transaction_date__gte=date_from,
            transaction_date__lte=date_to,
            to_branch__isnull=False,
        )
        .values('to_branch_id')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0')))
    )

    # Build lookup: office_id → total received
    received_map = {row['to_branch_id']: row['total'] for row in hq_received}

    # ── Every office gets a row, None if no transfers received ────────────────
    branch_summary = []
    for office in all_offices:
        branch_summary.append({
            'branch_name':  office.name,
            'total_amount': received_map.get(office.id, None),  # None → blank cell
        })

    # ── Grand total ───────────────────────────────────────────────────────────
    grand_total = sum(
        r['total_amount'] for r in branch_summary if r['total_amount']
    )

    context = {
        'date_from':      date_from,
        'date_to':        date_to,
        'branch_summary': branch_summary,
        'grand_total':    grand_total,
        **get_base_context(request),
    }
    return render(request, 'app/bank_transfer_expenses_result.html', context)



def branch_to_hq_expenses_filter(request):
    """
    Page 1: Date range filter form for branch-to-HQ transfer report.
    """
    today = date.today()
    default_from = today.replace(day=1)
    context = {
        'date_from': request.GET.get('date_from', default_from.strftime('%Y-%m-%d')),
        'date_to':   request.GET.get('date_to',   today.strftime('%Y-%m-%d')),
        **get_base_context(request),
    }
    return render(request, 'app/branch_to_hq_expenses_filter.html', context)


def branch_to_hq_expenses_result(request):
    """
    Page 2: Summary of transfers SENT by each branch to HQ within date range.

    Uses HQTransaction model:
        transaction_date  – DateField
        from_branch       – FK to Office  (the sending branch)
        amount            – DecimalField
    """
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to',   '')

    # ── Validate ──────────────────────────────────────────────────────────────
    if not date_from_str or not date_to_str:
        messages.error(request, 'Please select both start and end dates.')
        return redirect('branch_to_hq_expenses_filter')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('branch_to_hq_expenses_filter')

    if date_from > date_to:
        messages.error(request, '"Date from" cannot be after "Date to".')
        return redirect('branch_to_hq_expenses_filter')

    # ── All offices (branches) sorted alphabetically ──────────────────────────
    all_offices = Office.objects.exclude(name__iexact='HQ').order_by('name')

    # ── Transfers SENT by each branch to HQ in date range ─────────────────────
    branch_sent = (
        HQTransaction.objects.filter(
            transaction_date__gte=date_from,
            transaction_date__lte=date_to,
            from_branch__isnull=False,        # <── flipped: sender is the branch
        )
        .values('from_branch_id')             # <── group by sender
        .annotate(total=Coalesce(Sum('amount'), Decimal('0')))
    )

    # Build lookup: office_id → total sent
    sent_map = {row['from_branch_id']: row['total'] for row in branch_sent}

    # ── Every office gets a row, None if no transfers sent ────────────────────
    branch_summary = []
    for office in all_offices:
        branch_summary.append({
            'branch_name':  office.name,
            'total_amount': sent_map.get(office.id, None),
        })

    # ── Grand total ───────────────────────────────────────────────────────────
    grand_total = sum(
        r['total_amount'] for r in branch_summary if r['total_amount']
    )

    context = {
        'date_from':      date_from,
        'date_to':        date_to,
        'branch_summary': branch_summary,
        'grand_total':    grand_total,
        **get_base_context(request),
    }
    return render(request, 'app/branch_to_hq_expenses_result.html', context)





def expenses_statement_filter(request):
    """
    Page 1: Date range filter form for expenses statement report.
    """
    today = date.today()
    default_from = today.replace(day=1)  # first of current month
    context = {
        'date_from': request.GET.get('date_from', default_from.strftime('%Y-%m-%d')),
        'date_to':   request.GET.get('date_to',   today.strftime('%Y-%m-%d')),
        **get_base_context(request),
    }
    return render(request, 'app/expenses_statement_filter.html', context)


def expenses_statement_result(request):
    """
    Page 2: Summary of expenses grouped by ExpenseCategory (Account Name).

    Lists EVERY ExpenseCategory — even those with zero expenses in the period
    (shows blank amount cell, matching the screenshot style).

    Sums all Expense.amount values linked to each category via the
    transaction_type ForeignKey, filtered by transaction_date in the range.

    Expense model fields used:
        transaction_type  – ForeignKey to ExpenseCategory
        transaction_date  – DateField (actual transaction date)
        amount            – DecimalField
        (no office filter — all offices combined as per requirement)
    """
    date_from_str = request.GET.get('date_from', '')
    date_to_str   = request.GET.get('date_to',   '')

    # ── Validate ──────────────────────────────────────────────────────────────
    if not date_from_str or not date_to_str:
        messages.error(request, 'Please select both start and end dates.')
        return redirect('expenses_statement_filter')

    try:
        date_from = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('expenses_statement_filter')

    if date_from > date_to:
        messages.error(request, '"Date from" cannot be after "Date to".')
        return redirect('expenses_statement_filter')

    # ── All expense categories ordered alphabetically ─────────────────────────
    all_categories = ExpenseCategory.objects.all().order_by('name')

    # ── Aggregate expenses per category within date range ─────────────────────
    # Uses transaction_date (actual date); falls back to expense_date if null
    expense_totals = (
        Expense.objects.filter(
            transaction_date__gte=date_from,
            transaction_date__lte=date_to,
        )
        .values('transaction_type_id')
        .annotate(total=Coalesce(Sum('amount'), Decimal('0')))
    )

    # Build lookup: category_id → total amount
    totals_map = {row['transaction_type_id']: row['total'] for row in expense_totals}

    # ── Merge: every category gets a row ─────────────────────────────────────
    category_summary = []
    for cat in all_categories:
        amount = totals_map.get(cat.id, None)   # None → blank cell like screenshot
        category_summary.append({
            'category_name': cat.name,
            'total_amount':  amount,
        })

    # ── Grand total (only categories that have expenses) ──────────────────────
    grand_total = sum(
        r['total_amount'] for r in category_summary if r['total_amount']
    )

    context = {
        'date_from':        date_from,
        'date_to':          date_to,
        'category_summary': category_summary,
        'grand_total':      grand_total,
        **get_base_context(request),
    }
    return render(request, 'app/expenses_statement_result.html', context)







# ══════════════════════════════════════════════════════════════════════
#  PAGE 1 — Filter
# ══════════════════════════════════════════════════════════════════════


def general_financial_statement_filter(request):
    """
    Simple date-range filter form for the HQ financial statement.
    """
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/general_financial_statement_filter.html', context)


def general_financial_statement_report(request):

    if request.method != 'POST':
        return redirect('financial_statement_filter')

    date_from_str = request.POST.get('date_from', '')
    date_to_str   = request.POST.get('date_to',   '')

    # ── Validate dates ────────────────────────────────────────────────
    if not date_from_str or not date_to_str:
        messages.error(request, 'Please select both dates.')
        return redirect('financial_statement_filter')

    try:
        start_date = datetime.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('financial_statement_filter')

    if start_date > end_date:
        messages.error(request, '"Date from" cannot be after "Date to".')
        return redirect('financial_statement_filter')

    # ── Resolve office from session ───────────────────────────────────
    base_ctx        = get_base_context(request)
    selected_office = base_ctx['selected_office']

    if selected_office is None:
        messages.error(request, 'No office selected. Please select an office first.')
        return redirect('financial_statement_filter')

    date_from_display = start_date.strftime('%d/%m/%Y')
    date_to_display   = end_date.strftime('%d/%m/%Y')

    def _sum(qs, field):
        return qs.aggregate(t=Sum(field))['t'] or Decimal('0')

    # ══ OPENING BALANCE (before start_date, selected office only) ═════

    open_repayments = _sum(
        LoanRepayment.objects.filter(
            repayment_date__lt=start_date,
            loan_application__office=selected_office.name,
        ),
        'repayment_amount'
    )
    open_nyongeza = _sum(
        Nyongeza.objects.filter(
            date__lt=start_date,
            Office=selected_office,
        ),
        'amount'
    )
    open_external = _sum(
        OfficeTransaction.objects.filter(
            transaction_date__lt=start_date,
            office_to=selected_office,
        ),
        'amount'
    )
    open_loans = _sum(
        LoanApplication.objects.filter(
            application_date__lt=start_date,
            office=selected_office.name,
        ),
        'loan_amount'
    )
    open_expenses = _sum(
        Expense.objects.filter(
            expense_date__lt=start_date,
            office=selected_office.name,
        ),
        'amount'
    )
    open_salaries = _sum(
        Salary.objects.filter(
            salary_for_month__lt=start_date,
            fund_source=selected_office,
        ),
        'amount'
    )

    opening_balance = (
        (open_repayments + open_nyongeza + open_external)
        - (open_loans + open_expenses + open_salaries)
    )

    # ══ CURRENT PERIOD (selected office only) ════════════════════════

    pq_rep  = Q(repayment_date__gte=start_date,   repayment_date__lte=end_date,
                loan_application__office=selected_office.name)
    pq_nyo  = Q(date__gte=start_date,             date__lte=end_date,
                Office=selected_office)
    pq_exp  = Q(expense_date__gte=start_date,     expense_date__lte=end_date,
                office=selected_office.name)
    pq_sal  = Q(salary_for_month__gte=start_date, salary_for_month__lte=end_date,
                fund_source=selected_office)
    pq_trx_in  = Q(transaction_date__gte=start_date, transaction_date__lte=end_date,
                   office_to=selected_office)
    pq_trx_out = Q(transaction_date__gte=start_date, transaction_date__lte=end_date,
                   office_from=selected_office)

    # ══ INCOME ═══════════════════════════════════════════════════════

    total_repayments = _sum(LoanRepayment.objects.filter(pq_rep),        'repayment_amount')
    total_nyongeza   = _sum(Nyongeza.objects.filter(pq_nyo),             'amount')
    total_external   = _sum(OfficeTransaction.objects.filter(pq_trx_in), 'amount')

    income_subtotal = opening_balance + total_repayments + total_nyongeza + total_external

    # ══ OUTFLOW ══════════════════════════════════════════════════════

    total_money_to_branch = _sum(OfficeTransaction.objects.filter(pq_trx_out), 'amount')
    total_expenses        = _sum(Expense.objects.filter(pq_exp),               'amount')
    total_salaries        = _sum(Salary.objects.filter(pq_sal),                'amount')
    total_expenses        = total_expenses + total_salaries

    outflow_subtotal = total_money_to_branch + total_expenses

    # ══ CLOSING BALANCE ══════════════════════════════════════════════
    closing_balance = income_subtotal - outflow_subtotal

    context = {
        **base_ctx,

        'date_from_display':     date_from_display,
        'date_to_display':       date_to_display,

        # Opening
        'opening_balance':       opening_balance,

        # Income
        'total_repayments':      total_repayments,
        'total_nyongeza':        total_nyongeza,
        'total_external':        total_external,
        'income_subtotal':       income_subtotal,

        # Outflow
        'total_money_to_branch': total_money_to_branch,
        'total_expenses':        total_expenses,
        'total_salaries':        total_salaries,
        'outflow_subtotal':      outflow_subtotal,

        # Closing
        'closing_balance':       closing_balance,
    }

    return render(request, 'app/general_financial_statement_report.html', context)




def branch_financial_summary_filter(request):
    """Simple date-range form that submits to branch_financial_summary."""
    today = timezone.now().date()
    context = {
        'start_date': today.replace(day=1).strftime('%Y-%m-%d'),
        'end_date':   today.strftime('%Y-%m-%d'),
        **get_base_context(request),
    }
    return render(request, 'app/branch_financial_summary_filter.html', context)


# ══════════════════════════════════════════════════════════════
#  VIEW 2 — Report  (GET, receives start_date / end_date)
# ══════════════════════════════════════════════════════════════


def branch_financial_summary(request):
    
    start_date_str = request.GET.get('start_date', '')
    end_date_str   = request.GET.get('end_date',   '')

    if not start_date_str or not end_date_str:
        messages.error(request, 'Please select both dates.')
        return redirect('branch_financial_summary_filter')

    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date   = datetime.datetime.strptime(end_date_str,   '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'Invalid date format.')
        return redirect('branch_financial_summary_filter')

    if start_date > end_date:
        messages.error(request, '"Date from" cannot be after "Date to".')
        return redirect('branch_financial_summary_filter')

    offices = Office.objects.exclude(name__iexact='HQ').order_by('name')

    # ── Data buckets ──────────────────────────────────────────────────
    summary_data = {
        'opening_balance':     {},
        'mapato':              {},
        'hazina':              {},
        'nyongeza':            {},
        'transfers_in':        {},
        'income_subtotal':     {},
        'fomu':                {},
        'matumizi_ofisini':    {},
        'matumizi_kituo':      {},
        'matumizi_mkurugenzi': {},
        'makato_benki':        {},
        'outflow_subtotal':    {},
        'balance_cash':        {},
        'balance_benki':       {},
        'balance_total':       {},
    }

    totals = {k: Decimal('0.00') for k in summary_data}

    for office in offices:

        # ── OPENING BALANCE ──────────────────────────────────────────
        rep_b = LoanRepayment.objects.filter(
            loan_application__office=office.name,
            repayment_date__lt=start_date,
        ).exclude(
            loan_application__loan_type__iexact='Hazina'
        ).aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')

        hazina_rep_b = LoanRepayment.objects.filter(
            loan_application__office=office.name,
            repayment_date__lt=start_date,
            loan_application__loan_type__iexact='Hazina',
        ).aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')

        nyo_b = Nyongeza.objects.filter(
            Office=office,
            date__lt=start_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        trx_in_b = OfficeTransaction.objects.filter(
            office_to=office,
            transaction_date__lt=start_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        exp_b = Expense.objects.filter(
            office=office.name,
            expense_date__lt=start_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        loan_b = LoanApplication.objects.filter(
            office=office.name,
            application_date__lt=start_date,
        ).aggregate(t=Sum('loan_amount'))['t'] or Decimal('0')

        trx_out_b = OfficeTransaction.objects.filter(
            office_from=office,
            transaction_date__lt=start_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        bank_charge_b = BankCharge.objects.filter(
            office=office.name,
            expense_date__lt=start_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        opening_balance = (
            rep_b + hazina_rep_b + nyo_b + trx_in_b
        ) - (
            exp_b + loan_b + trx_out_b + bank_charge_b
        )

        summary_data['opening_balance'][office.id] = opening_balance
        totals['opening_balance'] += opening_balance

        # ── MAPATO (non-Hazina repayments only) ──────────────────────
        mapato = LoanRepayment.objects.filter(
            loan_application__office=office.name,
            repayment_date__gte=start_date,
            repayment_date__lte=end_date,
        ).exclude(
            loan_application__loan_type__iexact='Hazina'
        ).aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')

        summary_data['mapato'][office.id] = mapato
        totals['mapato'] += mapato

        # ── HAZINA (Hazina repayments only) ──────────────────────────
        hazina = LoanRepayment.objects.filter(
            loan_application__office=office.name,
            repayment_date__gte=start_date,
            repayment_date__lte=end_date,
            loan_application__loan_type__iexact='Hazina',
        ).aggregate(t=Sum('repayment_amount'))['t'] or Decimal('0')

        summary_data['hazina'][office.id] = hazina
        totals['hazina'] += hazina

        # ── NYONGEZA ─────────────────────────────────────────────────
        nyongeza = Nyongeza.objects.filter(
            Office=office,
            date__gte=start_date,
            date__lte=end_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        summary_data['nyongeza'][office.id] = nyongeza
        totals['nyongeza'] += nyongeza

        # ── FEDHA ZILIZOPOKELEWA (transfers received into office) ────
        transfers_in = OfficeTransaction.objects.filter(
            office_to=office,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        summary_data['transfers_in'][office.id] = transfers_in
        totals['transfers_in'] += transfers_in

        # ── INCOME SUBTOTAL ──────────────────────────────────────────
        income_subtotal = opening_balance + mapato + hazina + nyongeza + transfers_in
        summary_data['income_subtotal'][office.id] = income_subtotal
        totals['income_subtotal'] += income_subtotal

        # ── FOMU (Loan Disbursements) ─────────────────────────────────
        fomu = LoanApplication.objects.filter(
            office=office.name,
            application_date__gte=start_date,
            application_date__lte=end_date,
        ).aggregate(t=Sum('loan_amount'))['t'] or Decimal('0')

        summary_data['fomu'][office.id] = fomu
        totals['fomu'] += fomu

        # ── MATUMIZI OFISINI (office expenses only) ───────────────────
        matumizi_ofisini = Expense.objects.filter(
            office=office.name,
            expense_date__gte=start_date,
            expense_date__lte=end_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        summary_data['matumizi_ofisini'][office.id] = matumizi_ofisini
        totals['matumizi_ofisini'] += matumizi_ofisini

        # ── MATUMIZI BENKI-[KITUO] (branch-to-branch, excludes HQ) ───
        matumizi_kituo = OfficeTransaction.objects.filter(
            office_from=office,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
        ).exclude(
            office_to__name__iexact='HQ'
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        summary_data['matumizi_kituo'][office.id] = matumizi_kituo
        totals['matumizi_kituo'] += matumizi_kituo

        # ── MATUMIZI BENKI-[MKURUGENZI] (transfers to HQ only) ───────
        matumizi_mkurugenzi = OfficeTransaction.objects.filter(
            office_from=office,
            transaction_date__gte=start_date,
            transaction_date__lte=end_date,
            office_to__name__iexact='HQ',
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        summary_data['matumizi_mkurugenzi'][office.id] = matumizi_mkurugenzi
        totals['matumizi_mkurugenzi'] += matumizi_mkurugenzi

        # ── MAKATO BANK ───────────────────────────────────────────────
        makato_benki = BankCharge.objects.filter(
            office=office.name,
            expense_date__gte=start_date,
            expense_date__lte=end_date,
        ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

        summary_data['makato_benki'][office.id] = makato_benki
        totals['makato_benki'] += makato_benki

        # ── OUTFLOW SUBTOTAL ──────────────────────────────────────────
        outflow_subtotal = (
            fomu
            + matumizi_ofisini
            + matumizi_kituo
            + matumizi_mkurugenzi
            + makato_benki
        )
        summary_data['outflow_subtotal'][office.id] = outflow_subtotal
        totals['outflow_subtotal'] += outflow_subtotal

        # ── LIVE BALANCES (from BranchBalance) ───────────────────────
        latest = BranchBalance.objects.filter(
            branch=office
        ).order_by('-last_updated').first()

        bal_cash  = latest.office_balance if latest else Decimal('0')
        bal_bank  = latest.bank_balance   if latest else Decimal('0')
        bal_total = bal_cash + bal_bank

        summary_data['balance_cash'][office.id]  = bal_cash
        summary_data['balance_benki'][office.id] = bal_bank
        summary_data['balance_total'][office.id] = bal_total
        totals['balance_cash']  += bal_cash
        totals['balance_benki'] += bal_bank
        totals['balance_total'] += bal_total

    context = {
        **get_base_context(request),
        'offices':    offices,
        'summary':    summary_data,
        'totals':     totals,
        'start_date': start_date,
        'end_date':   end_date,
    }
    return render(request, 'app/branch_financial_summary_result.html', context)







def salary_slip_filter(request):
    """Month picker form for the salary slip / payroll list."""
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/salary_slip_filter.html', context)


# ══════════════════════════════════════════════════════════════════
#  VIEW 2 — Payroll List  (GET ?month=YYYY-MM)
# ══════════════════════════════════════════════════════════════════

def salary_slip_list(request):
    """
    Payroll list for a selected month — ALL offices combined.

    URL param: ?month=YYYY-MM

    Salary model fields used:
        employee          – ForeignKey to CustomUser
        fund_source       – ForeignKey to Office  (the employee's branch)
        salary_for_month  – DateField  (first day of the month)
        amount            – DecimalField  (basic salary)
        deduction         – DecimalField  (total deductions)
        net_salary        – DecimalField  (amount - deduction, or computed)
        payment_date      – DateField (nullable)
        receipt_number    – CharField (nullable)
        status            – CharField  ('paid' | 'pending' | 'unpaid')
    """
    month_param = request.GET.get('month', '')

    if not month_param:
        messages.error(request, 'Please select a month.')
        return redirect('salary_slip_filter')

    try:
        year, month = map(int, month_param.split('-'))
        # First and last day of the selected month
        start_date = datetime.date(year, month, 1)
        import calendar
        last_day   = calendar.monthrange(year, month)[1]
        end_date   = datetime.date(year, month, last_day)
    except (ValueError, AttributeError):
        messages.error(request, 'Invalid month format.')
        return redirect('salary_slip_filter')

    MONTH_NAMES = {
        1:'Jan', 2:'Feb', 3:'Mar',  4:'Apr',  5:'May',  6:'Jun',
        7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
    }
    month_label = f"{MONTH_NAMES[month]}-{year}"   # e.g. "Mar-2026"

    # ── Fetch all salaries for the month across ALL offices ────────────────
    salaries = (
        Salary.objects
        .filter(salary_for_month__gte=start_date, salary_for_month__lte=end_date)
        .select_related('employee', 'fund_source')
        .order_by('fund_source__name', 'employee__first_name')
    )

    # ── Group by branch/office & compute per-row values ───────────────────
    # Deduction is NOT on Salary model — it comes from CustomUser.deduction_amount
    grouped_salaries = defaultdict(list)
    total_basic     = Decimal('0')
    total_deduction = Decimal('0')
    total_net       = Decimal('0')

    for sal in salaries:
        branch = sal.fund_source.name if sal.fund_source else 'No Branch'

        # Basic salary: Salary.amount, fall back to employee.salary
        basic = sal.amount or Decimal('0')
        if basic == 0 and sal.employee:
            basic = sal.employee.salary or Decimal('0')

        # Deduction: CustomUser.deduction_amount
        deduction = Decimal('0')
        if sal.employee and sal.employee.deduction_amount:
            deduction = sal.employee.deduction_amount

        net = basic - deduction

        # Attach computed values as template-accessible attributes
        sal.basic_salary = basic
        sal.deduction    = deduction
        sal.net_salary   = net

        total_basic     += basic
        total_deduction += deduction
        total_net       += net

        grouped_salaries[branch].append(sal)

    # Sort branches alphabetically
    grouped_salaries = dict(sorted(grouped_salaries.items()))

    totals = {
        'basic_salary': total_basic,
        'deduction':    total_deduction,
        'net_salary':   total_net,
    }

    context = {
        'month_param':      month_param,
        'month_label':      month_label,
        'salaries':         salaries,
        'grouped_salaries': grouped_salaries,
        'totals':           totals,
        'start_date':       start_date,
        'end_date':         end_date,
        **get_base_context(request),
    }
    return render(request, 'app/salary_slip_list.html', context)


# ══════════════════════════════════════════════════════════════════
#  VIEW 3 — Mark All As Paid  (GET)
# ══════════════════════════════════════════════════════════════════

def salary_slip_pay_all(request):
    """
    Mark all unpaid/pending salaries for the month as 'paid'.
    Redirects back to the list.
    """
    month_param = request.GET.get('month', '')

    if not month_param:
        return redirect('salary_slip_filter')

    try:
        year, month = map(int, month_param.split('-'))
        import calendar
        start_date = datetime.date(year, month, 1)
        last_day   = calendar.monthrange(year, month)[1]
        end_date   = datetime.date(year, month, last_day)
    except ValueError:
        return redirect('salary_slip_filter')

    updated = Salary.objects.filter(
        salary_for_month__gte=start_date,
        salary_for_month__lte=end_date,
        payment_date__isnull=True,
    ).update(
        payment_date=datetime.date.today(),
    )

    messages.success(request, f'{updated} salary record(s) marked as paid.')
    return redirect(f"{__import__('django.urls', fromlist=['reverse']).reverse('salary_slip_list')}?month={month_param}")




from django.urls import reverse
def _branch_url(user_id):
    """Return the manage_admin_branches URL with user_id param."""
    return reverse('manage_admin_branches') + f'?user_id={user_id}'


# ══════════════════════════════════════════════════════════════════════
#  MAIN PAGE — select staff + manage their branches
# ══════════════════════════════════════════════════════════════════════

def manage_admin_branches(request):
    """
    Page 1 (and only page):
    - Dropdown of all staff (searchable)
    - When a staff is selected via ?user_id=X in the URL:
        * Shows their current designation (role) and current office_allocation
        * Checkbox grid of offices NOT yet in their UserOfficeAssignment
        * Table of already-assigned offices with Current? badge and Remove/Set Current actions
    """
    all_staff = CustomUser.objects.filter(
        is_active=True
    ).exclude(
        is_superuser=True
    ).order_by('last_name', 'first_name')

    selected_user     = None
    assigned_offices  = []
    unassigned_offices = []

    user_id = request.GET.get('user_id')
    if user_id:
        selected_user = get_object_or_404(CustomUser, pk=user_id)

        # All offices this user is assigned to (from UserOfficeAssignment)
        assigned_offices = UserOfficeAssignment.objects.filter(
            user=selected_user
        ).select_related('office').order_by('id')

        assigned_office_ids = assigned_offices.values_list('office_id', flat=True)

        # Offices NOT yet assigned to this user
        unassigned_offices = Office.objects.exclude(
            id__in=assigned_office_ids
        ).order_by('name')

        # Ensure the current office_allocation is also in UserOfficeAssignment
        # (auto-sync: if user has an office_allocation not yet in assignments, add it)
        if selected_user.office_allocation:
            UserOfficeAssignment.objects.get_or_create(
                user=selected_user,
                office=selected_user.office_allocation,
            )
            # Refresh
            assigned_offices = UserOfficeAssignment.objects.filter(
                user=selected_user
            ).select_related('office').order_by('id')
            assigned_office_ids = assigned_offices.values_list('office_id', flat=True)
            unassigned_offices  = Office.objects.exclude(
                id__in=assigned_office_ids
            ).order_by('name')

    return render(request, 'app/manage_admin_branches.html', {
        'all_staff':          all_staff,
        'selected_user':      selected_user,
        'assigned_offices':   assigned_offices,
        'unassigned_offices': unassigned_offices,
        **get_base_context(request),
    })


# ══════════════════════════════════════════════════════════════════════
#  ADD BRANCH(ES)
# ══════════════════════════════════════════════════════════════════════

def manage_admin_branches_add(request):
    """
    POST: Add one or more offices to a user's UserOfficeAssignment.
    Does NOT change office_allocation (the current branch).
    """
    if request.method != 'POST':
        return redirect('manage_admin_branches')

    user_id    = request.POST.get('user_id')
    office_ids = request.POST.getlist('offices')

    if not user_id or not office_ids:
        messages.warning(request, 'Please select at least one branch.')
        return redirect(_branch_url(user_id))

    user = get_object_or_404(CustomUser, pk=user_id)

    added = 0
    for oid in office_ids:
        office = get_object_or_404(Office, pk=oid)
        _, created = UserOfficeAssignment.objects.get_or_create(user=user, office=office)
        if created:
            added += 1

    if added:
        messages.success(request, f'{added} branch(es) added successfully.')
    else:
        messages.info(request, 'All selected branches were already assigned.')

    return redirect(_branch_url(user_id))


# ══════════════════════════════════════════════════════════════════════
#  SET CURRENT (change office_allocation)
# ══════════════════════════════════════════════════════════════════════

def manage_admin_branches_set_current(request):
    """
    POST: Change CustomUser.office_allocation to a different assigned branch.
    The old current branch remains in UserOfficeAssignment (still assigned).
    """
    if request.method != 'POST':
        return redirect('manage_admin_branches')

    user_id   = request.POST.get('user_id')
    office_id = request.POST.get('office_id')

    user   = get_object_or_404(CustomUser, pk=user_id)
    office = get_object_or_404(Office,      pk=office_id)

    # Only allow setting current if the office is already assigned
    if not UserOfficeAssignment.objects.filter(user=user, office=office).exists():
        messages.error(request, 'This branch is not assigned to this staff member.')
        return redirect(_branch_url(user_id))

    user.office_allocation = office
    user.save(update_fields=['office_allocation'])

    messages.success(request, f'Current branch changed to {office.name}.')
    return redirect(_branch_url(user_id))


# ══════════════════════════════════════════════════════════════════════
#  REMOVE BRANCH
# ══════════════════════════════════════════════════════════════════════

def manage_admin_branches_remove(request):
    """
    POST: Remove a UserOfficeAssignment.
    Cannot remove the current office_allocation — user must first set a different current.
    """
    if request.method != 'POST':
        return redirect('manage_admin_branches')

    user_id       = request.POST.get('user_id')
    assignment_id = request.POST.get('assignment_id')

    user       = get_object_or_404(CustomUser,           pk=user_id)
    assignment = get_object_or_404(UserOfficeAssignment, pk=assignment_id, user=user)

    # Guard: cannot remove the current branch
    if assignment.office == user.office_allocation:
        messages.error(
            request,
            f'Cannot remove "{assignment.office.name}" — it is the current branch. '
            f'Please set another branch as current first.'
        )
        return redirect(_branch_url(user_id))

    office_name = assignment.office.name
    assignment.delete()
    messages.success(request, f'Branch "{office_name}" removed successfully.')
    return redirect(_branch_url(user_id))







# ══════════════════════════════════════════════════════════════════════
#  PAGE 1 — Month picker  (GET)
# ══════════════════════════════════════════════════════════════════════

def payroll_filter(request):
    context = {
        **get_base_context(request),
    }
    return render(request, 'app/payroll_filter.html', context)


# ══════════════════════════════════════════════════════════════════════
#  PAGE 2 — Payroll table  (POST from filter)
# ══════════════════════════════════════════════════════════════════════

def payroll_report(request):
    """
    Monthly Payroll Report.

    Column definitions (matching screenshot exactly):
      Basic Salary      = user.salary
      Deduction amount  = user.deduction_amount   (fixed deduction stored on user)
      Salary Advance    = '-'  (extend if you have an advance model)
      Staff loan        = '-'  (extend if you have a staff loan model)
      NSSF              = user.nssf_amount         (per-employee NSSF stored on user)
      PAYE              = '-'  (not stored; extend if needed)
      Total Deduction   = deduction_amount + NSSF  (+ advance/loan/PAYE when present)
      Net Pay           = Basic Salary - Total Deduction
      Open Balance      = '-'  (extend as needed)
      Closing Advances  = '-'  (extend as needed)

    Already-paid staff for this month = Salary record exists with salary_for_month
    matching the 1st of the selected month. They appear pre-checked + disabled.

    Accepts:
      POST  payroll_month = 'MM-YYYY'   (from filter form)
      GET   pm            = 'MM-YYYY'   (redirect after submit, to re-show page)
    """
    payroll_month_str = ''

    if request.method == 'POST':
        payroll_month_str = request.POST.get('payroll_month', '').strip()
    else:
        # GET — used after successful submit redirect
        payroll_month_str = request.GET.get('pm', '').strip()

    if not payroll_month_str:
        return redirect('payroll_filter')

    try:
        parts = payroll_month_str.split('-')
        month = int(parts[0])
        year  = int(parts[1])
        salary_date = datetime.date(year, month, 1)
    except (ValueError, IndexError):
        return redirect('payroll_filter')

    # All active non-superuser staff ordered alphabetically
    all_staff = CustomUser.objects.filter(
        is_active=True, is_superuser=False
    ).order_by('last_name', 'first_name')

    # Staff IDs already paid this month
    paid_ids = set(
        Salary.objects.filter(
            salary_for_month=salary_date
        ).values_list('employee_id', flat=True)
    )

    staff_rows = []
    totals = {
        'basic_salary':     Decimal('0'),
        'deduction_amount': Decimal('0'),
        'salary_advance':   Decimal('0'),
        'staff_loan':       Decimal('0'),
        'nssf':             Decimal('0'),
        'paye':             Decimal('0'),
        'total_deduction':  Decimal('0'),
        'net_pay':          Decimal('0'),
        'open_balance':     Decimal('0'),
        'closing_advances': Decimal('0'),
    }

    for user in all_staff:
        basic  = user.salary or Decimal('0')
        deduct = user.deduction_amount or Decimal('0')

        # NSSF: read from nssf_amount field (add to model as shown above).
        # If you haven't added nssf_amount yet, falls back gracefully to 0.
        nssf = getattr(user, 'nssf_amount', None) or Decimal('0')

        # PAYE — not stored on model; set to 0 (extend if needed)
        paye = Decimal('0')

        # Salary advance / staff loan — extend if you have those models
        salary_advance = Decimal('0')
        staff_loan     = Decimal('0')

        total_deduction = deduct + nssf + paye + salary_advance + staff_loan
        net_pay         = max(basic - total_deduction, Decimal('0'))

        already_paid = user.id in paid_ids

        staff_rows.append({
            'user':             user,
            'basic_salary':     basic,
            'deduction_amount': deduct,
            'salary_advance':   salary_advance,  # Decimal 0 → shown as '-' in template
            'staff_loan':       staff_loan,
            'nssf':             nssf,
            'paye':             paye,
            'total_deduction':  total_deduction,
            'net_pay':          net_pay,
            'open_balance':     Decimal('0'),    # extend as needed
            'closing_advances': Decimal('0'),    # extend as needed
            'already_paid':     already_paid,
        })

        # Grand totals accumulate ALL rows (paid and unpaid)
        totals['basic_salary']     += basic
        totals['deduction_amount'] += deduct
        totals['salary_advance']   += salary_advance
        totals['staff_loan']       += staff_loan
        totals['nssf']             += nssf
        totals['paye']             += paye
        totals['total_deduction']  += total_deduction
        totals['net_pay']          += net_pay

    return render(request, 'app/payroll_report.html', {
        'payroll_month': payroll_month_str,
        'salary_date':   salary_date,
        'staff_rows':    staff_rows,
        'totals':        totals,
        **get_base_context(request),
    })


# ══════════════════════════════════════════════════════════════════════
#  SUBMIT — create Salary records for checked staff  (POST)
# ══════════════════════════════════════════════════════════════════════

def payroll_submit(request):
    """
    POST: Create Salary records for each selected (unpaid) staff member.

    Salary.amount is stored as NET PAY (Basic - Total Deductions).
    fund_source is resolved from the currently selected office in the session:
      - Uses filter_office if a specific branch is selected
      - Falls back to selected_office if HQ is active (filter_office is None)
    After each payment, bank_balance of fund_source branch is deducted.
    Already-paid staff are silently skipped (double-payment safe).
    """
    if request.method != 'POST':
        return redirect('payroll_filter')

    payroll_month_str = request.POST.get('payroll_month', '').strip()
    pay_ids           = request.POST.getlist('pay_ids')
    txn_date_str      = request.POST.get('transaction_date', '')

    if not payroll_month_str:
        return redirect('payroll_filter')

    try:
        parts = payroll_month_str.split('-')
        month = int(parts[0])
        year  = int(parts[1])
        salary_date = datetime.date(year, month, 1)
    except (ValueError, IndexError):
        return redirect('payroll_filter')

    try:
        txn_date = datetime.datetime.strptime(txn_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        txn_date = datetime.date.today()

    if not pay_ids:
        messages.warning(request, 'No staff selected. Please check at least one staff member.')
        return redirect(reverse('payroll_report') + f'?pm={payroll_month_str}')

    # ── Resolve fund source from selected office in session ───────────────
    # filter_office is None when HQ is selected (by design in get_base_context)
    # so we fall back to selected_office which is always the actual Office object
    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']    # None when HQ selected
    selected_office = base_ctx['selected_office']  # always the actual Office object

    fund_source = filter_office or selected_office

    if not fund_source:
        messages.error(request, 'Please select a branch before processing payroll.')
        return redirect(reverse('payroll_report') + f'?pm={payroll_month_str}')

    paid_count = 0
    skipped    = 0

    for uid in pay_ids:
        try:
            user = CustomUser.objects.get(pk=uid, is_active=True)
        except CustomUser.DoesNotExist:
            continue

        # Double-payment guard
        if Salary.objects.filter(employee=user, salary_for_month=salary_date).exists():
            skipped += 1
            continue

        basic  = user.salary or Decimal('0')
        deduct = user.deduction_amount or Decimal('0')
        nssf   = getattr(user, 'nssf_amount', None) or Decimal('0')
        paye   = Decimal('0')
        net    = max(basic - deduct - nssf - paye, Decimal('0'))

        Salary.objects.create(
            employee           = user,
            amount             = net,
            salary_for_month   = salary_date,
            payment_date       = txn_date,
            transaction_method = 'bank',
            fund_source        = fund_source,
            processed_by       = request.user,
        )

        # ── Deduct net salary from branch bank balance ─────────────────
        latest = BranchBalance.objects.filter(
            branch=fund_source
        ).order_by('-last_updated').first()

        current_office_bal = latest.office_balance if latest else Decimal('0')
        current_bank_bal   = latest.bank_balance   if latest else Decimal('0')

        BranchBalance.objects.create(
            branch         = fund_source,
            office_balance = current_office_bal,
            bank_balance   = max(current_bank_bal - net, Decimal('0')),
            updated_by     = request.user,
        )

        paid_count += 1

    if paid_count:
        messages.success(
            request,
            f'✓ Salary processed for {paid_count} staff member(s) '
            f'for {payroll_month_str}.'
        )
    if skipped:
        messages.info(
            request,
            f'{skipped} staff member(s) already paid this month — skipped.'
        )
    if not paid_count and not skipped:
        messages.warning(request, 'No valid staff IDs were processed.')

    return redirect(reverse('payroll_report') + f'?pm={payroll_month_str}')







# ══════════════════════════════════════════════════════════════════════
#  LIST — with search, from/to branch filter, pagination, page size
# ══════════════════════════════════════════════════════════════════════
from django.core.paginator import Paginator
def office_transaction_list(request):
    """
    Office Transaction list page.

    Query params:
      search      — text filter on description, processed_by name, receipt/id
      from_branch — Office.id to filter office_from
      to_branch   — Office.id to filter office_to
      page        — current page number (default 1)
      page_size   — rows per page (default 10; options 10/25/50/100)

    Columns:
      Date | Name (processed_by) | Description | Credit | Debit |
      Receipt (id as 000001) | Department (transaction_method) |
      From Branch | To Branch | Action (delete)

    The "Add Office Transaction" button renders without a URL — wire it
    via openAddTransactionModal() in the template JS.
    """
    search_query  = request.GET.get('search',      '').strip()
    from_branch_id= request.GET.get('from_branch', '').strip()
    to_branch_id  = request.GET.get('to_branch',   '').strip()
    page_num      = request.GET.get('page',         1)
    page_size_str = request.GET.get('page_size',   '10')

    try:
        page_size = int(page_size_str)
        if page_size not in (10, 25, 50, 100):
            page_size = 10
    except ValueError:
        page_size = 10

    qs = OfficeTransaction.objects.select_related(
        'office_from', 'office_to', 'processed_by'
    ).order_by('-transaction_date', '-id')

    # ── Text search ────────────────────────────────────────────────
    if search_query:
        qs = qs.filter(
            Q(description__icontains=search_query) |
            Q(processed_by__first_name__icontains=search_query) |
            Q(processed_by__last_name__icontains=search_query)  |
            Q(office_from__name__icontains=search_query) |
            Q(office_to__name__icontains=search_query)
        )

    # ── Branch filters ─────────────────────────────────────────────
    if from_branch_id:
        qs = qs.filter(office_from_id=from_branch_id)

    if to_branch_id:
        qs = qs.filter(office_to_id=to_branch_id)

    # ── Pagination ─────────────────────────────────────────────────
    total_count = qs.count()
    paginator   = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page_num)
    except Exception:
        page_obj = paginator.page(1)

    # Build smart page range: 1 2 3 … 10 11 12 … 35 36
    page_range = _build_page_range(paginator.num_pages, page_obj.number)

    all_offices = Office.objects.order_by('name')

    return render(request, 'app/office_transaction_list.html', {
        'transactions':   page_obj.object_list,
        'page_obj':       page_obj,
        'page_range':     page_range,
        'total_count':    total_count,
        'page_size':      page_size,
        'all_offices':    all_offices,
        'search_query':   search_query,
        'from_branch_id': from_branch_id,
        'to_branch_id':   to_branch_id,
        **get_base_context(request),
    })


# ══════════════════════════════════════════════════════════════════════
#  DELETE
# ══════════════════════════════════════════════════════════════════════

# def office_transaction_delete(request, pk):
#     """Delete an OfficeTransaction and redirect back to the list."""
#     if request.method == 'POST':
#         txn = get_object_or_404(OfficeTransaction, pk=pk)
#         txn.delete()
#         messages.success(request, f'Transaction #{pk:06d} deleted.')
#     return redirect('office_transaction_list')

def office_transaction_delete(request, pk):
    """Delete an OfficeTransaction and reverse its balance effect on both branches."""
    if request.method == 'POST':
        txn = get_object_or_404(OfficeTransaction, pk=pk)

        amount = txn.amount
        method = txn.transaction_method  # 'cash' or 'bank'
        office_from = txn.office_from
        office_to   = txn.office_to

        with transaction.atomic():
            # ── Restore sender (office_from) ──────────────────────────────
            balance_from = BranchBalance.objects.select_for_update().filter(branch=office_from).last()
            if balance_from:
                BranchBalance.objects.create(
                    branch=office_from,
                    office_balance=(
                        balance_from.office_balance + amount
                        if method == 'cash'
                        else balance_from.office_balance
                    ),
                    bank_balance=(
                        balance_from.bank_balance + amount
                        if method == 'bank'
                        else balance_from.bank_balance
                    ),
                    updated_by=request.user,
                )

            # ── Deduct from receiver (office_to) ──────────────────────────
            balance_to = BranchBalance.objects.select_for_update().filter(branch=office_to).last()
            if balance_to:
                BranchBalance.objects.create(
                    branch=office_to,
                    office_balance=(
                        balance_to.office_balance - amount
                        if method == 'cash'
                        else balance_to.office_balance
                    ),
                    bank_balance=(
                        balance_to.bank_balance - amount
                        if method == 'bank'
                        else balance_to.bank_balance
                    ),
                    updated_by=request.user,
                )

            # ── Delete the transaction record ──────────────────────────────
            txn.delete()

        messages.success(
            request,
            f'Transaction #{pk:06d} deleted. '
            f'TZS {amount:,.0f} ({method.title()}) reversed: '
            f'returned to {office_from.name}, deducted from {office_to.name}.'
        )

    return redirect('office_transaction_list')


# ══════════════════════════════════════════════════════════════════════
#  Helper: smart page range builder
# ══════════════════════════════════════════════════════════════════════
def _build_page_range(num_pages, current):
    """
    Returns a list of page numbers and '...' strings for the paginator.
    Always shows first/last 2 pages and 2 around current.
    """
    pages = set()
    pages.update([1, 2])
    pages.update([num_pages - 1, num_pages])
    pages.update([current - 1, current, current + 1])
    pages = sorted(p for p in pages if 1 <= p <= num_pages)

    result = []
    prev = None
    for p in pages:
        if prev is not None and p - prev > 1:
            result.append('...')
        result.append(p)
        prev = p
    return result










# ══════════════════════════════════════════════════════════════════════
#  LOANS OWED
# ══════════════════════════════════════════════════════════════════════

MONTH_NAMES = {
    1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun',
    7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'
}

def _label(y, m):
    return f"{MONTH_NAMES[m]}-{y}"

def _months_range(sy, sm, ey, em):
    """Inclusive list of (year, month) from (sy,sm) to (ey,em)."""
    out, y, m = [], sy, sm
    while (y, m) <= (ey, em):
        out.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


def loans_owed_summary(request):
    base_ctx      = get_base_context(request)
    filter_office = base_ctx['filter_office']

    qs = LoanApplication.objects.filter(
        status='Approved',
        repayment_amount_remaining__gt=0,
    )
    if filter_office:
        qs = qs.filter(office=filter_office.name)

    cash_count = qs.count()

    return render(request, 'app/loans_owed_summary.html', {
        **base_ctx,
        'total_loans_owed': cash_count,
        'cash_loans_count': cash_count,
    })




# ================================================================================
def _months_range(sy, sm, ey, em):
    months = []
    y, m = sy, sm
    while (y, m) <= (ey, em):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months


def _label(y, m):
    import datetime
    return datetime.date(y, m, 1).strftime('%b-%Y')


def _get_installment_amounts(inst, period, loan=None):
    """
    Returns a list of per-slot installment amounts applying the
    first-installment-largest rounding strategy (floor to nearest 1,000).
    If loan object is provided, uses loan_amount and total_interest_amount
    for precise calculation. Falls back to flat inst if not available.
    """
    from decimal import Decimal, ROUND_DOWN

    if not period or period <= 0:
        return [inst] * (period or 1)

    # Try to get precise P and total_interest from loan object
    if loan is not None:
        P              = loan.loan_amount or Decimal('0')
        total_interest = loan.total_interest_amount or Decimal('0')
    else:
        # Fallback: reconstruct from inst * period (loses precision)
        total_return   = inst * period
        # We can't split P and interest without more info, so use flat
        return [inst] * period

    def floor_1000(val):
        return (val / Decimal('1000')).to_integral_value(rounding=ROUND_DOWN) * Decimal('1000')

    rounded_installment = floor_1000((P + total_interest) / Decimal(period))

    # First installment absorbs remainder
    first_installment = (P + total_interest) - (rounded_installment * (period - 1))

    amounts = [first_installment] + [rounded_installment] * (period - 1)
    return amounts


def _get_topup_lump(loan, rmap):
    """
    Returns (lump_amount, lump_year, lump_month) if this loan has a topup
    whose old_balance_cleared was recorded as a LoanRepayment, else None.
    """
    from decimal import Decimal

    topup = loan.topups.order_by('topup_date').first()
    if not topup or not topup.old_balance_cleared:
        return None

    # Find the repayment record matching the topup clearing amount on topup date
    lump_y = topup.topup_date.year
    lump_m = topup.topup_date.month
    lump_amount = topup.old_balance_cleared

    # Confirm it exists in rmap (year, month) bucket
    if rmap.get((lump_y, lump_m), 0) >= lump_amount:
        return (lump_amount, lump_y, lump_m)

    return None


# def _distribute_payments(col_months, rmap, inst, frd, period, cur_y, cur_m, loan=None):
#     from decimal import Decimal

#     # ── Detect topup lump ────────────────────────────────────────────
#     topup_info = None
#     if loan is not None:
#         topup_info = _get_topup_lump(loan, rmap)

#     # ── Build adjusted rmap excluding the topup lump amount ─────────
#     clean_rmap = dict(rmap)
#     if topup_info:
#         lump_amount, lump_y, lump_m = topup_info
#         key = (lump_y, lump_m)
#         remaining_after_lump = clean_rmap.get(key, Decimal('0')) - lump_amount
#         if remaining_after_lump <= Decimal('0'):
#             clean_rmap.pop(key, None)
#         else:
#             clean_rmap[key] = remaining_after_lump

#     # ── Get per-slot installment amounts ────────────────────────────
#     slot_amounts = _get_installment_amounts(inst, period, loan=loan)

#     # ── Pool remaining (non-lump) payments ──────────────────────────
#     total_paid_pool = sum(clean_rmap.values(), Decimal('0'))

#     # ── Find lump_slot_offset: first slot NOT covered by normal payments
#     lump_slot_offset = None
#     if topup_info:
#         temp_remaining = total_paid_pool
#         lump_slot_offset = 0  # default
#         for offset in range(period):
#             slot_inst = slot_amounts[offset]
#             if slot_inst > 0 and temp_remaining >= slot_inst:
#                 temp_remaining -= slot_inst
#                 lump_slot_offset = offset + 1
#             else:
#                 lump_slot_offset = offset
#                 break

#     # ── Fill adjusted slots (only needed for normal loans) ──────────
#     adjusted = {}
#     if not topup_info:
#         remaining = total_paid_pool
#         for offset in range(period):
#             slot_inst = slot_amounts[offset]
#             if slot_inst > 0:
#                 applied   = min(remaining, slot_inst)
#                 remaining = max(remaining - slot_inst, Decimal('0'))
#             else:
#                 applied = Decimal('0')
#             adjusted[offset] = applied

#     # ── Build one cell per column month ─────────────────────────────
#     cells = []
#     col_totals_contribution = []

#     for (cy, cm) in col_months:
#         if not frd:
#             cells.append({'type': 'empty'})
#             col_totals_contribution.append((Decimal('0'), Decimal('0')))
#             continue

#         offset = (cy - frd.year) * 12 + (cm - frd.month)

#         if offset < 0 or offset >= period:
#             cells.append({'type': 'empty'})
#             col_totals_contribution.append((Decimal('0'), Decimal('0')))
#             continue

#         slot_inst = slot_amounts[offset]
#         past      = (cy, cm) <= (cur_y, cur_m)

#         # ════════════════════════════════════════════════════════════
#         # TOPUP LOAN RENDERING
#         # ════════════════════════════════════════════════════════════
#         if topup_info and lump_slot_offset is not None:

#             # ── Slots BEFORE lump — normal tick with actual paid value
#             if offset < lump_slot_offset:
#                 cells.append({
#                     'type': 'tick',
#                     'paid': slot_amounts[offset],
#                     'out':  slot_inst,
#                 })
#                 col_totals_contribution.append((slot_amounts[offset], Decimal('0')))

#             # ── LUMP SLOT — show actual lump amount ──────────────────
#             elif offset == lump_slot_offset:
#                 lump_amount = topup_info[0]
#                 shortfall   = max(slot_inst - lump_amount, Decimal('0'))
#                 if lump_amount >= slot_inst:
#                     cells.append({
#                         'type': 'tick',
#                         'paid': lump_amount,
#                         'out':  slot_inst,
#                     })
#                     col_totals_contribution.append((lump_amount, Decimal('0')))
#                 else:
#                     cells.append({
#                         'type': 'partial',
#                         'paid': lump_amount,
#                         'out':  shortfall,
#                     })
#                     col_totals_contribution.append((lump_amount, shortfall))

#             # ── ALL slots AFTER lump — tick with empty paid ──────────
#             else:
#                 cells.append({
#                     'type': 'tick',
#                     'paid': Decimal('0'),
#                     'out':  slot_inst,
#                 })
#                 col_totals_contribution.append((Decimal('0'), Decimal('0')))

#             continue

#         # ════════════════════════════════════════════════════════════
#         # NORMAL LOAN RENDERING
#         # ════════════════════════════════════════════════════════════
#         paid_this = adjusted.get(offset, Decimal('0'))
#         shortfall = max(slot_inst - paid_this, Decimal('0'))

#         if slot_inst <= 0:
#             cells.append({'type': 'empty'})
#             col_totals_contribution.append((Decimal('0'), Decimal('0')))

#         elif paid_this >= slot_inst:
#             cells.append({
#                 'type': 'tick',
#                 'paid': paid_this,
#                 'out':  slot_inst,
#             })
#             col_totals_contribution.append((paid_this, Decimal('0')))

#         elif paid_this > 0:
#             cells.append({
#                 'type': 'partial',
#                 'paid': paid_this,
#                 'out':  shortfall,
#             })
#             col_totals_contribution.append((paid_this, shortfall))

#         elif past:
#             cells.append({
#                 'type': 'out_only',
#                 'paid': Decimal('0'),
#                 'out':  slot_inst,
#             })
#             col_totals_contribution.append((Decimal('0'), slot_inst))

#         else:
#             cells.append({
#                 'type': 'future_out',
#                 'paid': Decimal('0'),
#                 'out':  slot_inst,
#             })
#             col_totals_contribution.append((Decimal('0'), Decimal('0')))

#     return cells, col_totals_contribution




from decimal import Decimal
import datetime
from dateutil.relativedelta import relativedelta


def _months_range(sy, sm, ey, em):
    result = []
    cur = datetime.date(sy, sm, 1)
    end = datetime.date(ey, em, 1)
    while cur <= end:
        result.append((cur.year, cur.month))
        cur += relativedelta(months=1)
    return result


def _label(y, m):
    return datetime.date(y, m, 1).strftime('%b-%Y').upper()


def _ceil1000(val):
    import math
    return Decimal(math.ceil(float(val) / 1000) * 1000)


def _get_topup_lump(loan):
    topups = list(loan.topups.order_by('topup_date'))
    if not topups:
        return None
    topup = topups[-1]
    lump_amount = topup.old_balance_cleared
    ref_date = topup.payment_month or topup.topup_date
    return (lump_amount, ref_date.year, ref_date.month)


def _get_installment_amounts(loan):
    period = loan.payment_period_months or 0
    if period <= 0:
        return []

    P              = loan.loan_amount            or Decimal('0')
    total_interest = loan.total_interest_amount  or Decimal('0')
    total_return   = loan.total_repayment_amount or (P + total_interest)

    if period == 1:
        return [total_return]

    std_monthly   = _ceil1000(total_return / period)
    std_principal = _ceil1000(P            / period)
    std_interest  = std_monthly - std_principal

    last_principal = P              - std_principal * (period - 1)
    last_interest  = total_interest - std_interest  * (period - 1)
    last_monthly   = last_principal + last_interest

    return [std_monthly] * (period - 1) + [last_monthly]


def _get_topup_lump(loan):
    topups = list(loan.topups.order_by('topup_date'))
    if not topups:
        return None
    topup = topups[-1]
    lump_amount = Decimal(topup.old_balance_cleared or 0)
    if lump_amount <= 0:
        return None
    ref_date = topup.payment_month or topup.topup_date
    if ref_date is None:
        return None
    return (lump_amount, ref_date.year, ref_date.month)


def _distribute_payments(col_months, rmap=None, inst=Decimal('0'), frd=None,
                         period=12, cur_y=None, cur_m=None, loan=None):
    """
    SHERIA:
    ─────────────────────────────────────────────────────────────────
    • Kila rejesho → funika nyuma kwanza, kisha mwezi wake, kisha mbele.
    • covered >= slot_inst → TICK (out=0).
    • Spillover slots → TICK tu, paid=0.
    • Mwezi wa malipo halisi → onyesha paid + out_baki ya slot hiyo.
    • out_baki = slot_inst - covered (si global balance).
    • Topup month → onyesha old_balance_cleared (si topup loan amount).
    • Topup month entry kwenye rmap inaondolewa — topup_lump inashughulikia.
    • Malipo nje ya period → DAIMA onyesha paid (hata kama bado kuna balance).
    ─────────────────────────────────────────────────────────────────
    """
    rmap  = rmap or {}
    today = datetime.date.today()

    # ── 1. FRD ───────────────────────────────────────────────────────────────
    if frd is None:
        if loan is not None:
            try:
                frd = _get_first_repayment_date(loan)
            except Exception:
                pass
        if frd is None:
            if rmap:
                first_ym = sorted(rmap.keys())[0]
                frd = datetime.date(first_ym[0], first_ym[1], 1)
            elif col_months:
                first_ym = sorted(col_months)[0]
                frd = datetime.date(first_ym[0], first_ym[1], 1)
            else:
                frd = today

    inst   = Decimal(inst)
    period = max(int(period), 1)
    cur_y  = int(cur_y or frd.year)
    cur_m  = int(cur_m or frd.month)

    # ── 2. SLOT AMOUNTS ───────────────────────────────────────────────────────
    if loan is not None:
        try:
            slot_amounts = [Decimal(x) for x in _get_installment_amounts(loan)]
        except Exception:
            slot_amounts = [inst] * period
    else:
        slot_amounts = [inst] * period

    while len(slot_amounts) < period:
        slot_amounts.append(inst)
    slot_amounts = slot_amounts[:period]

    total_loan_amount = sum(slot_amounts)

    # ── 3. TOPUP INFO ─────────────────────────────────────────────────────────
    topup_lump  = Decimal('0')
    topup_month = None

    if loan is not None:
        try:
            topup_info = _get_topup_lump(loan)
            if topup_info:
                amt, ty, tm = topup_info
                topup_lump  = Decimal(amt)
                topup_month = (int(ty), int(tm))
        except Exception:
            pass

    # ── 4. OFFSET HELPER ──────────────────────────────────────────────────────
    def ym_to_offset(y, m):
        return (int(y) - frd.year) * 12 + (int(m) - frd.month)

    # ── 5. NORMALIZE RMAP KEYS → (int, int) tuples ───────────────────────────
    clean_rmap = {}
    for key, val in rmap.items():
        try:
            amt = Decimal(val)
        except Exception:
            continue
        if amt <= 0:
            continue
        if isinstance(key, (list, tuple)) and len(key) == 2:
            nk = (int(key[0]), int(key[1]))
            clean_rmap[nk] = clean_rmap.get(nk, Decimal('0')) + amt

    # ── MUHIMU: Ondoa topup_month kutoka clean_rmap ───────────────────────────
    if topup_month is not None:
        clean_rmap.pop(topup_month, None)

    # ── 6. ALL PAYMENT EVENTS (clean_rmap + topup_lump) ──────────────────────
    all_payment_events = dict(clean_rmap)
    if topup_month is not None and topup_lump > 0:
        all_payment_events[topup_month] = (
            all_payment_events.get(topup_month, Decimal('0')) + topup_lump
        )

    sorted_payments = sorted(
        all_payment_events.items(),
        key=lambda x: (int(x[0][0]), int(x[0][1]))
    )

    # ── 7. RUNNING BALANCE (global, chronological) ────────────────────────────
    running_balance = total_loan_amount
    balance_after   = {}
    for (py, pm), paid in sorted_payments:
        running_balance         = max(running_balance - paid, Decimal('0'))
        balance_after[(py, pm)] = running_balance

    total_paid_ever = sum(v for _, v in sorted_payments)
    loan_fully_paid = total_paid_ever >= total_loan_amount

    # ── 8. SLOT COVERAGE ──────────────────────────────────────────────────────
    slot_covered = {i: Decimal('0') for i in range(period)}

    # Gawanya clean_rmap kwenye slots
    for (py, pm), pay_amt in sorted(clean_rmap.items()):
        pay_offset    = ym_to_offset(py, pm)
        capped_offset = max(0, min(pay_offset, period - 1))
        remaining     = pay_amt

        for off in range(capped_offset):
            if remaining <= 0: break
            deficit = slot_amounts[off] - slot_covered[off]
            if deficit <= 0: continue
            take = min(deficit, remaining)
            slot_covered[off] += take
            remaining         -= take

        for off in range(capped_offset, period):
            if remaining <= 0: break
            deficit = slot_amounts[off] - slot_covered[off]
            if deficit <= 0: continue
            take = min(deficit, remaining)
            slot_covered[off] += take
            remaining         -= take

    # Gawanya topup_lump kwenye slots
    if topup_month is not None and topup_lump > 0:
        topup_offset = max(0, min(
            ym_to_offset(topup_month[0], topup_month[1]), period - 1
        ))
        remaining = topup_lump

        for off in range(topup_offset):
            if remaining <= 0: break
            deficit = slot_amounts[off] - slot_covered[off]
            if deficit <= 0: continue
            take = min(deficit, remaining)
            slot_covered[off] += take
            remaining         -= take

        for off in range(topup_offset, period):
            if remaining <= 0: break
            deficit = slot_amounts[off] - slot_covered[off]
            if deficit <= 0: continue
            take = min(deficit, remaining)
            slot_covered[off] += take
            remaining         -= take

    # ── 9. JENGA CELLS ────────────────────────────────────────────────────────
    cells                   = []
    col_totals_contribution = []

    for (cy, cm) in col_months:
        cy, cm  = int(cy), int(cm)
        offset  = ym_to_offset(cy, cm)
        past    = (cy, cm) <= (cur_y, cur_m)

        rmap_this_month = clean_rmap.get((cy, cm), Decimal('0'))
        paid_this_month = all_payment_events.get((cy, cm), Decimal('0'))

        # Balance kabla ya mwezi huu
        prev_months = [
            (int(py), int(pm))
            for (py, pm), _ in sorted_payments
            if (int(py), int(pm)) < (cy, cm)
        ]
        balance_before = (
            balance_after[prev_months[-1]] if prev_months else total_loan_amount
        )
        balance_this = balance_after.get((cy, cm), balance_before)

        # ── NDANI YA PERIOD ───────────────────────────────────────────────────
        if 0 <= offset < period:
            slot_inst = slot_amounts[offset]
            covered   = slot_covered.get(offset, Decimal('0'))
            out_baki  = max(slot_inst - covered, Decimal('0'))

            if (topup_month is not None and (cy, cm) == topup_month):
                paid_display = topup_lump
            elif rmap_this_month > 0:
                paid_display = rmap_this_month
            else:
                paid_display = Decimal('0')

            if covered >= slot_inst:
                cells.append({
                    'type': 'tick',
                    'paid': paid_display,
                    'out' : Decimal('0'),
                })
                col_totals_contribution.append((rmap_this_month, Decimal('0')))

            elif paid_display > 0 or covered > 0:
                cells.append({
                    'type': 'partial',
                    'paid': paid_display,
                    'out' : out_baki,
                })
                col_totals_contribution.append((rmap_this_month, out_baki))

            elif past:
                cells.append({
                    'type': 'out_only',
                    'paid': Decimal('0'),
                    'out' : slot_inst,
                })
                col_totals_contribution.append((Decimal('0'), slot_inst))

            else:
                cells.append({
                    'type': 'future_out',
                    'paid': Decimal('0'),
                    'out' : slot_inst,
                })
                col_totals_contribution.append((Decimal('0'), slot_inst))

        # ── NJE YA PERIOD ─────────────────────────────────────────────────────
        else:
            if paid_this_month > 0:
                # ═══════════════════════════════════════════════════════════════
                # FIX: Malipo nje ya period DAIMA yanaonyeshwa kwenye Paid column.
                # Awali code ilikuwa inaweka tick tu kama loan_fully_paid au
                # balance_this <= 0, na extra_paid kama bado kuna balance —
                # lakini katika hali zote mbili, paid_this_month ilionekana.
                #
                # Tatizo lilikuwa: kama `paid_this_month` ilikuwa inatoka kwa
                # `all_payment_events` (ambayo ni 120k kwa JAN-2027) lakini
                # `rmap_this_month` (kutoka clean_rmap) ilikuwa 0 kwa sababu
                # fulani (k.m. key mismatch au topup logic), basi
                # col_totals_contribution iliongeza (0, ...) badala ya (120k, ...).
                #
                # Sasa tunatumia `paid_this_month` moja kwa moja kwa display
                # na `rmap_this_month` kwa col_totals (ili totals row iwe sahihi).
                # ═══════════════════════════════════════════════════════════════
                if balance_this <= Decimal('0') or loan_fully_paid:
                    cells.append({
                        'type': 'tick',
                        'paid': paid_this_month,
                        'out' : Decimal('0'),
                    })
                    col_totals_contribution.append((rmap_this_month, Decimal('0')))
                else:
                    # Malipo nje ya period → onyesha paid tu, out = 0 (si global balance)
                    cells.append({
                        'type': 'extra_paid',
                        'paid': paid_this_month,
                        'out' : Decimal('0'),
                    })
                    col_totals_contribution.append((rmap_this_month, Decimal('0')))

            elif balance_before > Decimal('0') and past and offset >= period:
                cells.append({
                    'type': 'out_only',
                    'paid': Decimal('0'),
                    'out' : balance_before,
                })
                col_totals_contribution.append((Decimal('0'), balance_before))

            else:
                cells.append({'type': 'empty'})
                col_totals_contribution.append((Decimal('0'), Decimal('0')))

    return cells, col_totals_contribution






def _distribute_payments_hama(col_months, rmap, inst, frd, period, cur_y, cur_m, bf_y, bf_m):
    if not col_months or not frd:
        return [], []

    if period > 1:
        total_return = inst * period
        std_monthly  = _ceil1000(total_return / period)
        last_monthly = total_return - std_monthly * (period - 1)
        slot_amounts = [std_monthly] * (period - 1) + [last_monthly]
    elif period == 1:
        slot_amounts = [inst]
    else:
        slot_amounts = []

    total_inst_sum = sum(slot_amounts) if slot_amounts else inst * period

    paid_upto_dec  = Decimal('0')
    paid_after_dec = {}

    for (cy, cm), amount in rmap.items():
        if (cy, cm) <= (bf_y, bf_m):
            paid_upto_dec += amount
        else:
            paid_after_dec[(cy, cm)] = (
                paid_after_dec.get((cy, cm), Decimal('0')) + amount
            )

    # ── total paid across ALL months (before + after December) ──
    total_paid_all = paid_upto_dec + sum(paid_after_dec.values())

    # ── December outstanding = what still remains after ALL payments ──
    out_brought_forward = max(total_inst_sum - total_paid_all, Decimal('0'))

    cells = []
    col_totals_contribution = []

    for idx, (cy, cm) in enumerate(col_months):
        if idx == 0:
            cells.append({
                'type': 'partial',
                'paid': paid_upto_dec,
                'out':  out_brought_forward,   # ← now reflects all payments
            })
            col_totals_contribution.append((paid_upto_dec, out_brought_forward))
        else:
            paid_this = paid_after_dec.get((cy, cm), Decimal('0'))
            if paid_this > 0:
                cells.append({'type': 'paid_only', 'paid': paid_this, 'out': Decimal('0')})
                col_totals_contribution.append((paid_this, Decimal('0')))
            else:
                cells.append({'type': 'empty'})
                col_totals_contribution.append((Decimal('0'), Decimal('0')))

    return cells, col_totals_contribution


def loans_owed_report(request):
    base_ctx        = get_base_context(request)
    filter_office   = base_ctx['filter_office']
    selected_office = base_ctx['selected_office']
    branch_name     = selected_office.name.upper() if selected_office else 'All Branches'

    today        = datetime.date.today()
    cur_y, cur_m = today.year, today.month

    win_start    = today - relativedelta(months=6)
    win_y, win_m = win_start.year, win_start.month

    loans_qs = LoanApplication.objects.filter(
        is_approved=False,
    ).select_related('client').prefetch_related('topups')

    if filter_office:
        loans_qs = loans_qs.filter(office=filter_office.name)

    loans_list = list(loans_qs)

    # ── Build rep_map using payment_month as authoritative key ──
    rep_map = {l.id: {} for l in loans_list}

    repayments = LoanRepayment.objects.filter(
        loan_application__in=loans_list
    ).values(
        'loan_application_id',
        'repayment_date',
        'repayment_amount',
        'payment_month',
    )
    for r in repayments:
        # Use payment_month if set, otherwise fall back to repayment_date
        ref = r['payment_month'] or r['repayment_date']
        key = (ref.year, ref.month)
        lid = r['loan_application_id']
        rep_map[lid][key] = rep_map[lid].get(key, Decimal('0')) + r['repayment_amount']

    # ── SECTION A — active loans grouped by first_repayment_date ─
    groups = {}
    for loan in loans_list:
        frd = loan.first_repayment_date
        if not frd:
            continue
        sy, sm = frd.year, frd.month
        if (sy, sm) < (win_y, win_m):
            continue
        groups.setdefault((sy, sm), []).append(loan)

    month_sections = []

    for (sy, sm) in sorted(groups.keys(), reverse=True):
        loans_in_group = sorted(
            groups[(sy, sm)],
            key=lambda l: (l.client.lastname or '', l.client.firstname or '')
        )

        col_end    = datetime.date(sy, sm, 1) + relativedelta(months=11)
        col_months = _months_range(sy, sm, col_end.year, col_end.month)
        assert len(col_months) == 12

        month_headers = [_label(y, m) for y, m in col_months]
        col_totals    = [{'paid': Decimal('0'), 'out': Decimal('0')} for _ in col_months]

        sec_paid = sec_loaned = sec_amount = sec_balance = Decimal('0')
        rows = []

        for loan in loans_in_group:
            client = loan.client
            name   = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
            inst   = loan.monthly_installment or Decimal('0')
            rmap   = rep_map.get(loan.id, {})
            period = loan.payment_period_months or 0

            cells, contributions = _distribute_payments(
                col_months, rmap, inst,
                loan.first_repayment_date,
                period, cur_y, cur_m,
                loan=loan,
            )

            for idx, (p, o) in enumerate(contributions):
                col_totals[idx]['paid'] += p
                col_totals[idx]['out']  += o

            total_paid = sum(c.get('paid', Decimal('0')) for c in cells)
            loaned     = loan.loan_amount            or Decimal('0')
            tot_amt    = loan.total_repayment_amount or Decimal('0')
            balance    = loan.repayment_amount_remaining or Decimal('0')

            rows.append({
                'loan_id':       loan.id,
                'name':          name,
                'check_no':      client.checkno or client.employmentcardno or '—',
                'loan_id_label': f"{(loan.office or '').lower()}-{loan.id}",
                'cells':         cells,
                'total_paid':    total_paid,
                'loaned_amount': loaned,
                'total_amount':  tot_amt,
                'balance':       balance,
                'is_fully_paid': balance <= Decimal('0'),
                'mobile':        client.phonenumber or '—',
            })

            sec_paid    += total_paid
            sec_loaned  += loaned
            sec_amount  += tot_amt
            sec_balance += balance

        month_sections.append({
            'key':           f"{sy}-{sm:02d}",
            'label':         _label(sy, sm).upper(),
            'month_headers': month_headers,
            'col_totals':    col_totals,
            'rows':          rows,
            'total_paid':    sec_paid,
            'total_loaned':  sec_loaned,
            'total_amount':  sec_amount,
            'total_balance': sec_balance,
        })

    # ── SECTION B — HAMA (overdue loans) ────────────────────────
    hama_cutoff = today - relativedelta(months=6)

    hama_loans = [
        l for l in loans_list
        if l.first_repayment_date
        and l.first_repayment_date <= hama_cutoff
    ]

    bf_y, bf_m      = cur_y - 1, 12
    hama_col_months = _months_range(cur_y - 1, 12, cur_y, 12)  # Dec → Dec, always 13
    hama_month_hdrs = [_label(y, m) for y, m in hama_col_months]

    hama_year_groups = {}
    for loan in hama_loans:
        yr = loan.first_repayment_date.year
        hama_year_groups.setdefault(yr, []).append(loan)

    hama_sections = []

    for yr in sorted(hama_year_groups.keys(), reverse=True):
        h_loans = sorted(
            hama_year_groups[yr],
            key=lambda l: (l.client.lastname or '', l.client.firstname or '')
        )

        h_col_totals = [
            {'paid': Decimal('0'), 'out': Decimal('0')}
            for _ in hama_col_months
        ]
        h_paid = h_loaned = h_interest = h_amount = h_balance = Decimal('0')
        rows = []

        for loan in h_loans:
            client = loan.client
            name   = f"{client.firstname} {client.middlename or ''} {client.lastname}".strip()
            inst   = loan.monthly_installment or Decimal('0')
            rmap   = rep_map.get(loan.id, {})
            period = loan.payment_period_months or 0

            cells, contributions = _distribute_payments_hama(
                hama_col_months, rmap, inst,
                loan.first_repayment_date,
                period, cur_y, cur_m,
                bf_y, bf_m,
            )

            for idx, (p, o) in enumerate(contributions):
                h_col_totals[idx]['paid'] += p
                h_col_totals[idx]['out']  += o

            total_paid = sum(c.get('paid', Decimal('0')) for c in cells)
            loaned     = loan.loan_amount            or Decimal('0')
            interest   = loan.total_interest_amount  or Decimal('0')
            tot_amt    = loan.total_repayment_amount or Decimal('0')
            balance    = loan.repayment_amount_remaining or Decimal('0')

            rows.append({
                'loan_id':       loan.id,
                'name':          name,
                'check_no':      client.checkno or client.employmentcardno or '—',
                'loan_id_label': f"{(loan.office or '').lower()}-{loan.id}",
                'cells':         cells,
                'total_paid':    total_paid,
                'loaned_amount': loaned,
                'interest':      interest,
                'total_amount':  tot_amt,
                'balance':       balance,
                'mobile':        client.phonenumber or '—',
            })

            h_paid     += total_paid
            h_loaned   += loaned
            h_interest += interest
            h_amount   += tot_amt
            h_balance  += balance

        hama_sections.append({
            'label':          f"HAMA-{yr}",
            'month_headers':  hama_month_hdrs,
            'col_totals':     h_col_totals,
            'rows':           rows,
            'total_paid':     h_paid,
            'total_loaned':   h_loaned,
            'total_interest': h_interest,
            'total_amount':   h_amount,
            'total_balance':  h_balance,
        })

    return render(request, 'app/loans_owed_report.html', {
        **base_ctx,
        'branch_name':    branch_name,
        'month_sections': month_sections,
        'hama_sections':  hama_sections,
    })


def loans_owed_approve(request):
    if request.method == 'POST':
        loan_ids = request.POST.getlist('loan_ids')
        if loan_ids:
            # Only approve loans that are fully paid (remaining balance = 0)
            approved = LoanApplication.objects.filter(
                id__in=loan_ids,
                repayment_amount_remaining__lte=Decimal('0'),
            ).update(is_approved=True)
            if approved:
                messages.success(request, f'{approved} loan(s) approved and removed from report.')
            else:
                messages.warning(request, 'Selected loans are not fully paid yet.')
        else:
            messages.warning(request, 'No loans were selected.')
    return redirect('loans_owed_report')






def loan_edit(request, loan_id):
    """
    Edit an existing loan application.
    Recalculates all derived fields (interest, repayment, schedule)
    whenever core fields change, exactly as the model's save() does.
    """
    loan = get_object_or_404(LoanApplication, id=loan_id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # ── Parse submitted values ────────────────────────────
                loan_amount_raw   = request.POST.get('loan_amount', '0').replace(',', '').strip()
                loan_amount       = Decimal(loan_amount_raw)
                interest_rate     = Decimal(request.POST.get('interest_rate', loan.interest_rate))
                payment_period    = int(request.POST.get('payment_period_months', loan.payment_period_months))
                loan_type         = request.POST.get('loan_type', loan.loan_type)
                loan_purpose      = request.POST.get('loan_purpose', loan.loan_purpose or '')
                transaction_method= request.POST.get('transaction_method', loan.transaction_method)
                application_date_str = request.POST.get('application_date')

                application_date = (
                    datetime.datetime.strptime(application_date_str, '%Y-%m-%d').date()
                    if application_date_str
                    else loan.application_date
                )

                # ── Recalculate derived fields ────────────────────────
                P = loan_amount
                I = interest_rate
                N = Decimal(str(payment_period))

                total_interest  = (I / Decimal('100')) * P
                total_repayment = P + total_interest
                monthly         = total_repayment / N

                # Recalculate first repayment date from new application_date
                if application_date.day <= 18:
                    first_repayment_date = application_date.replace(day=28)
                else:
                    first_repayment_date = (
                        application_date + relativedelta(months=1)
                    ).replace(day=28)

                # ── How much has already been paid ────────────────────
                total_paid = loan.repayments.aggregate(
                    s=models.Sum('repayment_amount')
                )['s'] or Decimal('0.00')

                new_remaining = max(
                    total_repayment.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP) - total_paid,
                    Decimal('0.00')
                )

                # ── Update loan fields ────────────────────────────────
                loan.loan_amount            = loan_amount
                loan.interest_rate          = interest_rate
                loan.payment_period_months  = payment_period
                loan.loan_type              = loan_type
                loan.loan_purpose           = loan_purpose
                loan.transaction_method     = transaction_method
                loan.application_date       = application_date
                loan.first_repayment_date   = first_repayment_date

                loan.total_interest_amount  = total_interest.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                loan.interest_amount        = loan.total_interest_amount
                loan.total_repayment_amount = total_repayment.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                loan.monthly_installment    = monthly.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                loan.repayment_amount_remaining = new_remaining

                loan.save(update_fields=[
                    'loan_amount', 'interest_rate', 'payment_period_months',
                    'loan_type', 'loan_purpose', 'transaction_method',
                    'application_date', 'first_repayment_date',
                    'total_interest_amount', 'interest_amount',
                    'total_repayment_amount', 'monthly_installment',
                    'repayment_amount_remaining',
                ])

                messages.success(request, f"✅ Loan #{loan.id} updated successfully.")
                return redirect('loan_repayment_schedule', loan_id=loan.id)

        except Exception as e:
            messages.error(request, f"Error updating loan: {str(e)}")

    return render(request, 'app/loan_edit.html', {
        'loan': loan,
        **get_base_context(request),
    })
    
    
    

# ─────────────────────────────────────────────────────────────────────────────
#  CLIENT EXCEL IMPORT
# ─────────────────────────────────────────────────────────────────────────────

import io
from django.http import HttpResponse

def client_import_excel(request):
    """Upload an Excel file and bulk-create clients into the selected branch."""
    context = {**get_base_context(request)}

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Tafadhali chagua faili la Excel.")
            return render(request, 'app/client_import_excel.html', context)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            selected_office = get_selected_office(request)
            created = 0
            skipped = 0
            errors = []

            # Read header row (row 1) to map columns
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

            def col(row_values, name):
                """Return value for a given column header name."""
                try:
                    idx = headers.index(name)
                    v = row_values[idx]
                    return str(v).strip() if v is not None else ''
                except (ValueError, IndexError):
                    return ''

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue  # skip empty rows

                checkno = col(row, 'checkno')
                if checkno and Client.objects.filter(checkno=checkno).exists():
                    skipped += 1
                    errors.append(f"Mstari {row_num}: checkno '{checkno}' tayari ipo – imepigwa kando.")
                    continue

                try:
                    def to_date(val):
                        if not val:
                            return None
                        if hasattr(val, 'date'):
                            return val.date()
                        from datetime import datetime as dt
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
                            try:
                                return dt.strptime(str(val).strip(), fmt).date()
                            except ValueError:
                                pass
                        return None

                    def to_int(val):
                        try:
                            return int(val)
                        except (TypeError, ValueError):
                            return None

                    client = Client(
                        firstname=col(row, 'firstname'),
                        middlename=col(row, 'middlename') or '',
                        lastname=col(row, 'lastname'),
                        phonenumber=col(row, 'phonenumber'),
                        date_of_birth=to_date(col(row, 'date_of_birth')),
                        marital_status=col(row, 'marital_status'),
                        employername=col(row, 'employername'),
                        idara=col(row, 'idara'),
                        kaziyako=col(row, 'kaziyako'),
                        employmentcardno=col(row, 'employmentcardno'),
                        tareheya_kuajiriwa=to_date(col(row, 'tareheya_kuajiriwa')),
                        umri_kazini=to_int(col(row, 'umri_kazini')),
                        tarehe_ya_kustaafu=to_date(col(row, 'tarehe_ya_kustaafu')),
                        region=col(row, 'region'),
                        district=col(row, 'district'),
                        street=col(row, 'street'),
                        checkno=checkno,
                        mkoa=col(row, 'mkoa'),
                        wilaya=col(row, 'wilaya'),
                        tarafa=col(row, 'tarafa'),
                        kata=col(row, 'kata'),
                        mtaa=col(row, 'mtaa'),
                        wategemezi_wako=col(row, 'wategemezi_wako'),
                        bank_name=col(row, 'bank_name'),
                        bank_branch=col(row, 'bank_branch'),
                        bank_account_number=col(row, 'bank_account_number'),
                        account_name=col(row, 'account_name'),
                        account_type=col(row, 'account_type'),
                        # Mdhamini
                        mdhamini_jina_kamili=col(row, 'mdhamini_jina_kamili'),
                        mdhamini_checkno=col(row, 'mdhamini_checkno'),
                        mdhamini_kitambulisho_kazi=col(row, 'mdhamini_kitambulisho_kazi'),
                        mdhamini_kazi=col(row, 'mdhamini_kazi'),
                        mdhamini_kituo_kazi=col(row, 'mdhamini_kituo_kazi'),
                        mdhamini_kata=col(row, 'mdhamini_kata'),
                        mdhamini_tarafa=col(row, 'mdhamini_tarafa'),
                        mdhamini_wilaya=col(row, 'mdhamini_wilaya'),
                        mdhamini_mkoa=col(row, 'mdhamini_mkoa'),
                        mdhamini_simu=col(row, 'mdhamini_simu'),
                        registered_by=request.user,
                        registered_office=selected_office,
                    )
                    client.save()
                    created += 1
                except Exception as e:
                    errors.append(f"Mstari {row_num}: Hitilafu – {e}")

            context['created'] = created
            context['skipped'] = skipped
            context['errors'] = errors
            context['done'] = True
            if created:
                messages.success(request, f"✅ Wateja {created} wameingizwa kwa mafanikio.")
            if skipped:
                messages.warning(request, f"⚠️ Wateja {skipped} wamepigwa kando (checkno tayari ipo).")

        except Exception as e:
            messages.error(request, f"Hitilafu ya kusoma faili: {e}")

    return render(request, 'app/client_import_excel.html', context)


def client_excel_sample(request):
    """Download a sample Excel file with correct column headers."""
    try:
        import openpyxl
    except ImportError:
        return HttpResponse("openpyxl haipo – sakinisha kwanza.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wateja"

    headers = [
        'firstname', 'middlename', 'lastname', 'phonenumber',
        'date_of_birth', 'marital_status',
        'employername', 'idara', 'kaziyako', 'employmentcardno',
        'tareheya_kuajiriwa', 'umri_kazini', 'tarehe_ya_kustaafu',
        'region', 'district', 'street',
        'checkno', 'mkoa', 'wilaya', 'tarafa', 'kata', 'mtaa',
        'wategemezi_wako',
        'bank_name', 'bank_branch', 'bank_account_number',
        'account_name', 'account_type',
        'mdhamini_jina_kamili', 'mdhamini_checkno',
        'mdhamini_kitambulisho_kazi', 'mdhamini_kazi',
        'mdhamini_kituo_kazi', 'mdhamini_kata', 'mdhamini_tarafa',
        'mdhamini_wilaya', 'mdhamini_mkoa', 'mdhamini_simu',
    ]
    ws.append(headers)

    # Sample row
    sample = [
        'Juma', 'Ali', 'Hassan', '0712345678',
        '1985-06-15', 'Married',
        'Serikali ya Mtaa', 'Idara ya Afya', 'Daktari', 'EMP001',
        '2010-01-01', '14', '2040-01-01',
        'Dar es Salaam', 'Ilala', 'Kariakoo',
        'CHK001', 'Dar es Salaam', 'Ilala', 'Tarafa A', 'Kata B', 'Mtaa C',
        '3',
        'CRDB', 'Kariakoo', '0123456789012', 'Juma Ali Hassan', 'savings',
        'Fatuma Said Musa', 'CHK002',
        'EMP002', 'Mwalimu', 'Shule ya Msingi Kariakoo', 'Kata D', 'Tarafa B',
        'Ilala', 'Dar es Salaam', '0798765432',
    ]
    ws.append(sample)

    # Style headers bold
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1a4b8c')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_wateja.xlsx"'
    return response


def client_export_excel(request):
    """Export all clients of the currently selected branch to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    selected_office = get_selected_office(request)

    # Filter by selected branch; if HQ / None, export all
    if selected_office:
        clients_qs = Client.objects.filter(registered_office=selected_office).order_by('id')
        sheet_title = selected_office.name[:31]   # Excel limit
        filename = f"wateja_{selected_office.name.replace(' ', '_')}.xlsx"
    else:
        clients_qs = Client.objects.all().order_by('id')
        sheet_title = "Wateja Wote"
        filename = "wateja_wote.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [
        'firstname', 'middlename', 'lastname', 'phonenumber',
        'date_of_birth', 'marital_status',
        'employername', 'idara', 'kaziyako', 'employmentcardno',
        'tareheya_kuajiriwa', 'umri_kazini', 'tarehe_ya_kustaafu',
        'region', 'district', 'street',
        'checkno', 'mkoa', 'wilaya', 'tarafa', 'kata', 'mtaa',
        'wategemezi_wako',
        'bank_name', 'bank_branch', 'bank_account_number',
        'account_name', 'account_type',
        'mdhamini_jina_kamili', 'mdhamini_checkno',
        'mdhamini_kitambulisho_kazi', 'mdhamini_kazi',
        'mdhamini_kituo_kazi', 'mdhamini_kata', 'mdhamini_tarafa',
        'mdhamini_wilaya', 'mdhamini_mkoa', 'mdhamini_simu',
        'registered_date',
    ]

    # ── Styles ──────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(fill_type='solid', fgColor='1a4b8c')
    center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left         = Alignment(horizontal='left', vertical='center')
    alt_fill     = PatternFill(fill_type='solid', fgColor='f0f5ff')
    thin         = Border(
        left=Side(style='thin', color='d0dbe8'),
        right=Side(style='thin', color='d0dbe8'),
        top=Side(style='thin', color='d0dbe8'),
        bottom=Side(style='thin', color='d0dbe8'),
    )

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin
    ws.row_dimensions[1].height = 28

    def safe(val):
        if val is None:
            return ''
        return val

    for i, client in enumerate(clients_qs, start=2):
        row = [
            safe(client.firstname),
            safe(client.middlename),
            safe(client.lastname),
            safe(client.phonenumber),
            client.date_of_birth.strftime('%Y-%m-%d') if client.date_of_birth else '',
            safe(client.marital_status),
            safe(client.employername),
            safe(client.idara),
            safe(client.kaziyako),
            safe(client.employmentcardno),
            client.tareheya_kuajiriwa.strftime('%Y-%m-%d') if client.tareheya_kuajiriwa else '',
            safe(client.umri_kazini),
            client.tarehe_ya_kustaafu.strftime('%Y-%m-%d') if client.tarehe_ya_kustaafu else '',
            safe(client.region),
            safe(client.district),
            safe(client.street),
            safe(client.checkno),
            safe(client.mkoa),
            safe(client.wilaya),
            safe(client.tarafa),
            safe(client.kata),
            safe(client.mtaa),
            safe(client.wategemezi_wako),
            safe(client.bank_name),
            safe(client.bank_branch),
            safe(client.bank_account_number),
            safe(client.account_name),
            safe(client.account_type),
            safe(client.mdhamini_jina_kamili),
            safe(client.mdhamini_checkno),
            safe(client.mdhamini_kitambulisho_kazi),
            safe(client.mdhamini_kazi),
            safe(client.mdhamini_kituo_kazi),
            safe(client.mdhamini_kata),
            safe(client.mdhamini_tarafa),
            safe(client.mdhamini_wilaya),
            safe(client.mdhamini_mkoa),
            safe(client.mdhamini_simu),
            client.registered_date.strftime('%Y-%m-%d') if client.registered_date else '',
        ]
        ws.append(row)
        fill = alt_fill if i % 2 == 0 else None
        for cell in ws[i]:
            cell.alignment = left
            cell.border    = thin
            if fill:
                cell.fill = fill
        ws.row_dimensions[i].height = 18

    # Column widths
    col_widths = [16,16,16,16,22,18,24,20,20,22,22,14,22,20,18,18,14,20,18,14,14,14,18,16,18,22,22,22,24,16,26,20,26,14,14,16,16,18,18]
    for col_num, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
